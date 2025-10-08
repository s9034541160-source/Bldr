import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import Outline
from collections import defaultdict
from openpyxl.workbook.defined_name import DefinedName

OUTPUT_FILE = os.path.join(os.getcwd(), 'Бюджет_исправленный_финал.xlsx')

# --- Полная иерархия статей (по скриншоту + современные группы) ---
ARTICLES_TREE = [
    {"name": "ДОХОДЫ", "children": [
        {"name": "Операционные доходы", "children": [
            {"code": "900000", "name": "Выручка всего", "children": [
                {"code": "900100", "name": "Выручка по СМР (ПБУ-2/2008)"},
                {"code": "900105", "name": "Корректировка выручки по ПБУ-2/2008"},
                {"code": "900102", "name": "Принято работ Заказчиком (КС-3)"}
            ]},
            {"code": "900110", "name": "Материалы и оборудование - всего", "children": [
                {"code": "900111", "name": "Материалы поставки Заказчика"},
                {"code": "900117", "name": "Компенсация материалов поставки Заказчика"},
                {"code": "900112", "name": "Материалы поставки Генподрядчика (централизованная поставка)"},
                {"code": "900118", "name": "Компенсация материалов поставки Генподрядчика (централизованная поставка)"},
                {"code": "900114", "name": "Материалы поставки Генподрядчика (прочие прямые поставки)"},
                {"code": "900115", "name": "Собственные материалы"},
                {"code": "900113", "name": "Оборудование"}
            ]},
            {"code": "900120", "name": "Подрядные работы", "children": [
                {"code": "900121", "name": "Собственными силами"},
                {"code": "900122", "name": "Субподряд"}
            ]},
            {"code": "900130", "name": "Страхование"},
            {"code": "900200", "name": "Услуги генподряда"},
            {"code": "900400", "name": "Выручка от реализации строительных материалов прочим организациям"},
            {"code": "900900", "name": "Прочая выручка"},
            {"code": "900901", "name": "Аренда земельных участков, зданий, помещений"},
            {"code": "900904", "name": "Аренда машин и механизмов"},
            {"code": "900902", "name": "Автоуслуги"},
            {"code": "900903", "name": "Прочее"}
        ]}
    ]},
    {"name": "РАСХОДЫ", "children": [
        {"name": "Операционные расходы", "children": [
            {"name": "Прямые расходы", "children": [
                {"name": "Материалы", "children": [
                    {"code": "200110", "name": "Материалы поставки Заказчика"},
                    {"code": "200120", "name": "Материалы поставки Генподрядчика (централизованная поставка)"},
                    {"code": "200130", "name": "Материалы поставки Генподрядчика (прочие прямые поставки)"},
                    {"code": "200140", "name": "Собственные материалы"},
                    {"code": "200150", "name": "Непроектные материалы + расходные материалы"},
                    {"code": "200160", "name": "Тара, упаковка"},
                    {"code": "200200", "name": "Оборудование"}
                ]},
                {"name": "ФОТ", "children": [
                    {"code": "200701", "name": "ФОТ рабочих"},
                    {"code": "200702", "name": "ФОТ ИТР"},
                    {"code": "200703", "name": "Выплаты социального характера"},
                    {"code": "200704", "name": "Страховые взносы в фонды и страхование от несчастных случаев на пр-ве"},
                    {"code": "200709", "name": "Обучение персонала, участие в семинарах"},
                    {"code": "200710", "name": "Аттестация персонала, повышение квалификации"},
                    {"code": "200712", "name": "Резерв на оплату отпусков"},
                    {"code": "200711", "name": "Прочие расходы на персонал"}
                ]},
                {"name": "Субподряд", "children": [
                    {"code": "200300", "name": "Услуги по генеральному подряду"},
                    {"code": "200400", "name": "Расходы на СМР, выполненные субподрядными орг-ми"},
                    {"code": "200450", "name": "Расходы на субподрядные работы"},
                    {"code": "200460", "name": "Генподрядное вознаграждение"},
                    {"code": "200500", "name": "Расходы на проектно-изыскат-кие работы, выполненные субподряд. орг-ми"}
                ]},
                {"name": "Машины и механизмы", "children": [
                    {"code": "201101", "name": "ГСМ"},
                    {"code": "201102", "name": "Запчасти"},
                    {"code": "201103", "name": "Текущий ремонт и техническое обслуживание (сторонее)"},
                    {"code": "201110", "name": "Текущий ремонт и техническое обслуживание (собств. силами)"},
                    {"code": "201106", "name": "ОСАГО"},
                    {"code": "201107", "name": "КАСКО"},
                    {"code": "201109", "name": "Услуги управления автотранспорта"},
                    {"code": "201108", "name": "Прочие расходы по обслуживанию и эксплуатации машин и механизмов"}
                ]},
                {"name": "Прочие прямые расходы", "children": [
                    {"code": "201701", "name": "ТМЦ стоимостью в пределах лимита"},
                    {"code": "201703", "name": "Прочие транспортные услуги"},
                    {"code": "201704", "name": "Услуги по предоставлению персонала"},
                    {"code": "201702", "name": "Прочие услуги производственного характера"}
                ]}
            ]},
            {"name": "Косвенные расходы", "children": [
                {"code": "250000", "name": "Общепроизводственные расходы"},
                {"code": "260000", "name": "Управленческие расходы и налоги"},
                {"code": "680000", "name": "Управленческие расходы и налоги"},
                {"name": "Амортизация", "children": [
                    {"code": "201600", "name": "Амортизация и списание объектов ОС"},
                    {"code": "201601", "name": "Амортизация МиМ и оборудования"},
                    {"code": "201604", "name": "Амортизация ВЗиС"},
                    {"code": "201605", "name": "Амортизация прочее"},
                    {"code": "201602", "name": "Списание объектов стоимость в пределах лимита"},
                    {"code": "201603", "name": "Списание стоимости неисключительных прав"}
                ]},
                {"name": "Командировочные расходы", "children": [
                    {"code": "201500", "name": "Командировочные расходы"},
                    {"code": "201501", "name": "Билеты"},
                    {"code": "201502", "name": "Проживание в гостинице"},
                    {"code": "201506", "name": "Услуги гостиничного комплекса"},
                    {"code": "201503", "name": "Суточные"},
                    {"code": "201505", "name": "Прочие командировочные расходы"}
                ]},
                {"name": "Охрана труда", "children": [
                    {"code": "200800", "name": "Затраты по охране труда"},
                    {"code": "200801", "name": "Спецодежда"},
                    {"code": "200802", "name": "Спецмолоко"},
                    {"code": "200803", "name": "Санитарно-гигиенические нормы"},
                    {"code": "200804", "name": "Медосмотр"},
                    {"code": "200805", "name": "Средства для охраны труда"},
                    {"code": "200807", "name": "Аттестация рабочих мест"},
                    {"code": "200806", "name": "Прочие по охране труда"}
                ]},
                {"name": "Связь, охрана, склад", "children": [
                    {"code": "202100", "name": "Услуги связи"},
                    {"name": "Охрана и безопасность", "children": [
                        {"code": "202200", "name": "Охрана и безопасность"},
                        {"code": "202201", "name": "Охранные услуги"},
                        {"code": "202202", "name": "Обслуживание охранных систем и сигнализации"}
                    ]},
                    {"code": "202500", "name": "Услуги собственного складского хозяйства"}
                ]},
                {"name": "Коммунальные", "children": [
                    {"code": "201400", "name": "Коммунальные производственные расходы"},
                    {"code": "201401", "name": "Электроэнергия"},
                    {"code": "201402", "name": "Вода и стоки (канализация)"},
                    {"code": "201403", "name": "Теплоэнергия"},
                    {"code": "201404", "name": "Вывоз мусора"},
                    {"code": "201405", "name": "Прочие коммунальные платежи"}
                ]},
                {"name": "Аренда", "children": [
                    {"code": "201200", "name": "Аренда имущества производственного назначения"},
                    {"code": "201201", "name": "Земля"},
                    {"code": "201202", "name": "Здания, помещения"},
                    {"code": "201203", "name": "Складские помещения и складское оборудование"},
                    {"code": "201205", "name": "Автотранспорт, специальные машины, механизмы и оборудование"},
                    {"code": "201208", "name": "Финансовая аренда (лизинг)"}
                ]},
                {"name": "Резервы", "children": [
                    {"code": "201800", "name": "Резервы"},
                    {"code": "201801", "name": "Резерв на гарантийное обслуживание и гарантийный ремонт"},
                    {"code": "201802", "name": "Резерв на покрытие предвиденных расходов"},
                    {"code": "201804", "name": "Резерв на компенсацию стоимости МТР"},
                    {"code": "201803", "name": "Резерв по ожидаемому убытку"}
                ]}
            ]},
            {"name": "Операционные расходы (прочие)", "children": [
                {"code": "800001", "name": "Операционные расходы - всего"},
                {"code": "800002", "name": "Операционная прибыль"}
            ]},
            # --- Современные группы ---
            {"name": "IT и цифровизация", "children": [
                {"code": "710100", "name": "BIM и цифровое проектирование"},
                {"code": "710200", "name": "Программное обеспечение"},
                {"code": "710300", "name": "Цифровые сервисы и облачные решения"}
            ]},
            {"name": "Экологические расходы", "children": [
                {"code": "720100", "name": "Утилизация отходов"},
                {"code": "720200", "name": "Экологический аудит и мониторинг"},
                {"code": "720300", "name": "Природоохранные мероприятия"}
            ]},
            {"name": "PR/GR/Маркетинг", "children": [
                {"code": "730100", "name": "PR и продвижение"},
                {"code": "730200", "name": "Взаимодействие с госорганами (GR)"},
                {"code": "730300", "name": "Маркетинговые исследования"}
            ]},
            {"name": "Обучение и развитие персонала", "children": [
                {"code": "740100", "name": "Корпоративное обучение"},
                {"code": "740200", "name": "Сертификация и повышение квалификации"},
                {"code": "740300", "name": "Тренинги по безопасности"}
            ]},
            {"name": "Безопасность и Compliance", "children": [
                {"code": "750100", "name": "Внутренний аудит"},
                {"code": "750200", "name": "Внутренний контроль и комплаенс"},
                {"code": "750300", "name": "Расходы на охрану"}
            ]},
            {"name": "Логистика", "children": [
                {"code": "760100", "name": "Транспортировка материалов"},
                {"code": "760200", "name": "Складские услуги"},
                {"code": "760300", "name": "Внешняя логистика"}
            ]},
            {"name": "Сервисное обслуживание после сдачи объекта", "children": [
                {"code": "770100", "name": "Гарантийные обязательства"},
                {"code": "770200", "name": "Постгарантийный сервис"},
                {"code": "770300", "name": "Рекламации и устранение дефектов"}
            ]},
            {"name": "Риски и непредвиденные расходы", "children": [
                {"code": "700100", "name": "Резерв на инфляцию"},
                {"code": "700200", "name": "Колебания цен на материалы и услуги"},
                {"code": "700300", "name": "Форс-мажор и непредвиденные расходы"}
            ]}
        ]}
    ]}
]

# Современные светлые стили для уровней
LEVEL_STYLES = [
    {  # 1 уровень (ДОХОДЫ/РАСХОДЫ)
        'fill': PatternFill(start_color='E3EAFD', end_color='E3EAFD', fill_type='solid'),
        'font': Font(bold=True, color='305496')
    },
    {  # 2 уровень
        'fill': PatternFill(start_color='F2F9F1', end_color='F2F9F1', fill_type='solid'),
        'font': Font(bold=True, color='2E7D32')
    },
    {  # 3 уровень
        'fill': PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid'),
        'font': Font(bold=False, color='305496')
    },
    {  # 4+ уровень
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'font': Font(bold=False, color='666666')
    },
]

BORDER = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

def write_articles_tree(ws, tree, level=1, row_counter=None, num_stack=None):
    if row_counter is None:
        row_counter = [2]  # Счетчик строк начинается с 2 (после заголовка)
    if num_stack is None:
        num_stack = []     # Стек для многоуровневой нумерации
    for idx, node in enumerate(tree, 1):
        if len(num_stack) < level:
            num_stack.append(idx)
        else:
            num_stack[level-1] = idx
            num_stack = num_stack[:level]
        num_str = '.'.join(str(n) for n in num_stack)
        row = row_counter[0]
        # Запись узла
        if 'code' in node:
            ws.append([num_str, node['code'], '    ' * (level-1) + node['name']])
        else:
            ws.append([num_str, '', '    ' * (level-1) + node['name']])
        # Применение стилей
        style_idx = min(level-1, len(LEVEL_STYLES)-1)
        style = LEVEL_STYLES[style_idx]
        for col in range(1, 4):
            ws.cell(row=row, column=col).font = style['font']
            ws.cell(row=row, column=col).fill = style['fill']
            ws.cell(row=row, column=col).border = BORDER
        row_counter[0] += 1
        # Дети
        if 'children' in node and node['children']:
            write_articles_tree(ws, node['children'], level+1, row_counter, num_stack.copy())
    return row_counter[0] - 1

def format_articles_sheet(ws):
    white_bold = Font(bold=True, color='305496')
    header_fill = PatternFill(start_color='E3EAFD', end_color='E3EAFD', fill_type='solid')
    border = BORDER
    center = Alignment(horizontal='center', vertical='center')
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = white_bold
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).alignment = center
        ws.cell(row=1, column=col).border = border
    for col in ws.columns:
        max_length = 0
        col_letter_name = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                # Преобразуем в строку и исключаем формулы
                cell_str = str(cell.value)
                if not cell_str.startswith('='):
                    max_length = max(max_length, len(cell_str))
        # Ограничиваем ширину
        adjusted_width = max(10, min(max_length + 2, 50))
        ws.column_dimensions[col_letter_name].width = adjusted_width
    ws.freeze_panes = ws['A2']
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 80

# --- Исправленные данные для ФОТ АУП (зарплаты и налоги в тысячах) ---
AUP_MIN = [
    ["Генеральный директор", 1, 80000, 112000],
    ["Первый ЗГД", 1, 200000, 280000],
    ["Главный бухгалтер", 1, 200000, 280000],
    ["Бухгалтерия", 2, 150000, 210000],
    ["Списание", 2, 150000, 210000],
    ["Отдел кадров", 2, 150000, 420000],
    ["ОТиТБ", 2, 120000, 168000],
    ["Комплектация", 1, 200000, 280000],
    ["Закупка/логистика", 1, 200000, 280000],
    ["ПТУ", 1, 200000, 280000],
    ["ЗамПТУ", 1, 200000, 280000],
    ["Инженер ПТУ", 2, 200000, 280000],
    ["КСП", 2, 200000, 280000],
    ["АТК", 2, 200000, 280000],
    ["Переватка", 2, 200000, 280000],
    ["ГИ", 2, 200000, 280000],
    ["Технолог", 2, 200000, 280000],
    ["СК", 2, 200000, 280000],
    ["Инженер по сварке", 2, 200000, 280000],
    ["IT", 1, 180000, 252000],
]

AUP_MARKET = [
    ["Генеральный директор", 1, 80000, 112000],
    ["Первый ЗГД", 1, 250000, 350000],
    ["Главный бухгалтер", 1, 250000, 350000],
    ["Бухгалтерия", 2, 180000, 252000],
    ["Списание", 2, 180000, 252000],
    ["Отдел кадров", 2, 180000, 504000],
    ["ОТиТБ", 2, 150000, 210000],
    ["Комплектация", 1, 250000, 350000],
    ["Закупка/логистика", 1, 250000, 350000],
    ["ПТУ", 1, 250000, 350000],
    ["ЗамПТУ", 1, 250000, 350000],
    ["Инженер ПТУ", 2, 250000, 350000],
    ["КСП", 2, 250000, 350000],
    ["АТК", 2, 250000, 350000],
    ["Переватка", 2, 250000, 350000],
    ["ГИ", 2, 250000, 350000],
    ["Технолог", 2, 250000, 350000],
    ["СК", 2, 250000, 350000],
    ["Инженер по сварке", 2, 250000, 350000],
    ["IT", 1, 180000, 252000],
]

AUP_HEADERS = ["№", "Должность", "Этапность", "На руки", "С Налогами"]

def fill_aup_sheet(ws, data):
    # Заголовки
    ws.append(AUP_HEADERS)
    # Данные
    for i, row in enumerate(data, 1):
        ws.append([i] + row)
    # Итоговая строка
    last = ws.max_row
    ws.append(['', 'Итого', '', f'=SUM(D2:D{last})', f'=SUM(E2:E{last})'])
    # Форматирование (наш стиль)
    white_bold = Font(bold=True, color='305496')
    header_fill = PatternFill(start_color='E3EAFD', end_color='E3EAFD', fill_type='solid')
    zebra_fill = PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid')
    border = BORDER
    center = Alignment(horizontal='center', vertical='center')
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = white_bold
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).alignment = center
        ws.cell(row=1, column=col).border = border
    for row in range(2, ws.max_row):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = border
            if row % 2 == 0:
                ws.cell(row=row, column=col).fill = zebra_fill
    # Итог жирный
    for col in range(1, ws.max_column + 1):
        ws.cell(row=ws.max_row, column=col).font = white_bold
        ws.cell(row=ws.max_row, column=col).border = border
    # Денежный формат для всех сумм
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = '#,##0 ₽'
        ws.cell(row=row, column=5).number_format = '#,##0 ₽'
    for col in ws.columns:
        max_length = 0
        col_letter_name = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_str = str(cell.value)
                if not cell_str.startswith('='):
                    max_length = max(max_length, len(cell_str))
        adjusted_width = max(10, min(max_length + 2, 50))
        ws.column_dimensions[col_letter_name].width = adjusted_width
    ws.freeze_panes = ws['A2']
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

FOT_OMP_DATA = [
    ["Руководитель", 1, 250000, 350000],
    ["Зам. Рук", 2, 300000, 420000],
    ["Комплектация (1)", 1, 250000, 350000],
    ["Зав.склада (1)", 1, 250000, 350000],
    ["Кладовщик (1)", 2, 300000, 420000],
    ["Бухгалтер (1)", 1, 180000, 252000],
    ["Списание (1)", 2, 200000, 280000],
    ["Отиз (1)", 1, 180000, 252000],
    ["Отдел кадров (1)", 1, 180000, 252000],
    ["ОТиТБ (1)", 1, 180000, 252000],
    ["Эколог (2)", 2, 300000, 420000],
    ["ПТО (1)", 1, 250000, 350000],
    ["Зам.ПТО (2)", 2, 300000, 420000],
    ["АТК (1)", 1, 250000, 350000],
    ["Инженер ГСМ", 2, 200000, 280000],
    ["Диспетчер", 2, 200000, 280000],
    ["Медик", 2, 200000, 280000],
    ["Переватка (1)", 1, 180000, 252000],
    ["Подготовка (1)", 1, 180000, 252000],
    ["IT (1A)", 1, 180000, 252000],
]

def fill_autobudget_sheet(wb, articles_tree):
    ws_src = wb.create_sheet('Исходники')
    ws = wb.create_sheet('АвтоБюджет')
    # --- Исходные данные ---
    input_rows = [
        ("Стоимость строительства", 100_000_000, "Смета"),
        ("Зарплата рабочего, руб/сут", 6_700, "Договор/Кадры"),
        ("Общая трудоёмкость, чел.-часы", 40_000, "ГПР/Смета"),
        ("Продолжительность, мес", 12, "ГПР/Договор"),
        ("Количество смен", 1, "ГПР/Производство"),
        ("Часов в смене", 10, "ГПР/Производство"),
        ("Смен в месяце", 30, "ГПР/Производство"),
        ("Режим вахты, дней", 45, "Договор/Кадры"),
        ("Коэффициент выработки", 0.8, "Расчёт/ГПР"),
        ("Стоимость комплекта СИЗ, руб", 30_000, "Смета/Договор"),
        ("Стоимость билета, руб", 15_000, "Договор/Кадры"),
        ("Стоимость питания, руб/сут/чел", 800, "Договор/Кадры"),
        ("Стоимость проживания, руб/сут/чел", 1_200, "Договор/Кадры"),
        ("Зарплата ИТР, руб/сут", 8_335, "Договор/Кадры"),
        ("Доля линейного ИТР от рабочих", 0.14, "Расчёт/ГПР"),
        ("Обязательный ИТР (шт.)", 8, "Договор/Кадры"),
    ]
    ws_src.append(["Исходные данные", "Значение", "Источник"])
    input_start = ws_src.max_row + 1
    for row in input_rows:
        ws_src.append(list(row))
    input_end = ws_src.max_row
    # --- Присваиваем имена ячейкам исходных данных ---
    input_names = [
        ("СТОИМОСТЬ_СТРОИТЕЛЬСТВА", 0),
        ("ЗП_РАБОЧЕГО_В_СУТКИ", 1),
        ("ТРУДОЕМКОСТЬ", 2),
        ("ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС", 3),
        ("СМЕН", 4),
        ("ЧАСОВ_В_СМЕНЕ", 5),
        ("СМЕН_В_МЕС", 6),
        ("РЕЖИМ_ВАХТЫ_ДНЕЙ", 7),
        ("КЭФ_ВЫРАБОТКИ", 8),
        ("СТОИМОСТЬ_СИЗ", 9),
        ("СТОИМОСТЬ_БИЛЕТА", 10),
        ("СТОИМОСТЬ_ПИТАНИЯ", 11),
        ("СТОИМОСТЬ_ПРОЖИВАНИЯ", 12),
        ("ЗП_ИТР_В_СУТКИ", 13),
        ("ДОЛЯ_ИТР", 14),
        ("ОБЯЗАТЕЛЬНЫЙ_ИТР", 15),
    ]
    for name, idx in input_names:
        cell = ws_src.cell(row=input_start+idx, column=2)
        ref = f"'{ws_src.title}'!${cell.column_letter}${cell.row}"
        ws.parent.defined_names.add(DefinedName(name, attr_text=ref))
    # --- Промежуточные расчёты ---
    ws_src.append(["Промежуточные расчёты", "Значение", "Формула"])
    calc_base = ws_src.max_row + 1
    calc_names = [
        ("КОЛ_РАБОЧИХ", 0),
        ("КОЛ_ИТР", 1),
        ("ФОТ_РАБОЧИХ", 2),
        ("ФОТ_ИТР", 3),
        ("ФОТ_ВСЕГО", 4),
        ("СТРАХОВЫЕ", 5),
        ("ВАХТ_РАБОЧИХ", 6),
        ("ВАХТ_ИТР", 7),
        ("БИЛЕТЫ", 8),
        ("ПРОЖИВАНИЕ", 9),
        ("ПИТАНИЕ", 10),
        ("СИЗ", 11),
    ]
    calc_rows = [
        ("Количество рабочих", '=ROUNDUP(ТРУДОЕМКОСТЬ/КЭФ_ВЫРАБОТКИ/(ЧАСОВ_В_СМЕНЕ*СМЕН*СМЕН_В_МЕС*ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС),0)', "Общая трудоёмкость / коэф.выработки / (часов в смене * смен * смен в мес * продолжительность)", '0'),
        ("Количество ИТР", '=ОБЯЗАТЕЛЬНЫЙ_ИТР+ROUNDUP(ДОЛЯ_ИТР*КОЛ_РАБОЧИХ,0)', "Обязательный ИТР + % от рабочих", '0'),
        ("ФОТ рабочих", '=КОЛ_РАБОЧИХ*ЗП_РАБОЧЕГО_В_СУТКИ*СМЕН*СМЕН_В_МЕС*ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС', "Кол-во рабочих * з/п за смену * смен * смен в мес * продолжительность", '#,##0 ₽'),
        ("ФОТ ИТР", '=КОЛ_ИТР*ЗП_ИТР_В_СУТКИ*ВАХТ_ИТР*РЕЖИМ_ВАХТЫ_ДНЕЙ', "Кол-во ИТР * з/п ИТР * вахт * дней вахты", '#,##0 ₽'),
        ("ФОТ всего", '=ФОТ_РАБОЧИХ+ФОТ_ИТР', "ФОТ рабочих + ФОТ ИТР", '#,##0 ₽'),
        ("Страховые взносы", '=ФОТ_ВСЕГО*0.32', "ФОТ всего * 32%", '#,##0 ₽'),
        ("Кол-во вахт рабочих", '=ROUNDUP(СМЕН_В_МЕС/РЕЖИМ_ВАХТЫ_ДНЕЙ*ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС,0)', "(смен в мес / режим вахты) * продолжительность", '0'),
        ("Кол-во вахт ИТР", '=ROUNDUP(СМЕН_В_МЕС/РЕЖИМ_ВАХТЫ_ДНЕЙ*ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС,0)', "(смен в мес / режим вахты) * продолжительность", '0'),
        ("Билеты", '=(КОЛ_РАБОЧИХ*2*ВАХТ_РАБОЧИХ+КОЛ_ИТР*2*ВАХТ_ИТР)*СТОИМОСТЬ_БИЛЕТА', "(рабочие*2*вахт + ИТР*2*вахт)*стоимость билета", '#,##0 ₽'),
        ("Проживание", '=(КОЛ_РАБОЧИХ*ВАХТ_РАБОЧИХ*РЕЖИМ_ВАХТЫ_ДНЕЙ+КОЛ_ИТР*ВАХТ_ИТР*РЕЖИМ_ВАХТЫ_ДНЕЙ)*СТОИМОСТЬ_ПРОЖИВАНИЯ', "(рабочие*вахт*дней вахты + ИТР*вахт*дней вахты)*стоимость проживания", '#,##0 ₽'),
        ("Питание", '=(КОЛ_РАБОЧИХ*ВАХТ_РАБОЧИХ*РЕЖИМ_ВАХТЫ_ДНЕЙ+КОЛ_ИТР*ВАХТ_ИТР*РЕЖИМ_ВАХТЫ_ДНЕЙ)*СТОИМОСТЬ_ПИТАНИЯ', "(рабочие*вахт*дней вахты + ИТР*вахт*дней вахты)*стоимость питания", '#,##0 ₽'),
        ("СИЗ", '=(КОЛ_РАБОЧИХ+КОЛ_ИТР)*2*СТОИМОСТЬ_СИЗ*ROUNDUP(ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС/12,0)', "(Кол-во рабочих + Кол-во ИТР) * 2 * стоимость комплекта СИЗ * количество комплектов на человека", '#,##0 ₽'),
    ]
    calc_start = ws_src.max_row + 1
    for name, formula, comment, _ in calc_rows:
        ws_src.append([name, formula, comment])
    # Присваиваем имена ячейкам расчётов
    for name, idx in calc_names:
        cell = ws_src.cell(row=calc_start+idx, column=2)
        ref = f"'{ws_src.title}'!${cell.column_letter}${cell.row}"
        ws.parent.defined_names.add(DefinedName(name, attr_text=ref))
    calc_end = ws_src.max_row
    # Оформление расчётов
    for col in range(1, 4):
        cell = ws_src.cell(row=calc_start-1, column=col)
        cell.font = Font(bold=True, color='305496')
        cell.fill = PatternFill(start_color='E3EAFD', end_color='E3EAFD', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    for i, (name, formula, comment, num_format) in enumerate(calc_rows):
        row = calc_start + i
        for col in range(1, 4):
            cell = ws_src.cell(row=row, column=col)
            cell.border = BORDER
            if col == 2:
                cell.number_format = num_format
            if col == 3:
                cell.number_format = '@'  # Формат текстовый для формул
    # --- Далее тело автобюджета (write_tree) — возвращаю рабочую версию: родитель суммирует только непосредственных детей, строки с кодом всегда получают формулу, суммы по месяцам и итогам корректны. ---
    # --- Таблица статей ---
    ws.append(["Код статьи", "Статья"] + [f"Месяц {i+1}" for i in range(12)] + ["Итого"])
    header_row = ws.max_row
    # --- Формулы для статей ---
    code_to_name = {
        '200701': 'ФОТ_РАБОЧИХ',
        '200702': 'ФОТ_ИТР',
        '200704': 'СТРАХОВЫЕ',
        '201501': 'БИЛЕТЫ',
        '201502': 'ПРОЖИВАНИЕ',
        '201503': 'ПИТАНИЕ',
        '200800': 'СИЗ',
        '900000': 'СТОИМОСТЬ_СТРОИТЕЛЬСТВА',
    }
    from openpyxl.utils import get_column_letter
    def write_tree(tree, level=1, num_stack=None, parent_rows=None):
        if num_stack is None:
            num_stack = []
        if parent_rows is None:
            parent_rows = []
        month_numbers = []
        for col in range(3, 15):
            header = ws.cell(row=header_row, column=col).value
            if header and isinstance(header, str) and header.startswith('Месяц'):
                try:
                    month_num = int(header.split(' ')[-1])
                except Exception:
                    month_num = col-2
            else:
                month_num = col-2
            month_numbers.append(month_num)
        for idx, node in enumerate(tree, 1):
            if len(num_stack) < level:
                num_stack.append(idx)
            else:
                num_stack[level-1] = idx
                num_stack = num_stack[:level]
            num_str = '.'.join(str(n) for n in num_stack)
            this_row = ws.max_row + 1
            child_rows = []
            is_leaf = not node.get('children')
            # Формулы для специальных кодов
            if is_leaf:
                if 'code' in node and node['code'] == '900000':
                    vals = [f'=IF({month_numbers[m]}<=ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС,СТОИМОСТЬ_СТРОИТЕЛЬСТВА/ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС,0)' for m in range(12)] + [f'=SUM(C{this_row}:N{this_row})']
                elif 'code' in node and node['code'] == '900121':
                    vals = [f'=IF({month_numbers[m]}<=ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС,СТОИМОСТЬ_СТРОИТЕЛЬСТВА/ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС,0)' for m in range(12)] + [f'=SUM(C{this_row}:N{this_row})']
                elif 'code' in node and node['code'] in code_to_name:
                    name = code_to_name[node['code']]
                    vals = [f'=IF({month_numbers[m]}<=ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС,{name}/ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС,0)' for m in range(12)] + [f'=SUM(C{this_row}:N{this_row})']
                else:
                    vals = ["" for _ in range(12)] + [""]
            else:
                if 'code' in node and node['code'] == '900000':
                    vals = [f'=СТОИМОСТЬ_СТРОИТЕЛЬСТВА/ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС' for _ in range(12)] + [f'=SUM(C{this_row}:N{this_row})']
                elif 'code' in node and node['code'] == '900121':
                    vals = [f'=СТОИМОСТЬ_СТРОИТЕЛЬСТВА/ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС' for _ in range(12)] + [f'=SUM(C{this_row}:N{this_row})']
                elif 'code' in node and node['code'] in code_to_name:
                    name = code_to_name[node['code']]
                    vals = [f'={name}/ПРОДОЛЖИТЕЛЬНОСТЬ_МЕС' for _ in range(12)] + [f'=SUM(C{this_row}:N{this_row})']
                else:
                    vals = ["" for _ in range(12)] + [""]
            ws.append([node.get('code', ''), node['name']] + vals)
            style_idx = min(level-1, len(LEVEL_STYLES)-1)
            style = LEVEL_STYLES[style_idx]
            for col in range(1, ws.max_column+1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = style['font']
                cell.fill = style['fill']
                cell.border = BORDER
                if col >= 3:
                    cell.number_format = '#,##0 ₽'
            if ws.max_row % 2 == 0:
                for col in range(1, ws.max_column+1):
                    ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid')
            # Рекурсия
            if 'children' in node and node['children']:
                before = ws.max_row
                write_tree(node['children'], level+1, num_stack.copy(), child_rows)
                after = ws.max_row
                # Суммируем только непосредственных детей
                if child_rows:  # Проверяем наличие дочерних элементов
                    for m in range(12):
                        col_letter = get_column_letter(3+m)
                        sum_formula = "+".join([f'{col_letter}{row}' for row in child_rows])
                        ws.cell(row=this_row, column=3+m).value = f'={sum_formula}'
                    # Итоговая колонка
                    ws.cell(row=this_row, column=15).value = f'=SUM(C{this_row}:N{this_row})'
            if parent_rows is not None:
                parent_rows.append(this_row)
    write_tree(articles_tree)
    # После write_tree(articles_tree) — добавляем строку итогов по бюджету
    # Найти строки с ДОХОДЫ и РАСХОДЫ
    income_row = None
    expense_row = None
    for row in range(header_row+1, ws.max_row+1):
        # Проверяем вторую колонку (Название статьи)
        name_cell_value = ws.cell(row=row, column=2).value 
        # Проверяем первую колонку (Код статьи) - должна быть пустой для корневых элементов
        code_cell_value = ws.cell(row=row, column=1).value
        
        if name_cell_value == "ДОХОДЫ" and (code_cell_value is None or code_cell_value == ''):
            income_row = row
        elif name_cell_value == "РАСХОДЫ" and (code_cell_value is None or code_cell_value == ''):
            expense_row = row
    # Добавить строки итогов
    income_total_row = None
    expense_total_row = None
    budget_total_row = None
    
    # Добавление строки "Итоговые доходы"
    if income_row is not None:
        # ИЗМЕНЕНИЕ ЗДЕСЬ: используем найденный income_row
        income_total_row = ws.max_row + 1
        # Генерируем формулы для колонок C по O (15 колонка - это Итого)
        ws.append(["", "Итоговые доходы"] + [f"={get_column_letter(col)}{income_row}" for col in range(3, 16)]) 
        # Форматирование Итоговых доходов
        style = LEVEL_STYLES[0]
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=income_total_row, column=col)
            cell.font = style['font']
            cell.fill = style['fill']
            cell.border = BORDER
            if col >= 3:
                cell.number_format = '#,##0 ₽'

    # Добавление строки "Итоговые расходы"
    if expense_row is not None:
        # ИЗМЕНЕНИЕ ЗДЕСЬ: используем найденный expense_row
        expense_total_row = ws.max_row + 1
        # Генерируем формулы для колонок C по O (15 колонка - это Итого)
        ws.append(["", "Итоговые расходы"] + [f"={get_column_letter(col)}{expense_row}" for col in range(3, 16)]) 
        # Форматирование Итоговых расходов
        style = LEVEL_STYLES[0]
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=expense_total_row, column=col)
            cell.font = style['font']
            cell.fill = style['fill']
            cell.border = BORDER
            if col >= 3:
                cell.number_format = '#,##0 ₽'
    # Итоговый бюджет (доходы - расходы) — теперь "Итого по бюджету" и формула по каждому месяцу
    last_row = ws.max_row
    if income_total_row and expense_total_row:
        month_formulas = [f'={get_column_letter(3+m)}{income_total_row}-{get_column_letter(3+m)}{expense_total_row}' for m in range(12)]
        total_formula = f'=O{income_total_row}-O{expense_total_row}'
    else:
        month_formulas = ['0' for _ in range(12)]
        total_formula = '0'
    ws.append(['', 'Итого по бюджету'] + month_formulas + [total_formula])
    budget_total_row = ws.max_row
    # Оформление итоговых строк
    for row in range(ws.max_row-2, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            cell = ws.cell(row=row, column=col)
            # "Итого по бюджету" — красный жирный, остальные — синий жирный
            cell.font = Font(bold=True, color='C00000') if row == ws.max_row else Font(bold=True, color='305496')
            cell.border = BORDER
            if col >= 3:
                cell.number_format = '#,##0 ₽'
    # Оформление строк ДОХОДЫ и РАСХОДЫ
    for row in range(header_row+1, ws.max_row+1):
        code_cell = ws.cell(row=row, column=1).value
        name_cell = ws.cell(row=row, column=2).value
        if not code_cell and name_cell:
            name_str = str(name_cell).strip().upper()
            if 'ДОХОД' in name_str or 'РАСХОД' in name_str:
                for col in range(1, ws.max_column+1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color='305496')
                    cell.border = BORDER
                    if col >= 3:
                        cell.number_format = '#,##0 ₽'

    # --- Оформление листа Исходники ---
    white_bold = Font(bold=True, color='305496')
    header_fill = PatternFill(start_color='E3EAFD', end_color='E3EAFD', fill_type='solid')
    zebra_fill = PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid')
    border = BORDER
    center = Alignment(horizontal='center', vertical='center')
    # Заголовки
    for col in range(1, ws_src.max_column + 1):
        ws_src.cell(row=1, column=col).font = white_bold
        ws_src.cell(row=1, column=col).fill = header_fill
        ws_src.cell(row=1, column=col).alignment = center
        ws_src.cell(row=1, column=col).border = border
    # Зебра и границы
    for row in range(2, ws_src.max_row + 1):
        for col in range(1, ws_src.max_column + 1):
            cell = ws_src.cell(row=row, column=col)
            cell.border = border
            if row % 2 == 0:
                cell.fill = zebra_fill
    # Форматы
    money_rows = {2,3,11,12,13,14,15}  # индексы (1-based) строк с денежными значениями в input_rows
    percent_rows = {10, 16}  # индексы (1-based) строк с процентным форматом (Коэф. выработки, Доля ИТР)
    for i, row in enumerate(range(input_start, input_end+1), 1):
        cell = ws_src.cell(row=row, column=2)
        if i in money_rows:
            cell.number_format = '#,##0 ₽'
        elif i in percent_rows:
            cell.number_format = '0.00%'
        else:
            cell.number_format = '0'
    # Форматы для расчётов
    for i, (name, formula, comment, num_format) in enumerate(calc_rows):
        row = calc_start + i
        cell = ws_src.cell(row=row, column=2)
        cell.number_format = num_format
    # Автоширина для Исходники
    for col in ws_src.columns:
        max_length = 0
        col_letter_name = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_str = str(cell.value)
                if not cell_str.startswith('='):
                    max_length = max(max_length, len(cell_str))
        adjusted_width = max(10, min(max_length + 2, 50))
        ws_src.column_dimensions[col_letter_name].width = adjusted_width
    ws_src.freeze_panes = ws_src['A2']
    # --- Автоширина для АвтоБюджет ---
    for col in ws.columns:
        max_length = 0
        col_letter_name = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                cell_str = str(cell.value)
                if not cell_str.startswith('='):
                    max_length = max(max_length, len(cell_str))
        # Для месяцев (C:O) ширина 15, для остальных авто
        if 3 <= col[0].column <= 15:
            ws.column_dimensions[col_letter_name].width = 15
        else:
            adjusted_width = max(10, min(max_length + 2, 50))
            ws.column_dimensions[col_letter_name].width = adjusted_width

    # --- Добавляю в Исходники ссылки на итоги из АвтоБюджет ---
    autobudget_sheet = ws.title
    if income_total_row:
        ws_src.append(["", "Итого доходы (АвтоБюджет)", f"='{autobudget_sheet}'!O{income_total_row}"])
    if expense_total_row:
        ws_src.append(["", "Итого расходы (АвтоБюджет)", f"='{autobudget_sheet}'!O{expense_total_row}"])
    if budget_total_row:
        ws_src.append(["", "Итого по бюджету (АвтоБюджет)", f"='{autobudget_sheet}'!O{budget_total_row}"])
    # Оформление итоговых строк в Исходники
    total_rows_added = sum([1 for x in [income_total_row, expense_total_row, budget_total_row] if x is not None])
    for row in range(ws_src.max_row-total_rows_added+1, ws_src.max_row+1):
        for col in range(1, ws_src.max_column+1):
            cell = ws_src.cell(row=row, column=col)
            cell.font = Font(bold=True, color='C00000') if row == ws_src.max_row else Font(bold=True, color='305496')
            cell.border = BORDER
            if col == 3:
                cell.number_format = '#,##0 ₽'
    # Форматы для исходных данных — корректировка
    for i, row in enumerate(range(input_start, input_end+1), 1):
        cell = ws_src.cell(row=row, column=2)
        if i == 1:  # Стоимость строительства
            cell.number_format = '#,##0 ₽'
        elif i == 2:  # Зарплата рабочего, руб/сут
            cell.number_format = '#,##0 ₽'
        elif i == 3:  # Общая трудоёмкость, чел.-часы
            cell.number_format = '0'
        elif i == 4:  # Продолжительность, мес
            cell.number_format = '0'
        elif i == 5:  # Количество смен
            cell.number_format = '0'
        elif i == 6:  # Часов в смене
            cell.number_format = '0'
        elif i == 7:  # Смен в месяце
            cell.number_format = '0'
        elif i == 8:  # Режим вахты, дней
            cell.number_format = '0'
        elif i == 9:  # Коэффициент выработки
            cell.number_format = '0.00%'
        elif i == 10:  # Стоимость комплекта СИЗ, руб
            cell.number_format = '#,##0 ₽'
        elif i == 11:  # Стоимость билета, руб
            cell.number_format = '#,##0 ₽'
        elif i == 12:  # Стоимость питания, руб/сут/чел
            cell.number_format = '#,##0 ₽'
        elif i == 13:  # Стоимость проживания, руб/сут/чел
            cell.number_format = '#,##0 ₽'
        elif i == 14:  # Зарплата ИТР, руб/сут
            cell.number_format = '#,##0 ₽'
        elif i == 15:  # Доля линейного ИТР от рабочих
            cell.number_format = '0.00%'
        elif i == 16:  # Обязательный ИТР (шт.)
            cell.number_format = '0'
        else:
            cell.number_format = '@'

def main():
    try:
        # Проверяем доступность директории для сохранения файла
        output_dir = os.path.dirname(OUTPUT_FILE)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            
        # Проверяем структуру данных
        if not ARTICLES_TREE:
            raise ValueError("Ошибка: Пустая структура статей")
        
        if not AUP_MIN or not AUP_MARKET or not FOT_OMP_DATA:
            raise ValueError("Ошибка: Не все данные по ФОТ загружены")
        
        print(f'Начинаем создание бюджета: {OUTPUT_FILE}')
        
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws_art = wb.create_sheet('Справочник статей')
        ws_art.append(['№', 'Код статьи', 'Статья'])
        write_articles_tree(ws_art, ARTICLES_TREE)
        format_articles_sheet(ws_art)
        # --- ФОТ АУП(1) ---
        aup_1 = [row for row in AUP_MIN if row[1] == 1]
        ws_aup_1 = wb.create_sheet('ФОТ АУП(1)')
        fill_aup_sheet(ws_aup_1, aup_1)
        # --- ФОТ АУП(2) ---
        max_by_role = defaultdict(lambda: [None, 0, 0, 0])
        for row in AUP_MIN:
            role = row[0]
            if row[1] in (1,2):
                if max_by_role[role][2] < row[2]:
                    max_by_role[role] = row
        aup_2 = list(max_by_role.values())
        ws_aup_2 = wb.create_sheet('ФОТ АУП(2)')
        fill_aup_sheet(ws_aup_2, aup_2)
        # --- ФОТ ОМП(1) ---
        omp_1 = [row for row in FOT_OMP_DATA if row[1] == 1]
        ws_omp_1 = wb.create_sheet('ФОТ ОМП(1)')
        fill_aup_sheet(ws_omp_1, [[row[0], row[1], row[2], row[3]] for row in omp_1])
        # --- ФОТ ОМП(2) ---
        # Для этапа 2 — этапность 1 и 2, с увеличением зарплат для этапа 2
        omp_2 = [row for row in FOT_OMP_DATA if row[1] in (1,2)]
        ws_omp_2 = wb.create_sheet('ФОТ ОМП(2)')
        fill_aup_sheet(ws_omp_2, [[row[0], row[1], row[2], row[3]] for row in omp_2])
        # --- АвтоБюджет ---
        fill_autobudget_sheet(wb, ARTICLES_TREE)
        wb.save(OUTPUT_FILE)
        print(f'Готово! Файл создан: {OUTPUT_FILE}')
        print(f'Размер файла: {os.path.getsize(OUTPUT_FILE) / 1024:.1f} KB')
        
    except PermissionError as e:
        print(f'Ошибка доступа: {e}')
        print('Проверьте, что файл не открыт в Excel и у вас есть права на запись')
    except ValueError as e:
        print(f'Ошибка данных: {e}')
    except Exception as e:
        print(f'Неожиданная ошибка: {e}')
        print('Попробуйте запустить скрипт ещё раз')

if __name__ == "__main__":
    main() 