"""
Улучшенный парсер бюджета на основе наработок из I:\нейросетки\стройтэк\budget
Интегрирует лучшие практики из budget_fixed.py
"""

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
from typing import Dict, Any, List, Optional, Union
import pandas as pd
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

class EnhancedBudgetParser:
    """Улучшенный парсер бюджета с поддержкой множественных форматов"""
    
    def __init__(self):
        self.budget_categories = {
            'income': {
                'name': 'ДОХОДЫ',
                'items': [
                    {'code': '900000', 'name': 'Выручка всего', 'type': 'total'},
                    {'code': '900121', 'name': 'Собственными силами', 'type': 'subtotal'},
                ]
            },
            'expenses': {
                'name': 'РАСХОДЫ',
                'items': [
                    {'code': '200701', 'name': 'ФОТ', 'type': 'labor'},
                    {'code': '200704', 'name': 'Страховые взносы', 'type': 'insurance'},
                    {'code': '201501', 'name': 'Билеты', 'type': 'travel'},
                    {'code': '201502', 'name': 'Проживание', 'type': 'travel'},
                    {'code': '201503', 'name': 'Питание', 'type': 'travel'},
                    {'code': '200800', 'name': 'СИЗ', 'type': 'equipment'},
                ]
            }
        }
        
        self.calculation_formulas = {
            'labor': lambda base_cost: base_cost * 0.12,  # 12% от базовой стоимости
            'insurance': lambda base_cost: base_cost * 0.03624,  # 30.2% от ФОТ
            'travel': lambda base_cost: base_cost * 0.00276,  # 0.276% от базовой стоимости
            'equipment': lambda base_cost: base_cost * 0.0024,  # 0.24% от базовой стоимости
        }
    
    def parse_excel_budget(self, file_path: str) -> Dict[str, Any]:
        """Парсинг Excel файла бюджета"""
        try:
            # Читаем Excel файл
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Ищем листы с бюджетными данными
            budget_sheets = []
            for sheet_name in wb.sheetnames:
                if any(keyword in sheet_name.lower() for keyword in ['бюджет', 'budget', 'финанс', 'финанс']):
                    budget_sheets.append(sheet_name)
            
            if not budget_sheets:
                # Берем первый лист
                budget_sheets = [wb.sheetnames[0]]
            
            budget_data = {
                'file_path': file_path,
                'parsed_at': datetime.now().isoformat(),
                'sheets': {},
                'total_income': 0,
                'total_expenses': 0,
                'net_profit': 0
            }
            
            for sheet_name in budget_sheets:
                ws = wb[sheet_name]
                sheet_data = self._parse_budget_sheet(ws, sheet_name)
                budget_data['sheets'][sheet_name] = sheet_data
                
                # Суммируем данные
                budget_data['total_income'] += sheet_data.get('total_income', 0)
                budget_data['total_expenses'] += sheet_data.get('total_expenses', 0)
            
            budget_data['net_profit'] = budget_data['total_income'] - budget_data['total_expenses']
            
            return budget_data
            
        except Exception as e:
            logger.error(f"Ошибка парсинга Excel бюджета: {e}")
            return {
                'file_path': file_path,
                'parsed_at': datetime.now().isoformat(),
                'error': str(e),
                'total_income': 0,
                'total_expenses': 0,
                'net_profit': 0
            }
    
    def _parse_budget_sheet(self, ws, sheet_name: str) -> Dict[str, Any]:
        """Парсинг листа бюджета"""
        sheet_data = {
            'sheet_name': sheet_name,
            'rows': [],
            'total_income': 0,
            'total_expenses': 0,
            'categories': defaultdict(list)
        }
        
        # Читаем данные построчно
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not any(cell for cell in row if cell is not None):
                continue
            
            row_data = self._parse_budget_row(row, row_num)
            if row_data:
                sheet_data['rows'].append(row_data)
                
                # Классифицируем по категориям
                if row_data.get('type') == 'income':
                    sheet_data['total_income'] += row_data.get('total_amount', 0)
                    sheet_data['categories']['income'].append(row_data)
                elif row_data.get('type') == 'expense':
                    sheet_data['total_expenses'] += row_data.get('total_amount', 0)
                    sheet_data['categories']['expense'].append(row_data)
        
        return sheet_data
    
    def _parse_budget_row(self, row: tuple, row_num: int) -> Optional[Dict[str, Any]]:
        """Парсинг строки бюджета"""
        # Фильтруем пустые строки
        if not any(cell for cell in row if cell is not None):
            return None
        
        row_data = {
            'row_number': row_num,
            'code': str(row[0]) if row[0] else '',
            'name': str(row[1]) if row[1] else '',
            'monthly_amounts': [],
            'total_amount': 0
        }
        
        # Извлекаем месячные суммы (колонки 2-13)
        monthly_amounts = []
        for i in range(2, min(14, len(row))):
            amount = row[i] if i < len(row) else 0
            if isinstance(amount, (int, float)) and amount != 0:
                monthly_amounts.append(float(amount))
            else:
                monthly_amounts.append(0.0)
        
        row_data['monthly_amounts'] = monthly_amounts
        row_data['total_amount'] = sum(monthly_amounts)
        
        # Определяем тип строки
        name_lower = row_data['name'].lower()
        if any(keyword in name_lower for keyword in ['выручка', 'доход', 'прибыль', 'revenue', 'income']):
            row_data['type'] = 'income'
        elif any(keyword in name_lower for keyword in ['расход', 'затрат', 'стоимость', 'expense', 'cost']):
            row_data['type'] = 'expense'
        else:
            row_data['type'] = 'other'
        
        # Определяем подтип
        if 'фот' in name_lower or 'зарплат' in name_lower:
            row_data['subtype'] = 'labor'
        elif 'страхов' in name_lower:
            row_data['subtype'] = 'insurance'
        elif any(keyword in name_lower for keyword in ['билет', 'проезд', 'транспорт']):
            row_data['subtype'] = 'travel'
        elif any(keyword in name_lower for keyword in ['сиз', 'оборудование', 'инструмент']):
            row_data['subtype'] = 'equipment'
        else:
            row_data['subtype'] = 'other'
        
        return row_data if row_data['total_amount'] > 0 else None
    
    def create_budget_template(self, output_path: str, base_cost: float = 100_000_000) -> Dict[str, Any]:
        """Создание шаблона бюджета с корректными формулами"""
        try:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Создаем лист с бюджетом
            ws = wb.create_sheet('АвтоБюджет_Улучшенный')
            
            # Заголовки
            headers = ["Код статьи", "Статья"] + [f"Месяц {i+1}" for i in range(12)] + ["Итого"]
            ws.append(headers)
            
            # Стили
            self._apply_budget_styles(ws)
            
            # Добавляем статьи бюджета
            current_row = 2
            parent_rows = {}
            
            for category_name, category_data in self.budget_categories.items():
                # Добавляем заголовок категории
                ws.append([category_data['name']] + ["" for _ in range(13)])
                parent_rows[category_name] = current_row
                current_row += 1
                
                # Добавляем статьи категории
                for item in category_data['items']:
                    ws.append([item['code'], item['name']] + ["" for _ in range(13)])
                    
                    # Добавляем формулы для расчета
                    if item['type'] in self.calculation_formulas:
                        base_amount = self.calculation_formulas[item['type']](base_cost)
                        monthly_amount = base_amount / 12
                        
                        # Формулы по месяцам
                        for month in range(1, 13):
                            col_letter = get_column_letter(2 + month)
                            ws[f"{col_letter}{current_row}"] = monthly_amount
                        
                        # Итоговая формула
                        ws[f"O{current_row}"] = f"=SUM(C{current_row}:N{current_row})"
                    
                    current_row += 1
                
                # Добавляем итоги по категории
                if category_name == 'income':
                    ws.append(["", "ИТОГО ДОХОДЫ"] + ["" for _ in range(13)])
                    ws[f"O{current_row}"] = f"=SUM(O{parent_rows[category_name]+1}:O{current_row-1})"
                elif category_name == 'expenses':
                    ws.append(["", "ИТОГО РАСХОДЫ"] + ["" for _ in range(13)])
                    ws[f"O{current_row}"] = f"=SUM(O{parent_rows[category_name]+1}:O{current_row-1})"
                
                current_row += 1
            
            # Добавляем итоговую строку
            ws.append(["", "ЧИСТАЯ ПРИБЫЛЬ"] + ["" for _ in range(13)])
            income_total_row = parent_rows['income'] + len(self.budget_categories['income']['items']) + 1
            expenses_total_row = parent_rows['expenses'] + len(self.budget_categories['expenses']['items']) + 1
            ws[f"O{current_row}"] = f"=O{income_total_row}-O{expenses_total_row}"
            
            # Сохраняем файл
            wb.save(output_path)
            
            return {
                'status': 'success',
                'output_path': output_path,
                'base_cost': base_cost,
                'created_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Ошибка создания шаблона бюджета: {e}")
            return {
                'status': 'error',
                'error': str(e),
                'output_path': output_path
            }
    
    def _apply_budget_styles(self, ws):
        """Применение стилей к листу бюджета"""
        # Стили
        white_bold = Font(bold=True, color='305496')
        header_fill = PatternFill(start_color='E3EAFD', end_color='E3EAFD', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Форматирование заголовков
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = white_bold
            cell.fill = header_fill
            cell.border = border
    
    def calculate_budget_from_estimate(self, estimate_data: Dict[str, Any]) -> Dict[str, Any]:
        """Расчет бюджета на основе данных сметы"""
        try:
            base_cost = estimate_data.get('total_cost', 0)
            
            if base_cost <= 0:
                return {
                    'status': 'error',
                    'error': 'Некорректная базовая стоимость сметы',
                    'base_cost': base_cost
                }
            
            # Рассчитываем статьи бюджета
            budget_items = {}
            
            # ФОТ (12% от базовой стоимости)
            budget_items['labor'] = {
                'name': 'ФОТ',
                'code': '200701',
                'amount': base_cost * 0.12,
                'percentage': 12.0
            }
            
            # Страховые взносы (30.2% от ФОТ)
            fot_amount = budget_items['labor']['amount']
            budget_items['insurance'] = {
                'name': 'Страховые взносы',
                'code': '200704',
                'amount': fot_amount * 0.302,
                'percentage': 30.2
            }
            
            # Билеты (0.36% от базовой стоимости)
            budget_items['travel_tickets'] = {
                'name': 'Билеты',
                'code': '201501',
                'amount': base_cost * 0.0036,
                'percentage': 0.36
            }
            
            # Проживание (1.44% от базовой стоимости)
            budget_items['travel_accommodation'] = {
                'name': 'Проживание',
                'code': '201502',
                'amount': base_cost * 0.0144,
                'percentage': 1.44
            }
            
            # Питание (0.96% от базовой стоимости)
            budget_items['travel_food'] = {
                'name': 'Питание',
                'code': '201503',
                'amount': base_cost * 0.0096,
                'percentage': 0.96
            }
            
            # СИЗ (0.24% от базовой стоимости)
            budget_items['equipment'] = {
                'name': 'СИЗ',
                'code': '200800',
                'amount': base_cost * 0.0024,
                'percentage': 0.24
            }
            
            # Вычисляем итоги
            total_expenses = sum(item['amount'] for item in budget_items.values())
            net_profit = base_cost - total_expenses
            
            return {
                'status': 'success',
                'base_cost': base_cost,
                'budget_items': budget_items,
                'total_expenses': total_expenses,
                'net_profit': net_profit,
                'profit_margin': (net_profit / base_cost * 100) if base_cost > 0 else 0,
                'calculated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Ошибка расчета бюджета: {e}")
            return {
                'status': 'error',
                'error': str(e),
                'base_cost': estimate_data.get('total_cost', 0)
            }
    
    def export_budget_to_excel(self, budget_data: Dict[str, Any], output_path: str) -> Dict[str, Any]:
        """Экспорт данных бюджета в Excel"""
        try:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Создаем лист с бюджетом
            ws = wb.create_sheet('Бюджет_Расчет')
            
            # Заголовки
            headers = ["Код", "Статья", "Сумма", "Процент от базовой стоимости"]
            ws.append(headers)
            
            # Добавляем данные
            current_row = 2
            for item_code, item_data in budget_data.get('budget_items', {}).items():
                ws.append([
                    item_data['code'],
                    item_data['name'],
                    item_data['amount'],
                    f"{item_data['percentage']}%"
                ])
                current_row += 1
            
            # Добавляем итоги
            ws.append(["", "ИТОГО РАСХОДЫ", budget_data.get('total_expenses', 0), ""])
            ws.append(["", "ЧИСТАЯ ПРИБЫЛЬ", budget_data.get('net_profit', 0), ""])
            ws.append(["", "МАРЖА ПРИБЫЛИ", f"{budget_data.get('profit_margin', 0):.2f}%", ""])
            
            # Применяем стили
            self._apply_budget_styles(ws)
            
            # Сохраняем файл
            wb.save(output_path)
            
            return {
                'status': 'success',
                'output_path': output_path,
                'exported_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Ошибка экспорта бюджета: {e}")
            return {
                'status': 'error',
                'error': str(e),
                'output_path': output_path
            }

# Функции для интеграции с существующей системой
def parse_budget_enhanced(file_path: str, **kwargs) -> Dict[str, Any]:
    """
    Улучшенная функция парсинга бюджета
    
    Args:
        file_path: Путь к файлу бюджета
        **kwargs: Дополнительные параметры
        
    Returns:
        Словарь с данными бюджета
    """
    parser = EnhancedBudgetParser()
    
    if file_path.endswith(('.xlsx', '.xls')):
        return parser.parse_excel_budget(file_path)
    else:
        return {
            'status': 'error',
            'error': 'Неподдерживаемый формат файла',
            'file_path': file_path
        }

def calculate_budget_from_estimate(estimate_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Расчет бюджета на основе данных сметы
    
    Args:
        estimate_data: Данные сметы
        
    Returns:
        Словарь с расчетом бюджета
    """
    parser = EnhancedBudgetParser()
    return parser.calculate_budget_from_estimate(estimate_data)

def create_budget_template(output_path: str, base_cost: float = 100_000_000) -> Dict[str, Any]:
    """
    Создание шаблона бюджета
    
    Args:
        output_path: Путь для сохранения шаблона
        base_cost: Базовая стоимость для расчета
        
    Returns:
        Словарь с результатом создания шаблона
    """
    parser = EnhancedBudgetParser()
    return parser.create_budget_template(output_path, base_cost)
