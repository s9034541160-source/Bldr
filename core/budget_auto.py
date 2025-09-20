"""
Automatic budget generator module with enhanced sophistication
"""

import json
import re
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
from dataclasses import dataclass
from enum import Enum

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Optional imports for Excel processing
HAS_EXCEL_LIBS = False
pd = None
np = None
Workbook = None
Font = None
PatternFill = None
dataframe_to_rows = None

try:
    import pandas as pd
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_EXCEL_LIBS = True
except ImportError:
    HAS_EXCEL_LIBS = False
    logger.warning("Excel libraries not available. Excel export will fallback to JSON.")

class CostType(Enum):
    MATERIALS = "materials"
    LABOR = "labor"
    EQUIPMENT = "equipment"
    OVERHEADS = "overheads"
    PROFIT = "profit"

@dataclass
class PositionCost:
    """Detailed cost breakdown for a position"""
    position_code: str
    description: str
    unit: str
    quantity: float
    base_rate: float
    materials_cost: float
    labor_cost: float
    equipment_cost: float
    total_direct_cost: float
    overheads: float
    profit: float
    total_cost: float

def load_gesn_csv(filename: str) -> Dict[str, Any]:
    """
    Load GESN rates from CSV file
    
    Args:
        filename: Path to CSV file
        
    Returns:
        Dictionary with GESN rates
    """
    try:
        if pd is not None and HAS_EXCEL_LIBS:
            df = pd.read_csv(filename, index_col=0)
            logger.info(f"Successfully loaded GESN rates from {filename} with {len(df)} entries")
            return df.to_dict('index')
        else:
            # Fallback implementation without pandas
            import csv
            rates = {}
            with open(filename, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    code = row.pop('code')
                    # Convert string values to floats
                    for key, value in row.items():
                        try:
                            row[key] = float(value)
                        except:
                            pass
                    rates[code] = row
            logger.info(f"Successfully loaded GESN rates from {filename} with {len(rates)} entries (fallback method)")
            return rates
    except FileNotFoundError:
        logger.error(f"GESN CSV file not found: {filename}")
        # Return empty dict with correct structure
        return {}
    except Exception as e:
        logger.error(f"Error loading GESN CSV: {e}")
        # Return empty dict with correct structure
        return {}

def auto_budget(estimate_data: Dict[str, Any], gesn_rates: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generate automatic budget based on estimate data and GESN rates with enhanced sophistication
    
    Args:
        estimate_data: Estimate data from document processing
        gesn_rates: GESN rates data
        
    Returns:
        Detailed budget with calculations
    """
    logger.info("Starting automatic budget calculation")
    
    # Initialize budget structure
    budget = {
        "project_name": estimate_data.get("project_name", "Не указан"),
        "created_date": datetime.now().isoformat(),
        "total_cost": 0.0,
        "sections": [],
        "materials": [],
        "labor": [],
        "equipment": [],
        "overheads": 0.0,
        "profit": 0.0,
        "regional_coefficients": {},
        "cost_breakdown": {
            "materials": 0.0,
            "labor": 0.0,
            "equipment": 0.0,
            "overheads": 0.0,
            "profit": 0.0
        },
        "risk_analysis": {
            "contingency_reserve": 0.0,
            "risk_adjusted_total": 0.0
        }
    }
    
    # Process estimate positions
    positions = estimate_data.get("positions", [])
    total_direct_costs = 0.0
    materials_total = 0.0
    labor_total = 0.0
    equipment_total = 0.0
    
    logger.info(f"Processing {len(positions)} estimate positions")
    
    for i, position in enumerate(positions):
        position_cost = calculate_position_cost(position, gesn_rates)
        budget["sections"].append(position_cost)
        total_direct_costs += position_cost["total_direct_cost"]
        materials_total += position_cost["materials_total"]
        labor_total += position_cost["labor_total"]
        equipment_total += position_cost["equipment_total"]
        logger.debug(f"Processed position {i+1}/{len(positions)}: {position_cost['position_code']} - Total: {position_cost['total_cost']:.2f}")
    
    # Apply regional coefficients
    regional_coeffs = estimate_data.get("regional_coefficients", {})
    budget["regional_coefficients"] = regional_coeffs
    
    regional_multiplier = 1.0
    for coeff_name, coeff_value in regional_coeffs.items():
        regional_multiplier *= (1 + coeff_value / 100)
        logger.info(f"Applied regional coefficient: {coeff_name} = {coeff_value}%")
    
    # Calculate overheads (typically 10-20% of direct costs)
    overheads_percentage = estimate_data.get("overheads_percentage", 15.0)
    budget["overheads"] = total_direct_costs * (overheads_percentage / 100)
    budget["cost_breakdown"]["overheads"] = budget["overheads"]
    logger.info(f"Calculated overheads: {budget['overheads']:.2f} ({overheads_percentage}%)")
    
    # Calculate profit (typically 5-15% of total with overheads)
    profit_percentage = estimate_data.get("profit_percentage", 10.0)
    total_with_overheads = total_direct_costs + budget["overheads"]
    budget["profit"] = total_with_overheads * (profit_percentage / 100)
    budget["cost_breakdown"]["profit"] = budget["profit"]
    logger.info(f"Calculated profit: {budget['profit']:.2f} ({profit_percentage}%)")
    
    # Detailed cost breakdown
    budget["cost_breakdown"]["materials"] = materials_total
    budget["cost_breakdown"]["labor"] = labor_total
    budget["cost_breakdown"]["equipment"] = equipment_total
    
    # Calculate final total before contingency
    pre_contingency_total = total_with_overheads + budget["profit"]
    
    # Apply regional coefficients to final total
    original_total = pre_contingency_total
    pre_contingency_total *= regional_multiplier
    logger.info(f"Applied regional coefficients to total cost: {original_total:.2f} → {pre_contingency_total:.2f}")
    
    # Add contingency reserve (typically 3-10% for construction projects)
    contingency_percentage = estimate_data.get("contingency_percentage", 5.0)
    contingency_reserve = pre_contingency_total * (contingency_percentage / 100)
    budget["risk_analysis"]["contingency_reserve"] = contingency_reserve
    budget["total_cost"] = pre_contingency_total + contingency_reserve
    budget["risk_analysis"]["risk_adjusted_total"] = budget["total_cost"]
    
    # Calculate ROI for validation (profit/cost * 100)
    if total_direct_costs > 0:
        roi = (budget["profit"] / total_direct_costs) * 100
    else:
        roi = 0.0
        logger.warning("Total direct costs is zero, setting ROI to 0%")
    budget["roi"] = roi
    logger.info(f"Calculated ROI: {roi:.2f}%")
    
    logger.info(f"Budget calculation completed. Total cost: {budget['total_cost']:.2f}, ROI: {roi:.2f}%")
    return budget

def calculate_position_cost(position: Dict[str, Any], gesn_rates: Dict[str, Any]) -> Dict[str, Any]:
    """
    Calculate cost for a single position using GESN rates with enhanced sophistication
    
    Args:
        position: Estimate position data
        gesn_rates: GESN rates data
        
    Returns:
        Position cost breakdown
    """
    position_code = position.get("code", "")
    quantity = position.get("quantity", 1.0)
    unit = position.get("unit", "шт")
    
    # Find matching GESN rate
    rate_data = gesn_rates.get(position_code, {})
    if not rate_data:
        # Try to find similar codes
        for code, data in gesn_rates.items():
            if position_code in code or code in position_code:
                rate_data = data
                logger.debug(f"Found similar GESN code: {position_code} → {code}")
                break
    
    # Base rate calculation
    base_rate = rate_data.get("base_rate", 0.0) if isinstance(rate_data, dict) else 0.0
    materials_cost = rate_data.get("materials_cost", 0.0) if isinstance(rate_data, dict) else 0.0
    labor_cost = rate_data.get("labor_cost", 0.0) if isinstance(rate_data, dict) else 0.0
    equipment_cost = rate_data.get("equipment_cost", 0.0) if isinstance(rate_data, dict) else 0.0
    
    # Calculate totals
    position_total = base_rate * quantity
    materials_total = materials_cost * quantity
    labor_total = labor_cost * quantity
    equipment_total = equipment_cost * quantity
    total_direct_cost = position_total + materials_total + labor_total + equipment_total
    
    # Calculate overheads and profit for this position
    overheads_percentage = 15.0  # Default overhead percentage
    overheads = total_direct_cost * (overheads_percentage / 100)
    
    profit_percentage = 10.0  # Default profit percentage
    profit = (total_direct_cost + overheads) * (profit_percentage / 100)
    
    total_cost = total_direct_cost + overheads + profit
    
    result = {
        "position_code": position_code,
        "description": position.get("description", "Не указано"),
        "unit": unit,
        "quantity": quantity,
        "base_rate": base_rate,
        "materials_cost": materials_cost,
        "labor_cost": labor_cost,
        "equipment_cost": equipment_cost,
        "position_total": position_total,
        "materials_total": materials_total,
        "labor_total": labor_total,
        "equipment_total": equipment_total,
        "total_direct_cost": total_direct_cost,
        "overheads": overheads,
        "profit": profit,
        "total_cost": total_cost
    }
    
    logger.debug(f"Position cost calculation: {position_code} - Quantity: {quantity}, Total: {result['total_cost']:.2f}")
    return result

def extract_financial_data(document_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract financial data from document metadata with real regex processing
    
    Args:
        document_data: Document metadata from stage 8
        
    Returns:
        Structured financial data
    """
    financial_data = {
        "positions": [],
        "total_estimated_cost": 0.0,
        "currency": "RUB",
        "regional_coefficients": {}
    }
    
    # Extract from document metadata
    metadata = document_data.get("meta", {})
    
    # Extract finances from metadata using real regex
    content = str(metadata)
    # Extract money amounts with regex
    money_patterns = [
        r'(\d+(?:\s*\d{3})*(?:\.\d+)?)\s*(?:млн\s*)?(?:руб|рублей)',
        r'(\d+(?:\.\d+)?)\s*млн'
    ]
    
    amounts = []
    for pattern in money_patterns:
        matches = re.findall(pattern, content, re.IGNORECASE)
        for match in matches:
            try:
                amount = float(match.replace(' ', ''))
                # Convert millions to rubles if needed
                if 'млн' in match or 'млн' in content:
                    amount *= 1000000
                amounts.append(amount)
            except ValueError:
                logger.warning(f"Could not parse amount: {match}")
                pass
    
    # Create positions from extracted amounts
    for i, amount in enumerate(amounts[:10]):  # Limit to 10 positions
        financial_data["positions"].append({
            "code": f"FIN-{i+1}",
            "description": f"Финансовая позиция {i+1}",
            "quantity": 1.0,
            "unit": "шт",
            "estimated_cost": amount
        })
    
    # Calculate total
    financial_data["total_estimated_cost"] = sum(amounts)
    logger.info(f"Extracted {len(amounts)} financial amounts, total: {financial_data['total_estimated_cost']:.2f} RUB")
    
    # Extract regional coefficients
    regional_patterns = {
        "ekaterinburg": r'(?:екатеринбург|свердловск)',
        "moscow": r'(?:москва|московск)',
        "novosibirsk": r'(?:новосибирск)'
    }
    
    for region, pattern in regional_patterns.items():
        if re.search(pattern, content, re.IGNORECASE):
            financial_data["regional_coefficients"][region] = 10.0  # Default coefficient
            logger.info(f"Detected regional coefficient for {region}")
    
    return financial_data

def export_budget_to_excel(budget: Dict[str, Any], filename: Optional[str] = None) -> str:
    """
    Export budget to Excel format with real formulas and formatting
    
    Args:
        budget: Budget data
        filename: Output filename (optional)
        
    Returns:
        Path to the generated Excel file
    """
    if filename is None:
        filename = f"budget_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Use real Excel export if libraries are available
    if HAS_EXCEL_LIBS and Workbook is not None:
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Смета"
                
                # Add header
                ws['A1'] = f"Смета: {budget.get('project_name', 'Проект')}"
                if Font is not None:
                    ws['A1'].font = Font(bold=True, size=14)
                
                # Add creation date
                ws['A2'] = f"Дата создания: {budget.get('created_date', datetime.now().isoformat())}"
                
                # Add regional coefficients
                row = 4
                if budget.get("regional_coefficients"):
                    ws[f'A{row}'] = "Региональные коэффициенты:"
                    if Font is not None:
                        ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    for coeff_name, coeff_value in budget["regional_coefficients"].items():
                        ws[f'A{row}'] = coeff_name
                        ws[f'B{row}'] = coeff_value
                        row += 1
                    row += 1
                
                # Create budget data DataFrame
                sections = budget.get("sections", [])
                if sections:
                    # Prepare data for DataFrame
                    data = []
                    for section in sections:
                        data.append({
                            "Код": section.get("position_code", ""),
                            "Наименование": section.get("description", ""),
                            "Ед. изм.": section.get("unit", ""),
                            "Количество": section.get("quantity", 0),
                            "Базовая расценка": section.get("base_rate", 0),
                            "Материалы": section.get("materials_cost", 0),
                            "Работа": section.get("labor_cost", 0),
                            "Оборудование": section.get("equipment_cost", 0),
                            "Итого позиция": section.get("total_direct_cost", 0),
                            "Накладные": section.get("overheads", 0),
                            "Прибыль": section.get("profit", 0),
                            "Итого": section.get("total_cost", 0)
                        })
                    
                    # Create DataFrame
                    if pd is not None:
                        df = pd.DataFrame(data)
                        
                        # Add DataFrame to worksheet
                        if dataframe_to_rows is not None:
                            for r in dataframe_to_rows(df, index=False, header=True):
                                ws.append(r)
                        
                        # Format header row
                        if Font is not None and PatternFill is not None:
                            for col in range(1, len(df.columns) + 1):
                                cell = ws.cell(row=row, column=col)
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        row += len(df) + 1
                        
                        # Add totals row
                        total_row = row
                        ws[f'A{total_row}'] = "ИТОГО:"
                        if Font is not None:
                            ws[f'A{total_row}'].font = Font(bold=True)
                        ws[f'I{total_row}'] = f"=SUM(I{row-len(df)}:I{total_row-1})"
                        ws[f'J{total_row}'] = f"=SUM(J{row-len(df)}:J{total_row-1})"
                        ws[f'K{total_row}'] = f"=SUM(K{row-len(df)}:K{total_row-1})"
                        ws[f'L{total_row}'] = f"=SUM(L{row-len(df)}:L{total_row-1})"
                        
                        # Add detailed cost breakdown
                        total_row += 2
                        ws[f'A{total_row}'] = "Детализация затрат:"
                        if Font is not None:
                            ws[f'A{total_row}'].font = Font(bold=True)
                        
                        cost_breakdown = budget.get("cost_breakdown", {})
                        total_row += 1
                        ws[f'A{total_row}'] = "Материалы:"
                        ws[f'B{total_row}'] = cost_breakdown.get("materials", 0)
                        total_row += 1
                        ws[f'A{total_row}'] = "Работа:"
                        ws[f'B{total_row}'] = cost_breakdown.get("labor", 0)
                        total_row += 1
                        ws[f'A{total_row}'] = "Оборудование:"
                        ws[f'B{total_row}'] = cost_breakdown.get("equipment", 0)
                        total_row += 1
                        ws[f'A{total_row}'] = "Накладные расходы:"
                        ws[f'B{total_row}'] = cost_breakdown.get("overheads", 0)
                        total_row += 1
                        ws[f'A{total_row}'] = "Прибыль:"
                        ws[f'B{total_row}'] = cost_breakdown.get("profit", 0)
                        
                        # Add overheads and profit
                        total_row += 2
                        ws[f'A{total_row}'] = "Накладные расходы (%):"
                        ws[f'B{total_row}'] = f"{budget.get('overheads_percentage', 15)}%"
                        total_row += 1
                        ws[f'A{total_row}'] = "Прибыль (%):"
                        ws[f'B{total_row}'] = f"{budget.get('profit_percentage', 10)}%"
                        
                        total_row += 2
                        ws[f'A{total_row}'] = "ИТОГО С НАКЛАДНЫМИ И ПРИБЫЛЬЮ:"
                        if Font is not None:
                            ws[f'A{total_row}'].font = Font(bold=True)
                        ws[f'L{total_row}'] = f"=L{total_row-7}+L{total_row-2}+L{total_row-1}"
                        if Font is not None:
                            ws[f'L{total_row}'].font = Font(bold=True)
                        
                        # Apply regional coefficients
                        if budget.get("regional_coefficients"):
                            total_row += 1
                            ws[f'A{total_row}'] = "С учетом региональных коэффициентов:"
                            if Font is not None:
                                ws[f'A{total_row}'].font = Font(bold=True)
                            
                            # Calculate regional multiplier
                            regional_multiplier = 1.0
                            for coeff_value in budget["regional_coefficients"].values():
                                regional_multiplier *= (1 + coeff_value / 100)
                            
                            ws[f'L{total_row}'] = f"=L{total_row-1}*{regional_multiplier}"
                            if Font is not None:
                                ws[f'L{total_row}'].font = Font(bold=True)
                        
                        # Add risk analysis
                        total_row += 2
                        ws[f'A{total_row}'] = "Анализ рисков:"
                        if Font is not None:
                            ws[f'A{total_row}'].font = Font(bold=True)
                        
                        risk_analysis = budget.get("risk_analysis", {})
                        total_row += 1
                        ws[f'A{total_row}'] = "Резерв на непредвиденные расходы (%):"
                        ws[f'B{total_row}'] = f"{budget.get('contingency_percentage', 5)}%"
                        total_row += 1
                        ws[f'A{total_row}'] = "Резерв на непредвиденные расходы:"
                        ws[f'B{total_row}'] = risk_analysis.get("contingency_reserve", 0)
                        total_row += 1
                        ws[f'A{total_row}'] = "ИТОГО С УЧЕТОМ РИСКОВ:"
                        if Font is not None:
                            ws[f'A{total_row}'].font = Font(bold=True)
                        ws[f'B{total_row}'] = risk_analysis.get("risk_adjusted_total", 0)
                        if Font is not None:
                            ws[f'B{total_row}'].font = Font(bold=True)
            
            # Save workbook
            wb.save(filename)
            logger.info(f"Budget exported to Excel: {filename}")
            return filename
        except Exception as e:
            logger.error(f"Error exporting budget to Excel: {e}")
            # Fallback to JSON if Excel export fails
            pass
    
    # Fallback to JSON export
    json_filename = filename.replace('.xlsx', '.json')
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(budget, f, ensure_ascii=False, indent=2)
    
    logger.info(f"Budget exported to JSON (fallback): {json_filename}")
    return json_filename

# Load real GESN rates from CSV file
SAMPLE_GESN_RATES = load_gesn_csv('data/gesn_rates.csv')
if not SAMPLE_GESN_RATES:
    # Fallback sample rates
    SAMPLE_GESN_RATES = {
        "ГЭСН 8-1-1": {
            "base_rate": 15000.0,
            "materials_cost": 8000.0,
            "labor_cost": 5000.0,
            "equipment_cost": 2000.0,
            "description": "Устройство бетонной подготовки"
        },
        "ГЭСН 8-1-2": {
            "base_rate": 25000.0,
            "materials_cost": 15000.0,
            "labor_cost": 7000.0,
            "equipment_cost": 3000.0,
            "description": "Устройство монолитных бетонных конструкций"
        },
        "ГЭСН 8-6-1.1": {
            "base_rate": 30000.0,
            "materials_cost": 20000.0,
            "labor_cost": 8000.0,
            "equipment_cost": 2000.0,
            "description": "Устройство сборных железобетонных фундаментов"
        }
    }
    logger.warning("Using fallback sample GESN rates. For production use, provide a real GESN CSV file.")