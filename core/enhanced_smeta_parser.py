"""
Улучшенный парсер смет на основе наработок из I:\нейросетки\стройтэк\прога
Интегрирует лучшие практики из adaptive_estimate_parser.py
"""

import pandas as pd
import re
import uuid
import os
import tempfile
from typing import List, Dict, Any, Optional, Tuple, Set
from pathlib import Path
from enum import Enum
import logging

# Импортируем openpyxl для работы с объединенными ячейками
try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None
    load_workbook = None

logger = logging.getLogger(__name__)

class EstimateFormat(Enum):
    """Типы форматов российских смет"""
    GRAND_SMETA = "grand_smeta"
    SMETA_RU = "smeta_ru" 
    AVK5 = "avk5"
    WIN_SMETA = "win_smeta"
    TURBO_SMETKA = "turbo_smetka"
    GOSSTROY_SMETA = "gosstroy_smeta"
    UNKNOWN = "unknown"

class EnhancedSmetaParser:
    """Улучшенный парсер для всех российских форматов смет"""
    
    def __init__(self):
        self._init_format_signatures()
        self._init_common_mappings()
        self.detected_format = EstimateFormat.UNKNOWN
        
    def _init_format_signatures(self):
        """Инициализация сигнатур различных форматов"""
        self.format_signatures = {
            EstimateFormat.GRAND_SMETA: {
                'keywords': ['гранд-смета', 'всего по позиции', 'итого по расценке', 'смр', 'гс', 'гсн', 'фсн', 'тер', 'фер', 'гэсн'],
                'header_patterns': ['наименование работ', 'обоснование', 'сметная стоимость', 'шифр', 'код ресурса', 'ед.изм.', 'ед. изм', 'кол-во', 'количество', 'цена', 'сумма', 'всего'],
                'cost_patterns': ['базисн', 'текущ', 'индекс', 'сметн', 'цены', 'расценк'],
                'structure_indicators': ['справочно', 'фот', 'нр', 'сп', 'ндс', 'итого']
            },
            EstimateFormat.SMETA_RU: {
                'keywords': ['smeta.ru', 'итого позиции', 'всего работ', 'смета точка ру', 'смета.ру', 'smeta', 'смета'],
                'header_patterns': ['работы и затраты', 'шифр норматива', 'стоимость всего', 'наименование', 'ед. изм.', 'кол-во', 'шифр', 'код', 'объем', 'цена за ед.', 'сумма'],
                'cost_patterns': ['стоимость в ценах', 'общая стоимость', 'текущие цены', 'базисные цены', 'сметные цены', 'цены'],
                'structure_indicators': ['нормативы', 'расценки', 'всего по позиции', 'итого по расценке', 'фот', 'нр', 'сп']
            },
            EstimateFormat.AVK5: {
                'keywords': ['авк', 'система авк', 'дбн', 'дсту', 'авк-5'],
                'header_patterns': ['найменування робіт', 'одиниця виміру', 'кількість', 'ціна', 'сума', 'шифр', 'код'],
                'cost_patterns': ['вартість', 'сума', 'ціна', 'цены'],
                'structure_indicators': ['заробітна плата', 'матеріали', 'нр', 'сп']
            },
            EstimateFormat.WIN_SMETA: {
                'keywords': ['win смета', 'винсмета', 'итого смета', 'winсмета'],
                'header_patterns': ['описание работ', 'ед.изм', 'объем', 'количество', 'цена', 'сумма', 'шифр', 'код'],
                'cost_patterns': ['сумма', 'стоимость руб', 'цена', 'сметн'],
                'structure_indicators': ['материальные ресурсы', 'трудозатраты', 'нр', 'сп']
            }
        }
    
    def _init_common_mappings(self):
        """Общие маппинги для всех форматов"""
        self.unit_mappings = {
            # Объемные
            'м3': 'м³', 'куб.м': 'м³', 'м³': 'м³',
            '100 м3': 'м³', 'м3 грунта': 'м³',
            # Площадные  
            'м2': 'м²', 'кв.м': 'м²', 'м²': 'м²',
            '100 м2': 'м²',
            # Линейные
            'м': 'м', 'п.м': 'м', 'м.п': 'м',
            # Массовые
            'кг': 'кг', 'т': 'т', 'тонн': 'т',
            # Штучные
            'шт': 'шт', 'комплект': 'комплект',
            # Временные
            'ч': 'ч', 'час': 'ч', 'чел.час': 'ч'
        }
        
        self.work_material_codes = {
            'work_prefixes': [
                'ТЕР', 'ФЕР', 'ГЭСН', 'ГЭСНр', 'ГЭСНм', 'ФСЭМ', 'ТСЭМ',
                'ТЕРм', 'ТЕРп', 'ФЕРм', 'ФЕРп', 'ГЭСН-2020', 'ФЕР-2020', 'ТЕР-2020',
                'ФССЦ-TER', 'ТСЭМ-2020', 'ФСЭМ-2020', 'ГЭСНр-2020', 'ГЭСНм-2020',
                'ССН', 'ВСН', 'РД', 'Инструкция', 'Методика', 'ЕНиР', 'ГОСТ', 'СНиП',
                'СП', 'СТО', 'ТУ', 'ОСТ', 'РД-11-02', 'РД-11-05', 'МДС', 'Ведомственные',
                'Минрегион', 'Минстрой', 'Приложение', 'Пункт', 'Часть', 'Раздел'
            ],
            'material_prefixes': [
                'МАТ', 'МАТЕРИАЛ', 'РЕСУРС', 'МАТЕРИАЛЬНЫЕ РЕСУРСЫ',
                'МАТЕРИАЛЫ', 'МАТЕРИАЛЬНЫЕ ЗАТРАТЫ', 'МАТЕРИАЛЬНЫЕ РЕСУРСЫ'
            ]
        }
    
    def detect_format(self, file_path: str) -> EstimateFormat:
        """Определение формата сметы по файлу"""
        try:
            # Читаем первые строки файла для анализа
            if file_path.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file_path, nrows=20)
                text_content = ' '.join(df.astype(str).values.flatten())
            elif file_path.endswith('.csv'):
                df = pd.read_csv(file_path, nrows=20, encoding='utf-8')
                text_content = ' '.join(df.astype(str).values.flatten())
            else:
                with open(file_path, 'r', encoding='utf-8') as f:
                    text_content = f.read()[:5000]  # Первые 5000 символов
            
            # Анализируем сигнатуры
            for format_type, signatures in self.format_signatures.items():
                score = 0
                
                # Проверяем ключевые слова
                for keyword in signatures['keywords']:
                    if keyword.lower() in text_content.lower():
                        score += 2
                
                # Проверяем паттерны заголовков
                for pattern in signatures['header_patterns']:
                    if pattern.lower() in text_content.lower():
                        score += 1
                
                # Проверяем паттерны стоимости
                for pattern in signatures['cost_patterns']:
                    if pattern.lower() in text_content.lower():
                        score += 1
                
                # Если набрали достаточно очков, считаем формат определенным
                if score >= 3:
                    self.detected_format = format_type
                    logger.info(f"Определен формат: {format_type.value} (очков: {score})")
                    return format_type
            
            logger.warning("Не удалось определить формат сметы")
            return EstimateFormat.UNKNOWN
            
        except Exception as e:
            logger.error(f"Ошибка определения формата: {e}")
            return EstimateFormat.UNKNOWN
    
    def _process_merged_cells(self, file_path: str) -> str:
        """🚨 КРИТИЧЕСКИЙ ФИКС: Обработка объединенных ячеек"""
        if not HAS_OPENPYXL:
            logger.warning("openpyxl недоступен, пропускаем обработку объединенных ячеек")
            return file_path
        
        try:
            # 1. Загружаем файл через openpyxl
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            logger.info(f"🔍 Обработка объединенных ячеек в файле: {file_path}")
            
            # 2. Обрабатываем все объединенные диапазоны
            merged_ranges = list(sheet.merged_cells.ranges)
            logger.info(f"📊 Найдено {len(merged_ranges)} объединенных диапазонов")
            
            for merged_range in merged_ranges:
                # Получаем границы объединенного диапазона
                min_row, min_col, max_row, max_col = merged_range.bounds
                
                # Получаем значение из первой ячейки
                top_left_cell = sheet.cell(row=min_row, column=min_col)
                top_left_value = top_left_cell.value
                
                if top_left_value is not None:
                    # Присваиваем значение всем ячейкам в диапазоне
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            if cell.value is None:
                                cell.value = top_left_value
                                logger.debug(f"📝 Заполнена ячейка {cell.coordinate}: {top_left_value}")
            
            # 3. Разъединяем все ячейки
            for merged_range in merged_ranges:
                sheet.unmerge_cells(str(merged_range))
            
            # 4. Сохраняем временный файл
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file_path = temp_file.name
            temp_file.close()
            
            workbook.save(temp_file_path)
            logger.info(f"💾 Создан временный файл без объединенных ячеек: {temp_file_path}")
            
            return temp_file_path
            
        except Exception as e:
            logger.error(f"❌ Ошибка обработки объединенных ячеек: {e}")
            return file_path  # Возвращаем оригинальный файл
    
    def parse_excel_estimate(self, file_path: str) -> Dict[str, Any]:
        """Парсинг Excel сметы с улучшенной логикой"""
        temp_file_path = None
        try:
            # 🚨 КРИТИЧЕСКИЙ ФИКС: Обрабатываем объединенные ячейки
            temp_file_path = self._process_merged_cells(file_path)
            
            # Определяем формат
            format_type = self.detect_format(temp_file_path)
            
            # Читаем Excel файл (теперь без объединенных ячеек!)
            df = pd.read_excel(temp_file_path, sheet_name=None)
            
            # Выбираем основной лист
            main_sheet = None
            if len(df) == 1:
                main_sheet = list(df.values())[0]
            else:
                # Ищем лист с наибольшим количеством данных
                max_rows = 0
                for sheet_name, sheet_df in df.items():
                    if len(sheet_df) > max_rows:
                        max_rows = len(sheet_df)
                        main_sheet = sheet_df
            
            if main_sheet is None:
                raise ValueError("Не удалось найти основной лист с данными")
            
            # Очищаем данные
            main_sheet = main_sheet.dropna(how='all')
            
            # Определяем структуру колонок
            columns_mapping = self._detect_columns_structure(main_sheet, format_type)
            
            # Парсим позиции
            positions = self._parse_positions(main_sheet, columns_mapping, format_type)
            
            # Вычисляем итоги
            total_cost = sum(pos.get('total_cost', 0) for pos in positions)
            
            return {
                'format': format_type.value,
                'positions': positions,
                'total_cost': total_cost,
                'positions_count': len(positions),
                'metadata': {
                    'file_path': file_path,
                    'parsed_at': pd.Timestamp.now().isoformat(),
                    'columns_mapping': columns_mapping,
                    'merged_cells_processed': temp_file_path != file_path
                }
            }
            
        except Exception as e:
            logger.error(f"Ошибка парсинга Excel файла: {e}")
            return {
                'format': 'unknown',
                'positions': [],
                'total_cost': 0,
                'positions_count': 0,
                'error': str(e)
            }
        finally:
            # 🧹 Очищаем временный файл
            if temp_file_path and temp_file_path != file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                    logger.debug(f"🗑️ Удален временный файл: {temp_file_path}")
                except Exception as e:
                    logger.warning(f"⚠️ Не удалось удалить временный файл: {e}")
    
    def _detect_columns_structure(self, df: pd.DataFrame, format_type: EstimateFormat) -> Dict[str, str]:
        """Определение структуры колонок"""
        columns_mapping = {}
        
        # Получаем названия колонок
        columns = df.columns.tolist()
        
        # Паттерны для поиска колонок
        patterns = {
            'code': ['шифр', 'код', 'обоснование', 'норматив', 'ресурс'],
            'description': ['наименование', 'описание', 'работы', 'затраты', 'название'],
            'unit': ['ед.изм', 'ед. изм', 'единица', 'измерения', 'изм'],
            'quantity': ['кол-во', 'количество', 'объем', 'к-во', 'кол'],
            'price': ['цена', 'стоимость', 'руб', 'рублей', 'цена за ед'],
            'total': ['сумма', 'итого', 'всего', 'общая стоимость', 'стоимость всего']
        }
        
        for pattern_name, pattern_list in patterns.items():
            for col in columns:
                col_lower = str(col).lower()
                for pattern in pattern_list:
                    if pattern in col_lower:
                        columns_mapping[pattern_name] = col
                        break
                if pattern_name in columns_mapping:
                    break
        
        return columns_mapping
    
    def _parse_positions(self, df: pd.DataFrame, columns_mapping: Dict[str, str], format_type: EstimateFormat) -> List[Dict[str, Any]]:
        """Парсинг позиций сметы"""
        positions = []
        
        for idx, row in df.iterrows():
            try:
                # Извлекаем данные по маппингу колонок
                position = {
                    'id': str(uuid.uuid4()),
                    'row_number': idx + 1
                }
                
                # Код/шифр
                if 'code' in columns_mapping:
                    position['code'] = str(row.get(columns_mapping['code'], '')).strip()
                
                # Описание
                if 'description' in columns_mapping:
                    position['description'] = str(row.get(columns_mapping['description'], '')).strip()
                
                # Единица измерения
                if 'unit' in columns_mapping:
                    unit = str(row.get(columns_mapping['unit'], '')).strip()
                    position['unit'] = self.unit_mappings.get(unit, unit)
                
                # Количество
                if 'quantity' in columns_mapping:
                    quantity = row.get(columns_mapping['quantity'])
                    if pd.notna(quantity):
                        try:
                            position['quantity'] = float(quantity)
                        except (ValueError, TypeError):
                            position['quantity'] = 0.0
                    else:
                        position['quantity'] = 0.0
                
                # Цена
                if 'price' in columns_mapping:
                    price = row.get(columns_mapping['price'])
                    if pd.notna(price):
                        try:
                            position['price'] = float(price)
                        except (ValueError, TypeError):
                            position['price'] = 0.0
                    else:
                        position['price'] = 0.0
                
                # Общая стоимость
                if 'total' in columns_mapping:
                    total = row.get(columns_mapping['total'])
                    if pd.notna(total):
                        try:
                            position['total_cost'] = float(total)
                        except (ValueError, TypeError):
                            position['total_cost'] = 0.0
                    else:
                        position['total_cost'] = 0.0
                
                # Вычисляем общую стоимость если не указана
                if position.get('total_cost', 0) == 0 and position.get('quantity', 0) > 0 and position.get('price', 0) > 0:
                    position['total_cost'] = position['quantity'] * position['price']
                
                # Определяем тип позиции
                position['type'] = self._determine_position_type(position, format_type)
                
                # Фильтруем пустые позиции
                if (position.get('description', '').strip() and 
                    position.get('description', '').strip() not in ['', 'nan', 'None']) or \
                   (position.get('code', '').strip() and 
                    position.get('code', '').strip() not in ['', 'nan', 'None']):
                    positions.append(position)
                
            except Exception as e:
                logger.warning(f"Ошибка парсинга строки {idx + 1}: {e}")
                continue
        
        return positions
    
    def _determine_position_type(self, position: Dict[str, Any], format_type: EstimateFormat) -> str:
        """Определение типа позиции (работа, материал, оборудование)"""
        code = position.get('code', '').upper()
        description = position.get('description', '').upper()
        
        # Проверяем префиксы работ
        for prefix in self.work_material_codes['work_prefixes']:
            if code.startswith(prefix) or prefix in description:
                return 'work'
        
        # Проверяем префиксы материалов
        for prefix in self.work_material_codes['material_prefixes']:
            if code.startswith(prefix) or prefix in description:
                return 'material'
        
        # Проверяем по ключевым словам в описании
        work_keywords = ['работы', 'устройство', 'монтаж', 'установка', 'сборка', 'демонтаж']
        material_keywords = ['материал', 'бетон', 'цемент', 'арматура', 'кирпич', 'металл']
        
        for keyword in work_keywords:
            if keyword in description:
                return 'work'
        
        for keyword in material_keywords:
            if keyword in description:
                return 'material'
        
        return 'unknown'
    
    def parse_csv_estimate(self, file_path: str) -> Dict[str, Any]:
        """Парсинг CSV сметы"""
        try:
            # Читаем CSV файл
            df = pd.read_csv(file_path, encoding='utf-8')
            
            # Определяем формат
            format_type = self.detect_format(file_path)
            
            # Определяем структуру колонок
            columns_mapping = self._detect_columns_structure(df, format_type)
            
            # Парсим позиции
            positions = self._parse_positions(df, columns_mapping, format_type)
            
            # Вычисляем итоги
            total_cost = sum(pos.get('total_cost', 0) for pos in positions)
            
            return {
                'format': format_type.value,
                'positions': positions,
                'total_cost': total_cost,
                'positions_count': len(positions),
                'metadata': {
                    'file_path': file_path,
                    'parsed_at': pd.Timestamp.now().isoformat(),
                    'columns_mapping': columns_mapping
                }
            }
            
        except Exception as e:
            logger.error(f"Ошибка парсинга CSV файла: {e}")
            return {
                'format': 'unknown',
                'positions': [],
                'total_cost': 0,
                'positions_count': 0,
                'error': str(e)
            }
    
    def parse_text_estimate(self, content: str) -> Dict[str, Any]:
        """Парсинг текстовой сметы"""
        try:
            # Определяем формат по содержимому
            format_type = EstimateFormat.UNKNOWN
            
            # Ищем сигнатуры форматов
            for fmt, signatures in self.format_signatures.items():
                score = 0
                for keyword in signatures['keywords']:
                    if keyword.lower() in content.lower():
                        score += 1
                if score >= 2:
                    format_type = fmt
                    break
            
            # Извлекаем позиции из текста
            positions = self._extract_positions_from_text(content, format_type)
            
            # Вычисляем итоги
            total_cost = sum(pos.get('total_cost', 0) for pos in positions)
            
            return {
                'format': format_type.value,
                'positions': positions,
                'total_cost': total_cost,
                'positions_count': len(positions),
                'metadata': {
                    'parsed_at': pd.Timestamp.now().isoformat(),
                    'content_length': len(content)
                }
            }
            
        except Exception as e:
            logger.error(f"Ошибка парсинга текстовой сметы: {e}")
            return {
                'format': 'unknown',
                'positions': [],
                'total_cost': 0,
                'positions_count': 0,
                'error': str(e)
            }
    
    def _extract_positions_from_text(self, content: str, format_type: EstimateFormat) -> List[Dict[str, Any]]:
        """Извлечение позиций из текста"""
        positions = []
        
        # Разбиваем текст на строки
        lines = content.split('\n')
        
        for line_num, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            
            # Ищем паттерны позиций
            position = self._parse_line_as_position(line, line_num + 1)
            if position:
                positions.append(position)
        
        return positions
    
    def _parse_line_as_position(self, line: str, line_num: int) -> Optional[Dict[str, Any]]:
        """Парсинг строки как позиции сметы"""
        # Паттерны для извлечения данных
        patterns = {
            'gesn_code': r'(?:ГЭСН|ФЕР|ТЕР)\s+(\d+(?:-\d+)*(?:\.\d+)*)',
            'cost': r'(\d+(?:\.\d+)?)\s*(?:руб\.?|рублей)',
            'quantity': r'(\d+(?:\.\d+)?)\s*(?:м[²³]|м2|м3|шт|кг|т|м|ч)',
            'description': r'([А-Яа-я\s]+(?:работ|устройств|монтаж|установк))'
        }
        
        position = {
            'id': str(uuid.uuid4()),
            'row_number': line_num,
            'description': line
        }
        
        # Извлекаем код ГЭСН/ФЕР
        gesn_match = re.search(patterns['gesn_code'], line, re.IGNORECASE)
        if gesn_match:
            position['code'] = gesn_match.group(0)
        
        # Извлекаем стоимость
        cost_matches = re.findall(patterns['cost'], line)
        if cost_matches:
            try:
                position['total_cost'] = float(cost_matches[-1])  # Берем последнюю стоимость
            except (ValueError, TypeError):
                position['total_cost'] = 0.0
        
        # Извлекаем количество
        quantity_match = re.search(patterns['quantity'], line)
        if quantity_match:
            try:
                position['quantity'] = float(quantity_match.group(1))
            except (ValueError, TypeError):
                position['quantity'] = 0.0
        
        # Определяем тип позиции
        position['type'] = self._determine_position_type(position, EstimateFormat.UNKNOWN)
        
        return position if position.get('code') or position.get('total_cost', 0) > 0 else None

# Функция для интеграции с существующей системой
def parse_estimate_enhanced(file_path: str, **kwargs) -> Dict[str, Any]:
    """
    Улучшенная функция парсинга сметы
    
    Args:
        file_path: Путь к файлу сметы
        **kwargs: Дополнительные параметры
        
    Returns:
        Словарь с данными сметы
    """
    parser = EnhancedSmetaParser()
    
    if file_path.endswith(('.xlsx', '.xls')):
        return parser.parse_excel_estimate(file_path)
    elif file_path.endswith('.csv'):
        return parser.parse_csv_estimate(file_path)
    else:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return parser.parse_text_estimate(content)
