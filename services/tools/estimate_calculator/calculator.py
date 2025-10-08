"""
Core-логика сметного калькулятора
"""
import pandas as pd
import json
import os
from typing import Dict, List, Tuple
from pathlib import Path

from .models import (
    EstimateRequest, EstimateResponse, CostBreakdown, 
    GESNPrice, SIZNorm, RegionCoeff, VolumeItem
)


class EstimateCalculator:
    """Сметный калькулятор с расчетом маржи %"""
    
    def __init__(self, data_dir: str = None):
        """Инициализация калькулятора"""
        if data_dir is None:
            data_dir = Path(__file__).parent / "data"
        else:
            data_dir = Path(data_dir)
            
        self.data_dir = data_dir
        self.gesn_prices: Dict[str, GESNPrice] = {}
        self.siz_norms: Dict[str, List[SIZNorm]] = {}
        self.region_coeffs: Dict[str, RegionCoeff] = {}
        
        # Загружаем справочники
        self._load_gesn_prices()
        self._load_siz_norms()
        self._load_region_coeffs()
    
    def _load_gesn_prices(self):
        """Загрузка цен по ГЭСН из Excel"""
        try:
            gesn_file = self.data_dir / "gesn_2024.xlsx"
            if gesn_file.exists():
                df = pd.read_excel(gesn_file)
                for _, row in df.iterrows():
                    price = GESNPrice(
                        code=str(row.get('code', '')),
                        name=str(row.get('name', '')),
                        unit=str(row.get('unit', '')),
                        price_material=float(row.get('price_material', 0)),
                        price_labour=float(row.get('price_labour', 0)),
                        price_machine=float(row.get('price_machine', 0)),
                        total_price=float(row.get('price_material', 0)) + 
                                   float(row.get('price_labour', 0)) + 
                                   float(row.get('price_machine', 0))
                    )
                    self.gesn_prices[price.code] = price
            else:
                # Создаем тестовые данные если файл не найден
                self._create_test_gesn_data()
        except Exception as e:
            print(f"Ошибка загрузки ГЭСН: {e}")
            self._create_test_gesn_data()
    
    def _load_siz_norms(self):
        """Загрузка норм СИЗ из Excel"""
        try:
            siz_file = self.data_dir / "siz_2024.xlsx"
            if siz_file.exists():
                df = pd.read_excel(siz_file)
                for _, row in df.iterrows():
                    profession = str(row.get('profession', ''))
                    if profession not in self.siz_norms:
                        self.siz_norms[profession] = []
                    
                    norm = SIZNorm(
                        profession=profession,
                        item=str(row.get('item', '')),
                        unit=str(row.get('unit', '')),
                        norm=float(row.get('norm', 0)),
                        price=float(row.get('price', 0))
                    )
                    self.siz_norms[profession].append(norm)
            else:
                # Создаем тестовые данные СИЗ
                self._create_test_siz_data()
        except Exception as e:
            print(f"Ошибка загрузки СИЗ: {e}")
            self._create_test_siz_data()
    
    def _load_region_coeffs(self):
        """Загрузка районных коэффициентов из JSON"""
        try:
            coeff_file = self.data_dir / "region_coeff.json"
            if coeff_file.exists():
                with open(coeff_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    for item in data:
                        coeff = RegionCoeff(**item)
                        self.region_coeffs[coeff.region] = coeff
            else:
                # Создаем тестовые данные
                self._create_test_region_data()
        except Exception as e:
            print(f"Ошибка загрузки районных коэффициентов: {e}")
            self._create_test_region_data()
    
    def _create_test_gesn_data(self):
        """Создание тестовых данных ГЭСН"""
        test_data = [
            GESNPrice(code="01-01-001-01", name="Бетон тяжелый", unit="м³", 
                     price_material=4500, price_labour=1200, price_machine=800),
            GESNPrice(code="01-01-002-01", name="Арматура", unit="т", 
                     price_material=65000, price_labour=1500, price_machine=200),
            GESNPrice(code="01-01-003-01", name="Опалубка", unit="м²", 
                     price_material=800, price_labour=300, price_machine=100),
        ]
        for price in test_data:
            self.gesn_prices[price.code] = price
    
    def _create_test_siz_data(self):
        """Создание тестовых данных СИЗ"""
        self.siz_norms = {
            "Бетонщик": [
                SIZNorm(profession="Бетонщик", item="Каска", unit="шт", norm=1, price=500),
                SIZNorm(profession="Бетонщик", item="Перчатки", unit="пар", norm=12, price=150),
                SIZNorm(profession="Бетонщик", item="Костюм", unit="шт", norm=2, price=2000),
            ],
            "Арматурщик": [
                SIZNorm(profession="Арматурщик", item="Каска", unit="шт", norm=1, price=500),
                SIZNorm(profession="Арматурщик", item="Перчатки", unit="пар", norm=15, price=200),
                SIZNorm(profession="Арматурщик", item="Костюм", unit="шт", norm=2, price=2000),
            ]
        }
    
    def _create_test_region_data(self):
        """Создание тестовых данных районных коэффициентов"""
        self.region_coeffs = {
            "Москва": RegionCoeff(region="Москва", coeff=1.0, description="Центральный регион"),
            "СПб": RegionCoeff(region="СПб", coeff=1.0, description="Северо-Западный регион"),
            "Мурманск": RegionCoeff(region="Мурманск", coeff=1.4, description="Крайний Север"),
            "Норильск": RegionCoeff(region="Норильск", coeff=1.8, description="Крайний Север"),
        }
    
    def calculate_estimate(self, request: EstimateRequest) -> EstimateResponse:
        """Основной метод расчета сметы"""
        
        # 1. Расчет прямых затрат
        direct_costs = 0.0
        payroll = 0.0
        
        for volume in request.volumes:
            if volume.code in self.gesn_prices:
                gesn = self.gesn_prices[volume.code]
                direct_costs += volume.qty * gesn.total_price
                payroll += volume.qty * gesn.price_labour
            else:
                # Если код не найден, используем средние цены
                direct_costs += volume.qty * 5000  # Средняя цена за единицу
                payroll += volume.qty * 1500       # Средняя зарплата за единицу
        
        # 2. Расчет вахтовой надбавки
        shift_days = int(request.shift_pattern.split('/')[0])
        shift_bonus = payroll * (shift_days / 30) * 0.2
        
        # 3. Расчет северного коэффициента
        region_coeff = self.region_coeffs.get(request.region, RegionCoeff(region=request.region, coeff=1.0, description=""))
        northern_coeff = payroll * (request.north_coeff - 1.0)
        
        # 4. Расчет командировочных
        travel_costs = request.travel_days * request.daily_allowance
        
        # 5. Расчет СИЗ (упрощенный)
        siz_costs = self._calculate_siz_costs(request.volumes)
        
        # 6. Расчет накладных расходов (18% от зарплаты)
        overhead = payroll * 0.18
        
        # 7. Общая стоимость
        total_cost = direct_costs + shift_bonus + northern_coeff + travel_costs + siz_costs + overhead
        
        # 8. Расчет маржи
        if request.contract_price:
            margin_pct = ((request.contract_price - total_cost) / request.contract_price) * 100
        else:
            # Если договорная цена не указана, используем коэффициент 1.2 (20% маржа)
            contract_price = total_cost * 1.2
            margin_pct = ((contract_price - total_cost) / contract_price) * 100
        
        # 9. Создание детализации
        cost_breakdown = CostBreakdown(
            direct_costs=direct_costs,
            payroll=payroll,
            shift_bonus=shift_bonus,
            northern_coeff=northern_coeff,
            travel_costs=travel_costs,
            siz_costs=siz_costs,
            overhead=overhead,
            total_cost=total_cost
        )
        
        # 10. Создание Excel файла
        file_path = self._create_excel_report(request, cost_breakdown, margin_pct)
        
        return EstimateResponse(
            file_path=file_path,
            total_cost=total_cost,
            margin_pct=margin_pct,
            cost_breakdown=cost_breakdown,
            region=request.region,
            shift_pattern=request.shift_pattern
        )
    
    def _calculate_siz_costs(self, volumes: List[VolumeItem]) -> float:
        """Расчет стоимости СИЗ"""
        total_siz_cost = 0.0
        
        # Упрощенный расчет: 2% от прямых затрат
        for volume in volumes:
            if volume.code in self.gesn_prices:
                gesn = self.gesn_prices[volume.code]
                direct_cost = volume.qty * gesn.total_price
                total_siz_cost += direct_cost * 0.02
        
        return total_siz_cost
    
    def _create_excel_report(self, request: EstimateRequest, breakdown: CostBreakdown, margin_pct: float) -> str:
        """Создание Excel отчета"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import uuid
            
            # Создаем новую книгу
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Смета"
            
            # Заголовок
            ws['A1'] = "СМЕТА НА ВЫПОЛНЕНИЕ РАБОТ"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Информация о заказе
            ws['A3'] = f"Регион: {request.region}"
            ws['A4'] = f"График вахты: {request.shift_pattern}"
            ws['A5'] = f"Северный коэффициент: {request.north_coeff}"
            
            # Таблица объемов работ
            row = 7
            ws[f'A{row}'] = "№"
            ws[f'B{row}'] = "Код ГЭСН"
            ws[f'C{row}'] = "Наименование"
            ws[f'D{row}'] = "Ед.изм."
            ws[f'E{row}'] = "Количество"
            ws[f'F{row}'] = "Цена за ед."
            ws[f'G{row}'] = "Стоимость"
            
            # Стилизация заголовков
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                cell = ws[f'{col}{row}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            row += 1
            total_volume_cost = 0
            
            for i, volume in enumerate(request.volumes, 1):
                ws[f'A{row}'] = i
                ws[f'B{row}'] = volume.code
                ws[f'C{row}'] = volume.name
                ws[f'D{row}'] = volume.unit
                ws[f'E{row}'] = volume.qty
                
                if volume.code in self.gesn_prices:
                    price = self.gesn_prices[volume.code].total_price
                else:
                    price = 5000  # Средняя цена
                
                ws[f'F{row}'] = price
                cost = volume.qty * price
                ws[f'G{row}'] = cost
                total_volume_cost += cost
                row += 1
            
            # Итого по объемам
            ws[f'F{row}'] = "Итого:"
            ws[f'G{row}'] = total_volume_cost
            ws[f'F{row}'].font = Font(bold=True)
            ws[f'G{row}'].font = Font(bold=True)
            row += 2
            
            # Детализация затрат
            ws[f'A{row}'] = "ДЕТАЛИЗАЦИЯ ЗАТРАТ"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            row += 2
            
            costs = [
                ("Прямые затраты", breakdown.direct_costs),
                ("Зарплата", breakdown.payroll),
                ("Вахтовая надбавка", breakdown.shift_bonus),
                ("Северный коэффициент", breakdown.northern_coeff),
                ("Командировочные", breakdown.travel_costs),
                ("СИЗ", breakdown.siz_costs),
                ("Накладные расходы", breakdown.overhead),
            ]
            
            for name, cost in costs:
                ws[f'A{row}'] = name
                ws[f'B{row}'] = f"{cost:,.2f} ₽"
                row += 1
            
            # Общая стоимость
            ws[f'A{row}'] = "ОБЩАЯ СТОИМОСТЬ"
            ws[f'B{row}'] = f"{breakdown.total_cost:,.2f} ₽"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            row += 1
            
            # Маржа
            ws[f'A{row}'] = f"МАРЖА: {margin_pct:.1f}%"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Автоподбор ширины колонок
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Сохранение файла
            file_id = str(uuid.uuid4())[:8]
            filename = f"estimate_{file_id}.xlsx"
            file_path = f"/tmp/{filename}"
            
            # Создаем директорию если не существует
            os.makedirs("/tmp", exist_ok=True)
            
            wb.save(file_path)
            return file_path
            
        except Exception as e:
            print(f"Ошибка создания Excel: {e}")
            return f"/tmp/estimate_error_{uuid.uuid4()[:8]}.xlsx"
