#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TELEGRAM BOT WITH AIOGRAM
========================
Modern async Telegram bot using aiogram library
"""

import asyncio
import logging
import os
import sys
import base64
import requests
from dotenv import load_dotenv
from collections import defaultdict, deque
from datetime import datetime, timedelta
from typing import Optional, Tuple
import io

"""
Ensure logger is configured before any usage below (including in optional imports).
"""
# Configure logging EARLY
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Ensure project root on sys.path for 'core' imports
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

# Import compressed conversation history
try:
    from core.agents.conversation_history_compressed import compressed_conversation_history
    CONVERSATION_HISTORY_AVAILABLE = True
    logger.info("[TG] Compressed conversation history loaded")
except ImportError as e:
    logger.warning(f"[TG] Compressed conversation history not available: {e}")
    compressed_conversation_history = None
    CONVERSATION_HISTORY_AVAILABLE = False

# Load environment variables
load_dotenv()

# (already configured above)

# Bot configuration
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', 'YOUR_TELEGRAM_BOT_TOKEN_HERE')
API_BASE = os.getenv('API_BASE', 'http://localhost:8000')
API_USER = os.getenv('API_USER', 'admin')
API_PASSWORD = os.getenv('API_PASSWORD', 'admin')

# Chat management
CHAT_HISTORY = defaultdict(lambda: deque(maxlen=10))  # Last 10 messages per chat
# MESSAGE_QUEUE will be created after aiogram imports (event loop safety)
MESSAGE_QUEUE = None  # type: ignore
PROCESSING_CHATS = set()  # Chats currently being processed

def get_auth_headers():
    """Get JWT token and return headers"""
    headers = {"Content-Type": "application/json"}
    try:
        login_resp = requests.post(f"{API_BASE}/token", 
                                 data={"username": API_USER, "password": API_PASSWORD},
                                 timeout=10)
        if login_resp.status_code == 200:
            token = login_resp.json().get('access_token')
            if token:
                headers['Authorization'] = f'Bearer {token}'
                logger.info(f"[TG] JWT token acquired successfully")
            else:
                logger.error(f"[TG] No access_token in login response")
        else:
            logger.error(f"[TG] Login failed: {login_resp.status_code} - {login_resp.text}")
    except Exception as e:
        logger.error(f"[TG] Error getting JWT token: {e}")
    
    return headers

def post_with_auth(url: str, json_payload: dict, timeout: int = 1800) -> requests.Response:
    """
    Perform POST with JWT. If unauthorized, refresh token once and retry.
    """
    headers = get_auth_headers()
    resp = requests.post(url, json=json_payload, headers=headers, timeout=timeout)
    if resp.status_code == 401:
        logger.warning("[TG] 401 Unauthorized, refreshing JWT and retrying once")
        headers = get_auth_headers()
        resp = requests.post(url, json=json_payload, headers=headers, timeout=timeout)
    return resp

def escape_markdown_v2(text: str) -> str:
    """Escape text for Telegram MarkdownV2."""
    if not isinstance(text, str):
        text = str(text)
    # Escape special characters per MarkdownV2 spec
    to_escape = r"_ * [ ] ( ) ~ ` > # + - = | { } . !"
    for ch in to_escape.split():
        text = text.replace(ch, f"\\{ch}")
    return text

async def safe_reply(message, text: str, prefer_markdown: bool = True):
    """
    Safely send a reply: try MarkdownV2 (escaped), fallback to HTML, then plain text.
    """
    try:
        if prefer_markdown:
            escaped = escape_markdown_v2(text)
            return await message.reply(escaped, parse_mode="MarkdownV2")
        else:
            return await message.reply(text)
    except Exception as e_md:
        logger.warning(f"[TG] Markdown send failed: {e_md}")
        try:
            return await message.reply(text, parse_mode="HTML")
        except Exception as e_html:
            logger.warning(f"[TG] HTML send failed: {e_html}")
            return await message.reply(str(text))

def _split_text_chunks(text: str, max_len: int = 3500) -> list:
    if not text:
        return [""]
    chunks = []
    start = 0
    length = len(text)
    while start < length:
        end = min(start + max_len, length)
        # try to split on newline for nicer formatting
        if end < length:
            nl = text.rfind("\n", start, end)
            if nl != -1 and nl > start + 100:
                end = nl + 1
        chunks.append(text[start:end])
        start = end
    return chunks

async def reply_chunked_plain(message, text: str, max_len: int = 3500):
    for chunk in _split_text_chunks(text, max_len=max_len):
        if chunk.strip():
            await message.reply(chunk)

def detect_document_type(filename: Optional[str], header_bytes: bytes) -> str:
    """Detect document type by extension or simple magic bytes."""
    ext = (filename or '').lower().rsplit('.', 1)[-1] if (filename and '.' in filename) else ''
    if ext in {"txt","csv","md","json","pdf","docx","xlsx","rtf"}:
        return ext
    # simple magic checks
    sig = header_bytes[:8]
    if sig.startswith(b"%PDF"):
        return "pdf"
    if sig[:2] == b"PK":
        # Could be docx/xlsx/zip; no quick disambiguation without reading [Content_Types].xml
        return "zip"
    return ext or "bin"

def try_extract_text(data: bytes, doc_type: str) -> Tuple[str, str]:
    """
    Attempt to extract text from common types. Returns (text, method).
    method is a short label describing extractor used or 'none'.
    """
    text = ""
    method = "none"
    try:
        if doc_type in {"txt","csv","md","json"}:
            text = data.decode('utf-8', errors='ignore')
            method = f"decode_{doc_type}"
        elif doc_type == "pdf":
            try:
                import PyPDF2  # type: ignore
                r = PyPDF2.PdfReader(io.BytesIO(data))  # type: ignore
                pages = []
                for p in r.pages[:20]:  # limit pages
                    try:
                        pages.append(p.extract_text() or "")
                    except Exception:
                        pass
                text = "\n\n".join(pages)
                method = "pypdf2"
            except Exception as e:
                method = f"pdf_error:{e}"
        elif doc_type == "docx":
            try:
                import io as _io
                from docx import Document  # type: ignore
                doc = Document(_io.BytesIO(data))
                text = "\n".join([p.text for p in doc.paragraphs])
                method = "python-docx"
            except Exception as e:
                method = f"docx_error:{e}"
        elif doc_type == "xlsx":
            try:
                import io as _io
                from openpyxl import load_workbook  # type: ignore
                wb = load_workbook(filename=_io.BytesIO(data), read_only=True, data_only=True)
                lines = []
                for ws in wb.worksheets[:3]:  # limit sheets
                    for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
                        vals = [str(c) if c is not None else '' for c in row]
                        if any(vals):
                            lines.append("\t".join(vals))
                text = "\n".join(lines)
                method = "openpyxl"
            except Exception as e:
                method = f"xlsx_error:{e}"
        elif doc_type == "rtf":
            try:
                # naive strip RTF braces/controls
                raw = data.decode('utf-8', errors='ignore')
                import re as _re
                text = _re.sub(r"\\[a-zA-Z]+-?\d* ?|[{}]", "", raw)
                method = "rtf_strip"
            except Exception as e:
                method = f"rtf_error:{e}"
    except Exception as e:
        method = f"extract_error:{e}"
    # sanitize and clamp
    if text:
        text = text.replace('\x00','').strip()
        if len(text) > 10000:
            text = text[:10000]
    return text, method

def add_to_chat_history(chat_id: int, message: str, role: str = "user"):
    """Add message to chat history (both local and compressed)"""
    # Add to local history for immediate access
    CHAT_HISTORY[chat_id].append({
        "role": role,
        "content": message,
        "timestamp": datetime.now()
    })
    
    # Add to compressed persistent history
    if CONVERSATION_HISTORY_AVAILABLE and compressed_conversation_history:
        try:
            compressed_conversation_history.add_message(str(chat_id), {
                "role": role,
                "content": message
            })
            logger.info(f"[TG] Added to compressed history for chat {chat_id}")
        except Exception as e:
            logger.error(f"[TG] Error adding to compressed history: {e}")

def get_chat_history(chat_id: int) -> list:
    """Get chat history for context (from compressed history if available)"""
    if CONVERSATION_HISTORY_AVAILABLE and compressed_conversation_history:
        try:
            # Get formatted history from compressed system
            formatted_history = compressed_conversation_history.get_formatted_history(str(chat_id), max_tokens=1000)
            if formatted_history:
                logger.info(f"[TG] Using compressed history for chat {chat_id}")
                return formatted_history.split('\n')
        except Exception as e:
            logger.error(f"[TG] Error getting compressed history: {e}")
    
    # Fallback to local history
    return list(CHAT_HISTORY[chat_id])

from typing import Any

async def queue_message(message: Any):
    """Add message to processing queue"""
    await MESSAGE_QUEUE.put(message)
    logger.info(f"[TG] Message queued for chat {message.chat.id}")

async def process_message_queue():
    """Process messages from queue"""
    while True:
        try:
            message = await MESSAGE_QUEUE.get()
            chat_id = message.chat.id
            
            # Skip if chat is already being processed
            if chat_id in PROCESSING_CHATS:
                # Re-enqueue and wait briefly to preserve order without dropping messages
                logger.info(f"[TG] Chat {chat_id} busy, requeueing message")
                await asyncio.sleep(0.1)
                await MESSAGE_QUEUE.put(message)
                continue
                
            PROCESSING_CHATS.add(chat_id)
            try:
                await process_message(message)
            finally:
                PROCESSING_CHATS.discard(chat_id)
                
        except Exception as e:
            logger.error(f"[TG] Error in message queue processing: {e}")

# Initialize aiogram
try:
    from aiogram import Bot, Dispatcher, types, F
    from aiogram.filters import Command
    from aiogram.types import Message, CallbackQuery
    from aiogram.fsm.context import FSMContext
    from aiogram.fsm.state import State, StatesGroup
    from aiogram.fsm.storage.memory import MemoryStorage
    
    # Create bot and dispatcher
    bot = Bot(token=TELEGRAM_BOT_TOKEN)
    storage = MemoryStorage()
    dp = Dispatcher(storage=storage)
    # Now we can create the queue safely
    if MESSAGE_QUEUE is None:
        MESSAGE_QUEUE = asyncio.Queue()
    
    logger.info(f"[TG] Aiogram bot initialized with token: {TELEGRAM_BOT_TOKEN[:10]}...")
    
except ImportError as e:
    logger.error(f"[TG] Failed to import aiogram: {e}")
    bot = None
    dp = None

# Debug handler for all messages
@dp.message()
async def debug_handler(message: types.Message):
    """Debug handler to log all incoming messages"""
    try:
        logger.info(f"[TG] ===== DEBUG MESSAGE RECEIVED =====")
        logger.info(f"[TG] Message ID: {message.message_id}")
        logger.info(f"[TG] Chat ID: {message.chat.id}")
        logger.info(f"[TG] User ID: {message.from_user.id if message.from_user else 'None'}")
        logger.info(f"[TG] Username: {message.from_user.username if message.from_user else 'None'}")
        logger.info(f"[TG] Text: '{message.text}'")
        logger.info(f"[TG] Content type: {message.content_type}")
        logger.info(f"[TG] Date: {message.date}")
        logger.info(f"[TG] Full message object: {message}")
        logger.info(f"[TG] ===== END DEBUG =====")
        
        # Add to chat history and queue for processing
        add_to_chat_history(message.chat.id, message.text or "Media message", "user")
        logger.info(f"[TG] Added to chat history, queuing message...")
        await queue_message(message)
        logger.info(f"[TG] Message queued successfully")
        
    except Exception as e:
        logger.error(f"[TG] Error in debug handler: {e}")
        import traceback
        logger.error(f"[TG] Traceback: {traceback.format_exc()}")

async def process_message(message: types.Message):
    """Process different types of messages"""
    try:
        if message.text:
            await handle_text(message)
        elif message.voice:
            await handle_voice(message)
        elif message.photo:
            await handle_photo(message)
        elif message.document:
            await handle_document(message)
        else:
            await message.reply("❌ Неподдерживаемый тип сообщения")
            
    except Exception as e:
        logger.error(f"[TG] Error processing message: {e}")
        await message.reply(f"❌ Ошибка обработки: {e}")

async def handle_text(message: types.Message):
    """Handle text messages"""
    try:
        text = message.text or ''
        logger.info(f"[TG] Processing text: '{text}'")
        
        await message.bot.send_chat_action(message.chat.id, "typing")
        
        # Get chat history for context
        chat_history = get_chat_history(message.chat.id)
        logger.info(f"[TG] Chat history: {len(chat_history)} messages")
        
        # Get history stats if available
        history_stats = {}
        if CONVERSATION_HISTORY_AVAILABLE and compressed_conversation_history:
            try:
                history_stats = compressed_conversation_history.get_history_stats(str(message.chat.id))
                logger.info(f"[TG] History stats: {history_stats}")
            except Exception as e:
                logger.error(f"[TG] Error getting history stats: {e}")
        
        headers = get_auth_headers()
        payload = {
            'message': text,
            'context_search': True,
            'max_context': 3,
            'agent_role': 'coordinator',
            'chat_history': chat_history,  # Add chat history
            'request_context': {
                'channel': 'telegram',
                'chat_id': message.chat.id,
                'user_id': message.from_user.id if message.from_user else None,
                'message_id': message.message_id,
                'queue_position': MESSAGE_QUEUE.qsize(),
                'history_stats': history_stats
            }
        }
        
        logger.info(f"[TG] Sending to API: {API_BASE}/api/ai/chat")
        resp = post_with_auth(f"{API_BASE}/api/ai/chat", json_payload=payload, timeout=1800)
        
        if resp.status_code == 200:
            data = resp.json()
            if isinstance(data, str):
                response = data or 'Нет ответа'
            else:
                response = data.get('response') or 'Нет ответа'
            logger.info(f"[TG] API response: {response[:100]}...")
            
            # Add response to chat history
            add_to_chat_history(message.chat.id, response, "assistant")
            
            # Always chunk plain text to avoid Telegram parse/length issues
            await reply_chunked_plain(message, response)
        else:
            error_msg = resp.text if resp.text else f"HTTP {resp.status_code}"
            logger.error(f"[TG] API error: {error_msg}")
            await safe_reply(message, f"❌ Ошибка API: {error_msg}", prefer_markdown=False)
            
    except Exception as e:
        logger.error(f"[TG] Error in handle_text: {e}")
        await safe_reply(message, f"❌ Ошибка: {e}", prefer_markdown=False)

async def handle_voice(message: types.Message):
    """Handle voice messages"""
    try:
        logger.info(f"[TG] Processing voice message")
        await message.bot.send_chat_action(message.chat.id, "typing")
        await safe_reply(message, "🎤 Получил голосовое сообщение. Транскрибирую и анализирую...", prefer_markdown=False)
        
        # Get voice file bytes (aiogram v3)
        voice_file = await message.bot.get_file(message.voice.file_id)
        vbuf = io.BytesIO()
        await message.bot.download(voice_file, destination=vbuf)
        voice_bytes = vbuf.getvalue()
        voice_b64 = base64.b64encode(voice_bytes).decode('utf-8')
        
        headers = get_auth_headers()
        payload = {
            'message': 'Голосовое сообщение',
            'voice_data': voice_b64,
            'context_search': True,
            'max_context': 3,
            'agent_role': 'coordinator',
            'request_context': {
                'channel': 'telegram',
                'chat_id': message.chat.id,
                'message_id': message.message_id,
                'reply_to_message_id': (message.reply_to_message.message_id if message.reply_to_message else None),
                'user_id': message.from_user.id if message.from_user else None
            }
        }
        resp = post_with_auth(f"{API_BASE}/api/ai/chat", json_payload=payload, timeout=1800)
        if resp.status_code == 200:
            response_data = resp.json()
            if isinstance(response_data, str):
                ai_response = response_data
            else:
                ai_response = response_data.get('response', '')
            
            if "Whisper not available" in ai_response:
                final_response = "🎤 Голосовое сообщение получено, но система распознавания речи не установлена.\n\nДля работы с голосовыми сообщениями установите Whisper:\n```bash\npip install openai-whisper\n```"
            elif "Audio transcription failed" in ai_response:
                final_response = "🎤 Голосовое сообщение получено, но не удалось распознать речь.\n\nВозможные причины:\n• Низкое качество записи\n• Слишком много шума\n• Неподдерживаемый формат файла\n\nПопробуйте отправить текстовое сообщение."
            else:
                final_response = f"🎤 Голосовое сообщение обработано:\n\n{ai_response}"
            
            # Long, complex content is sent in plain text and chunked to avoid MarkdownV2 issues
            await reply_chunked_plain(message, final_response)
        else:
            error_msg = resp.text if resp.text else f"HTTP {resp.status_code}"
            await safe_reply(message, f"❌ Ошибка обработки голосового сообщения: {error_msg}", prefer_markdown=False)
            
    except Exception as e:
        logger.error(f"[TG] Error in handle_voice: {e}")
        await safe_reply(message, f"❌ Ошибка: {e}", prefer_markdown=False)

async def handle_photo(message: types.Message):
    """Handle photo messages"""
    try:
        logger.info(f"[TG] Processing photo message")
        await message.bot.send_chat_action(message.chat.id, "typing")
        await safe_reply(message, "📸 Получил фото. Анализирую содержимое...", prefer_markdown=False)
        
        # Get photo file (highest resolution)
        photo = message.photo[-1]
        photo_file = await message.bot.get_file(photo.file_id)
        pbuf = io.BytesIO()
        await message.bot.download(photo_file, destination=pbuf)
        photo_bytes = pbuf.getvalue()
        photo_b64 = base64.b64encode(photo_bytes).decode('utf-8')
        
        headers = get_auth_headers()
        payload = {
            'message': 'Фотография',
            'image_data': photo_b64,
            'context_search': True,
            'max_context': 3,
            'agent_role': 'coordinator',
            'request_context': {
                'channel': 'telegram',
                'chat_id': message.chat.id,
                'message_id': message.message_id,
                'reply_to_message_id': (message.reply_to_message.message_id if message.reply_to_message else None),
                'user_id': message.from_user.id if message.from_user else None
            }
        }
        
        resp = post_with_auth(f"{API_BASE}/api/ai/chat", json_payload=payload, timeout=1800)
        if resp.status_code == 200:
            response_data = resp.json()
            if isinstance(response_data, str):
                ai_response = response_data
            else:
                ai_response = response_data.get('response', '')
            
            if "обработка временно недоступна" in ai_response:
                final_response = "📸 Фото получено, но система анализа изображений не настроена.\n\nДля работы с изображениями установите необходимые библиотеки:\n```bash\npip install opencv-python pytesseract pillow\n```"
            elif "Ошибка анализа изображения" in ai_response:
                final_response = "📸 Фото получено, но не удалось проанализировать изображение.\n\nВозможные причины:\n• Неподдерживаемый формат файла\n• Поврежденное изображение\n• Отсутствуют необходимые библиотеки\n\nПопробуйте отправить другое изображение."
            else:
                final_response = f"📸 Фото обработано:\n\n{ai_response}"
            
            # Always send plain text in chunks to avoid Telegram Markdown/HTML entity errors
            await reply_chunked_plain(message, final_response)
        else:
            error_msg = resp.text if resp.text else f"HTTP {resp.status_code}"
            await safe_reply(message, f"❌ Ошибка обработки изображения: {error_msg}", prefer_markdown=False)
            
    except Exception as e:
        logger.error(f"[TG] Error in handle_photo: {e}")
        await safe_reply(message, f"❌ Ошибка: {e}", prefer_markdown=False)

async def handle_document(message: types.Message):
    """Handle document messages"""
    try:
        logger.info(f"[TG] Processing document message")
        await message.bot.send_chat_action(message.chat.id, "typing")
        await safe_reply(message, "📄 Получил документ. Обрабатываю...", prefer_markdown=False)
        
        # Get document file
        doc_file = await message.bot.get_file(message.document.file_id)
        dbuf = io.BytesIO()
        await message.bot.download(doc_file, destination=dbuf)
        doc_bytes = dbuf.getvalue()
        doc_b64 = base64.b64encode(doc_bytes).decode('utf-8')
        file_name = message.document.file_name or ''
        doc_type = detect_document_type(file_name, doc_bytes)
        extracted_text, extract_method = try_extract_text(doc_bytes, doc_type)
        
        headers = get_auth_headers()
        payload = {
            'message': 'Документ',
            'document_data': doc_b64,
            'document_name': file_name,
            'document_type': doc_type,
            'document_size': len(doc_bytes),
            'document_text': extracted_text if extracted_text else None,
            'extract_method': extract_method,
            'context_search': True,
            'max_context': 3,
            'agent_role': 'coordinator',
            'request_context': {
                'channel': 'telegram',
                'chat_id': message.chat.id,
                'message_id': message.message_id,
                'reply_to_message_id': (message.reply_to_message.message_id if message.reply_to_message else None),
                'user_id': message.from_user.id if message.from_user else None
            }
        }
        
        resp = post_with_auth(f"{API_BASE}/api/ai/chat", json_payload=payload, timeout=1800)
        if resp.status_code == 200:
            data = resp.json()
            if isinstance(data, str):
                out = data or 'Нет ответа'
            else:
                # Prefer coordinator synthesized response if provided
                out = data.get('response') or data.get('final_response') or 'Нет ответа'
            # Always chunk plain text to avoid Telegram parse/length issues
            await reply_chunked_plain(message, out)
        else:
            await safe_reply(message, f"❌ Ошибка обработки документа: {resp.text}", prefer_markdown=False)
            
    except Exception as e:
        logger.error(f"[TG] Error in handle_document: {e}")
        await safe_reply(message, f"❌ Ошибка: {e}", prefer_markdown=False)

# Command handlers
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    """Handle /start command"""
    try:
        welcome_text = """🤖 Добро пожаловать в SuperBuilder AI!

Я могу помочь с:
• Анализом смет и документов
• Созданием технических документов
• Поиском нормативной документации
• Обработкой файлов (PDF, DOCX, XLSX)
• Голосовыми сообщениями
• И многим другим!

Команды:
/start - приветствие
/history - статистика истории
/clear - очистить историю

Просто отправьте сообщение или файл."""
        
        await safe_reply(message, welcome_text, prefer_markdown=False)
        logger.info(f"[TG] Start command handled for chat {message.chat.id}")
    except Exception as e:
        logger.error(f"[TG] Error in start command: {e}")
        await message.reply("❌ Ошибка при обработке команды")

@dp.message(Command("history"))
async def cmd_history(message: types.Message):
    """Handle /history command - show conversation history stats"""
    try:
        if not CONVERSATION_HISTORY_AVAILABLE or not compressed_conversation_history:
            await safe_reply(message, "❌ История разговоров недоступна", prefer_markdown=False)
            return
            
        chat_id = str(message.chat.id)
        stats = compressed_conversation_history.get_history_stats(chat_id)
        
        history_text = f"""📊 Статистика истории разговора:

• Всего сообщений: {stats.get('total_messages', 0)}
• Токенов: {stats.get('total_tokens', 0)}
• Суммари: {stats.get('summary_count', 0)}
• Недавних сообщений: {stats.get('recent_messages', 0)}

История сохраняется между сессиями и автоматически сжимается для экономии токенов."""
        
        await safe_reply(message, history_text, prefer_markdown=False)
        logger.info(f"[TG] History command handled for chat {message.chat.id}")
    except Exception as e:
        logger.error(f"[TG] Error in history command: {e}")
        await safe_reply(message, "❌ Ошибка при получении истории", prefer_markdown=False)

@dp.message(Command("clear"))
async def cmd_clear(message: types.Message):
    """Handle /clear command - clear conversation history"""
    try:
        if not CONVERSATION_HISTORY_AVAILABLE or not compressed_conversation_history:
            await safe_reply(message, "❌ История разговоров недоступна", prefer_markdown=False)
            return
            
        chat_id = str(message.chat.id)
        compressed_conversation_history.clear_history(chat_id)
        
        # Also clear local history
        if message.chat.id in CHAT_HISTORY:
            CHAT_HISTORY[message.chat.id].clear()
        
        await safe_reply(message, "✅ История разговора очищена", prefer_markdown=False)
        logger.info(f"[TG] Clear command handled for chat {message.chat.id}")
    except Exception as e:
        logger.error(f"[TG] Error in clear command: {e}")
        await safe_reply(message, "❌ Ошибка при очистке истории", prefer_markdown=False)

async def main():
    """Main function to start the bot"""
    if not bot or not dp:
        logger.error("[TG] Bot or dispatcher not initialized")
        return
        
    logger.info(f"[TG] Starting aiogram bot...")
    logger.info(f"[TG] API_BASE: {API_BASE}")
    logger.info(f"[TG] API_USER: {API_USER}")
    
    try:
        # Start message queue processor
        logger.info("[TG] Starting message queue processor...")
        queue_task = asyncio.create_task(process_message_queue())
        
        # Start polling with error handling
        logger.info("[TG] Starting polling...")
        await dp.start_polling(bot, skip_updates=True)
    except Exception as e:
        logger.error(f"[TG] Error starting bot: {e}")
        import traceback
        logger.error(f"[TG] Traceback: {traceback.format_exc()}")
    finally:
        logger.info("[TG] Closing bot session...")
        await bot.session.close()

if __name__ == '__main__':
    if TELEGRAM_BOT_TOKEN and TELEGRAM_BOT_TOKEN != 'YOUR_TELEGRAM_BOT_TOKEN_HERE':
        asyncio.run(main())
    else:
        logger.error("[TG] Telegram bot token not configured")
