# namespace:financial
from typing import Any, Dict, List, Optional
import time
import os
from pathlib import Path
from core.tools.base_tool import ToolManifest, ToolInterface, ToolParam, ToolParamType

# Импортируем улучшенные парсеры
try:
    from core.enhanced_smeta_parser import parse_estimate_enhanced
    from core.enhanced_budget_parser import EnhancedBudgetParser, calculate_budget_from_estimate
    HAS_ENHANCED_PARSERS = True
except ImportError:
    HAS_ENHANCED_PARSERS = False

# Создаем интерфейс для координатора
coordinator_interface = ToolInterface(
    purpose='Создание бюджета на основе одной или нескольких смет с автоматическим расчетом всех статей расходов',
    input_requirements={
        'estimate_files': ToolParam(
            name='estimate_files',
            type=ToolParamType.ARRAY,
            required=True,
            description='Список путей к файлам смет (.xlsx, .xls, .csv)'
        ),
        'project_name': ToolParam(
            name='project_name',
            type=ToolParamType.STRING,
            required=True,
            description='Название проекта'
        )
    },
    execution_flow=[
        'Валидация входных файлов смет',
        'Парсинг всех смет и извлечение данных',
        'Объединение данных из всех смет',
        'Расчет статей бюджета на основе общей стоимости',
        'Создание детализированного бюджета',
        'Экспорт в Excel с формулами',
        'Возврат полного отчета'
    ],
    output_format={
        'status': 'success|error',
        'data': {
            'total_estimate_cost': 'float',
            'budget_breakdown': 'object',
            'excel_file_path': 'string',
            'recommendations': 'array'
        }
    },
    usage_guidelines={
        'for_coordinator': [
            'Используйте для создания бюджета на основе реальных смет',
            'Поддерживает множественные файлы смет',
            'Автоматически рассчитывает все статьи расходов'
        ],
        'for_models': [
            'Инструмент создает полный бюджет на основе смет',
            'Результат сохраняется в Excel файл',
            'Включает детализацию по всем статьям'
        ]
    }
)

manifest = ToolManifest(
    name='create_budget_from_estimates',
    version='1.0.0',
    title='💰 Создание бюджета из смет',
    description='Создание детализированного бюджета на основе одной или нескольких смет с автоматическим расчетом всех статей расходов',
    category='financial',
    ui_placement='dashboard',
    enabled=True,
    system=False,
    entrypoint='tools.custom.create_budget_from_estimates:execute',
    coordinator_interface=coordinator_interface,
    outputs=['total_estimate_cost', 'budget_breakdown', 'excel_file_path', 'recommendations'],
    params=[
        ToolParam(
            name='estimate_files',
            type=ToolParamType.ARRAY,
            required=True,
            description='Список путей к файлам смет',
            ui={
                'placeholder': 'Выберите файлы смет...',
                'fileTypes': ['.xlsx', '.xls', '.csv']
            }
        ),
        ToolParam(
            name='project_name',
            type=ToolParamType.STRING,
            required=True,
            description='Название проекта',
            ui={
                'placeholder': 'Введите название проекта...',
                'maxLength': 200
            }
        ),
        ToolParam(
            name='output_directory',
            type=ToolParamType.STRING,
            required=False,
            description='Директория для сохранения бюджета',
            ui={
                'placeholder': 'Путь для сохранения (по умолчанию: exports/)'
            }
        ),
        ToolParam(
            name='include_monthly_planning',
            type=ToolParamType.BOOLEAN,
            required=False,
            default=True,
            description='Включить месячное планирование (12 месяцев)',
            ui={
                'label': 'Месячное планирование'
            }
        ),
        ToolParam(
            name='profit_margin_percent',
            type=ToolParamType.NUMBER,
            required=False,
            default=15.0,
            description='Процент прибыли от стоимости смет',
            ui={
                'min': 0,
                'max': 50,
                'step': 0.1,
                'suffix': '%'
            }
        )
    ]
)

def execute(**kwargs) -> Dict[str, Any]:
    """Создание бюджета на основе смет"""
    start_time = time.time()
    
    try:
        # Валидация параметров
        estimate_files = kwargs.get('estimate_files', [])
        if not estimate_files:
            return {
                'status': 'error',
                'error': 'Не указаны файлы смет',
                'execution_time': time.time() - start_time
            }
        
        project_name = kwargs.get('project_name', '').strip()
        if not project_name:
            return {
                'status': 'error',
                'error': 'Название проекта не может быть пустым',
                'execution_time': time.time() - start_time
            }
        
        # Дополнительные параметры
        output_directory = kwargs.get('output_directory', 'exports')
        include_monthly_planning = kwargs.get('include_monthly_planning', True)
        profit_margin_percent = max(0, min(50, kwargs.get('profit_margin_percent', 15.0)))
        
        # Создаем директорию для экспорта
        output_dir = Path(output_directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Парсим все сметы
        all_estimates = []
        total_cost = 0.0
        parsed_files = 0
        
        for file_path in estimate_files:
            if not os.path.exists(file_path):
                continue
                
            try:
                if HAS_ENHANCED_PARSERS:
                    # Используем улучшенный парсер
                    estimate_data = parse_estimate_enhanced(file_path)
                else:
                    # Fallback на базовый парсинг
                    estimate_data = _parse_estimate_basic(file_path)
                
                if estimate_data.get('positions'):
                    all_estimates.append(estimate_data)
                    total_cost += estimate_data.get('total_cost', 0)
                    parsed_files += 1
                    
            except Exception as e:
                print(f"Ошибка парсинга файла {file_path}: {e}")
                continue
        
        if parsed_files == 0:
            return {
                'status': 'error',
                'error': 'Не удалось обработать ни одного файла сметы',
                'execution_time': time.time() - start_time
            }
        
        # Создаем объединенные данные сметы
        combined_estimate = {
            'project_name': project_name,
            'total_cost': total_cost,
            'files_count': parsed_files,
            'estimates': all_estimates
        }
        
        # Создаем бюджет на основе сметы
        if HAS_ENHANCED_PARSERS:
            budget_parser = EnhancedBudgetParser()
            budget_result = budget_parser.calculate_budget_from_estimate(combined_estimate)
        else:
            budget_result = _calculate_budget_basic(combined_estimate, profit_margin_percent)
        
        if budget_result.get('status') != 'success':
            return {
                'status': 'error',
                'error': f"Ошибка расчета бюджета: {budget_result.get('error', 'Unknown error')}",
                'execution_time': time.time() - start_time
            }
        
        # Создаем Excel файл с бюджетом
        excel_file_path = output_dir / f"Бюджет_{project_name.replace(' ', '_')}.xlsx"
        
        if HAS_ENHANCED_PARSERS:
            # Создаем детализированный Excel с месячным планированием
            excel_result = _create_detailed_budget_excel(
                budget_result, combined_estimate, excel_file_path, include_monthly_planning
            )
        else:
            # Создаем базовый Excel
            excel_result = _create_basic_budget_excel(budget_result, combined_estimate, excel_file_path)
        
        # Генерируем рекомендации
        recommendations = _generate_budget_recommendations(budget_result, combined_estimate)
        
        execution_time = time.time() - start_time
        
        return {
            'status': 'success',
            'data': {
                'total_estimate_cost': total_cost,
                'budget_breakdown': budget_result.get('budget_items', {}),
                'total_expenses': budget_result.get('total_expenses', 0),
                'net_profit': budget_result.get('net_profit', 0),
                'profit_margin': budget_result.get('profit_margin', 0),
                'excel_file_path': str(excel_file_path),
                'parsed_files_count': parsed_files,
                'recommendations': recommendations
            },
            'execution_time': execution_time,
            'result_type': 'budget_creation',
            'result_title': f'💰 Бюджет проекта: {project_name}',
            'metadata': {
                'project_name': project_name,
                'total_cost': total_cost,
                'files_processed': parsed_files,
                'created_at': time.strftime('%Y-%m-%d %H:%M:%S')
            }
        }
        
    except Exception as e:
        return {
            'status': 'error',
            'error': str(e),
            'execution_time': time.time() - start_time
        }

def _parse_estimate_basic(file_path: str) -> Dict[str, Any]:
    """Базовый парсинг сметы (fallback)"""
    try:
        import pandas as pd
        
        if file_path.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_path)
        elif file_path.endswith('.csv'):
            df = pd.read_csv(file_path, encoding='utf-8')
        else:
            return {'positions': [], 'total_cost': 0}
        
        # Простое извлечение данных
        positions = []
        total_cost = 0
        
        for _, row in df.iterrows():
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[-1]):
                try:
                    cost = float(row.iloc[-1])
                    positions.append({
                        'description': str(row.iloc[0]),
                        'total_cost': cost
                    })
                    total_cost += cost
                except (ValueError, TypeError):
                    continue
        
        return {
            'positions': positions,
            'total_cost': total_cost,
            'format': 'basic'
        }
        
    except Exception as e:
        return {'positions': [], 'total_cost': 0, 'error': str(e)}

def _calculate_budget_basic(estimate_data: Dict[str, Any], profit_margin: float) -> Dict[str, Any]:
    """Базовый расчет бюджета (fallback)"""
    try:
        total_cost = estimate_data.get('total_cost', 0)
        
        # Простые расчеты
        labor_cost = total_cost * 0.12  # 12% ФОТ
        insurance_cost = labor_cost * 0.302  # 30.2% страховые
        travel_cost = total_cost * 0.0276  # 2.76% командировочные
        equipment_cost = total_cost * 0.0024  # 0.24% СИЗ
        
        total_expenses = labor_cost + insurance_cost + travel_cost + equipment_cost
        net_profit = total_cost - total_expenses
        profit_margin_percent = (net_profit / total_cost * 100) if total_cost > 0 else 0
        
        return {
            'status': 'success',
            'base_cost': total_cost,
            'total_expenses': total_expenses,
            'net_profit': net_profit,
            'profit_margin': profit_margin_percent,
            'budget_items': {
                'labor': {'name': 'ФОТ', 'amount': labor_cost, 'percentage': 12.0},
                'insurance': {'name': 'Страховые взносы', 'amount': insurance_cost, 'percentage': 30.2},
                'travel': {'name': 'Командировочные', 'amount': travel_cost, 'percentage': 2.76},
                'equipment': {'name': 'СИЗ', 'amount': equipment_cost, 'percentage': 0.24}
            }
        }
        
    except Exception as e:
        return {'status': 'error', 'error': str(e)}

def _create_detailed_budget_excel(budget_result: Dict[str, Any], estimate_data: Dict[str, Any], 
                                 excel_path: Path, include_monthly: bool) -> Dict[str, Any]:
    """Создание детализированного Excel файла с бюджетом"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Основной лист с бюджетом
        ws = wb.create_sheet('Бюджет_Детализированный')
        
        # Заголовки
        headers = ["Статья", "Сумма (руб.)", "Процент от сметы", "Примечание"]
        if include_monthly:
            headers.extend([f"Месяц {i+1}" for i in range(12)])
            headers.append("Итого")
        
        ws.append(headers)
        
        # Стили
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматирование заголовков
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Добавляем данные
        current_row = 2
        total_cost = estimate_data.get('total_cost', 0)
        budget_items = budget_result.get('budget_items', {})
        
        for item_code, item_data in budget_items.items():
            row_data = [
                item_data['name'],
                item_data['amount'],
                f"{item_data['percentage']:.2f}%",
                f"Расчет на основе сметы"
            ]
            
            if include_monthly:
                monthly_amount = item_data['amount'] / 12
                for month in range(12):
                    row_data.append(monthly_amount)
                row_data.append(item_data['amount'])
            
            ws.append(row_data)
            current_row += 1
        
        # Итоговая строка
        total_expenses = budget_result.get('total_expenses', 0)
        net_profit = budget_result.get('net_profit', 0)
        
        ws.append(["ИТОГО РАСХОДЫ", total_expenses, f"{(total_expenses/total_cost*100):.2f}%", ""])
        ws.append(["ЧИСТАЯ ПРИБЫЛЬ", net_profit, f"{(net_profit/total_cost*100):.2f}%", ""])
        
        # Применяем стили ко всем ячейкам
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(excel_path)
        
        return {'status': 'success', 'file_path': str(excel_path)}
        
    except Exception as e:
        return {'status': 'error', 'error': str(e)}

def _create_basic_budget_excel(budget_result: Dict[str, Any], estimate_data: Dict[str, Any], 
                              excel_path: Path) -> Dict[str, Any]:
    """Создание базового Excel файла с бюджетом"""
    try:
        import pandas as pd
        
        # Создаем DataFrame с данными бюджета
        budget_items = budget_result.get('budget_items', {})
        data = []
        
        for item_code, item_data in budget_items.items():
            data.append({
                'Статья': item_data['name'],
                'Сумма (руб.)': item_data['amount'],
                'Процент': f"{item_data['percentage']:.2f}%"
            })
        
        # Добавляем итоги
        data.append({
            'Статья': 'ИТОГО РАСХОДЫ',
            'Сумма (руб.)': budget_result.get('total_expenses', 0),
            'Процент': f"{(budget_result.get('total_expenses', 0)/estimate_data.get('total_cost', 1)*100):.2f}%"
        })
        
        data.append({
            'Статья': 'ЧИСТАЯ ПРИБЫЛЬ',
            'Сумма (руб.)': budget_result.get('net_profit', 0),
            'Процент': f"{(budget_result.get('net_profit', 0)/estimate_data.get('total_cost', 1)*100):.2f}%"
        })
        
        df = pd.DataFrame(data)
        df.to_excel(excel_path, index=False, sheet_name='Бюджет')
        
        return {'status': 'success', 'file_path': str(excel_path)}
        
    except Exception as e:
        return {'status': 'error', 'error': str(e)}

def _generate_budget_recommendations(budget_result: Dict[str, Any], estimate_data: Dict[str, Any]) -> List[str]:
    """Генерация рекомендаций по бюджету"""
    recommendations = []
    
    total_cost = estimate_data.get('total_cost', 0)
    profit_margin = budget_result.get('profit_margin', 0)
    
    if profit_margin > 80:
        recommendations.append("🎯 Отличная маржа прибыли! Проект очень рентабелен.")
    elif profit_margin > 50:
        recommendations.append("✅ Хорошая маржа прибыли. Проект рентабелен.")
    elif profit_margin > 20:
        recommendations.append("⚠️ Умеренная маржа прибыли. Рассмотрите оптимизацию затрат.")
    else:
        recommendations.append("🚨 Низкая маржа прибыли! Необходима оптимизация бюджета.")
    
    if total_cost > 100_000_000:
        recommendations.append("💰 Крупный проект. Рекомендуется детальное планирование по этапам.")
    elif total_cost > 10_000_000:
        recommendations.append("📊 Средний проект. Учитывайте сезонные факторы.")
    else:
        recommendations.append("🔧 Небольшой проект. Можно использовать упрощенное планирование.")
    
    return recommendations
