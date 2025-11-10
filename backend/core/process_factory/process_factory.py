"""
ProcessFactory - генератор кода для бизнес-процессов
"""

from __future__ import annotations

import logging
import json
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional
import subprocess

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass(slots=True, kw_only=True)
class ProcessSpecification:
    """Описание бизнес-процесса для генерации кода."""

    process_id: str
    process_name: str
    description: str
    process_type: str = "standard"
    inputs: List[Dict[str, Any]] = field(default_factory=list)
    outputs: List[Dict[str, Any]] = field(default_factory=list)
    steps: List[Dict[str, Any]] = field(default_factory=list)


class ProcessFactory:
    """Фабрика для генерации кода бизнес-процессов"""
    
    def __init__(self, base_path: str = "backend/modules"):
        """
        Args:
            base_path: Базовый путь для модулей процессов
        """
        self.base_path = Path(base_path)
        self.templates_path = Path("backend/core/process_factory/templates")
    
    def create_process(
        self,
        *,
        process_id: str,
        process_name: str,
        description: str,
        process_type: str = "standard",
        inputs: Optional[List[Dict[str, str]]] = None,
        outputs: Optional[List[Dict[str, str]]] = None,
        steps: Optional[List[Dict[str, Any]]] = None,
        validate: bool = True,
        register_git: bool = False,
    ) -> Dict[str, Any]:
        """
        Создание нового бизнес-процесса
        
        Args:
            process_id: ID процесса (например, F1.01)
            process_name: Название процесса
            description: Описание процесса
            process_type: Тип процесса (standard, document, calculation, approval, integration)
            inputs: Список входных данных
            outputs: Список выходных данных
            steps: Список шагов процесса
            
        Returns:
            Информация о созданных файлах
        """
        process_dir = self.base_path / process_id.lower().replace(".", "_")
        process_dir.mkdir(parents=True, exist_ok=True)
        
        created_files = []
        
        # Генерация структуры модуля
        files_to_create = [
            ("__init__.py", self._generate_init(process_id, process_name)),
            ("agent.py", self._generate_agent(process_id, process_name, description, steps)),
            ("schemas.py", self._generate_schemas(process_id, inputs, outputs)),
            ("service.py", self._generate_service(process_id, process_name, steps)),
            ("api.py", self._generate_api(process_id, process_name, process_type)),
            ("README.md", self._generate_readme(process_id, process_name, description, inputs, outputs, steps)),
        ]
        
        for filename, content in files_to_create:
            file_path = process_dir / filename
            file_path.write_text(content, encoding="utf-8")
            created_files.append(str(file_path))
            logger.info(f"Created {file_path}")
        
        validation_report: Optional[Dict[str, Any]] = None
        if validate:
            validation_report = self.validate_process(process_dir)

        if register_git:
            self.register_process_in_git(created_files)

        return {
            "process_id": process_id,
            "process_name": process_name,
            "directory": str(process_dir),
            "files": created_files,
            "validation": validation_report,
        }
    
    def create_process_from_spec(self, spec: ProcessSpecification) -> Dict[str, Any]:
        """Создание процесса по спецификации."""
        return self.create_process(
            process_id=spec.process_id,
            process_name=spec.process_name,
            description=spec.description,
            process_type=spec.process_type,
            inputs=spec.inputs,
            outputs=spec.outputs,
            steps=spec.steps,
            validate=True,
            register_git=False,
        )

    def create_process_from_excel(self, excel_path: str | Path) -> Dict[str, Any]:
        """Создание процесса на основе Excel-описания."""
        specification = self.parse_excel_spec(excel_path)
        return self.create_process_from_spec(specification)

    def _generate_init(self, process_id: str, process_name: str) -> str:
        """Генерация __init__.py"""
        return f'''"""
{process_name} - {process_id}
"""

from .agent import {self._to_class_name(process_id)}Agent
from .service import {self._to_class_name(process_id)}Service
from .schemas import *

__all__ = ["{self._to_class_name(process_id)}Agent", "{self._to_class_name(process_id)}Service"]
'''
    
    def _generate_agent(
        self,
        process_id: str,
        process_name: str,
        description: str,
        steps: Optional[List[Dict[str, Any]]]
    ) -> str:
        """Генерация агента процесса"""
        class_name = self._to_class_name(process_id)
        
        steps_code = ""
        if steps:
            for i, step in enumerate(steps, 1):
                step_name = step.get("name", f"Step {i}")
                steps_code += f'        # {step_name}\n'
        
        return f'''"""
Агент для процесса {process_id}: {process_name}
"""

from backend.core.agent import Agent
from typing import Dict, Any
from .service import {class_name}Service


class {class_name}Agent(Agent):
    """Агент для выполнения процесса {process_id}"""
    
    def __init__(self):
        super().__init__(
            agent_id="{process_id.lower()}",
            name="{process_name}",
            description="{description}"
        )
        self.service = {class_name}Service()
    
    def execute(self, task: str, context: Dict[str, Any]) -> Dict[str, Any]:
        """Выполнение процесса {process_id}"""
        # Парсинг входных данных из контекста
        inputs = context.get("inputs", {{}})
        
        # Выполнение шагов процесса
{steps_code}
        # Вызов сервиса для выполнения бизнес-логики
        result = self.service.execute(inputs)
        
        return {{
            "process_id": "{process_id}",
            "task": task,
            "result": result,
            "context": context
        }}
'''
    
    def _generate_schemas(
        self,
        process_id: str,
        inputs: Optional[List[Dict[str, str]]],
        outputs: Optional[List[Dict[str, str]]]
    ) -> str:
        """Генерация Pydantic схем"""
        class_name = self._to_class_name(process_id)
        
        input_fields = ""
        if inputs:
            for inp in inputs:
                field_name = inp.get("name", "field")
                field_type = inp.get("type", "str")
                input_fields += f'    {field_name}: {field_type}\n'
        else:
            input_fields = "    # Добавьте поля входных данных\n"
        
        output_fields = ""
        if outputs:
            for out in outputs:
                field_name = out.get("name", "field")
                field_type = out.get("type", "str")
                output_fields += f'    {field_name}: {field_type}\n'
        else:
            output_fields = "    # Добавьте поля выходных данных\n"
        
        return f'''"""
Pydantic схемы для процесса {process_id}
"""

from pydantic import BaseModel
from typing import Optional


class {class_name}Input(BaseModel):
    """Входные данные для процесса {process_id}"""
{input_fields}


class {class_name}Output(BaseModel):
    """Выходные данные процесса {process_id}"""
{output_fields}
'''
    
    def _generate_service(
        self,
        process_id: str,
        process_name: str,
        steps: Optional[List[Dict[str, Any]]]
    ) -> str:
        """Генерация сервиса процесса"""
        class_name = self._to_class_name(process_id)
        
        steps_code = ""
        if steps:
            for i, step in enumerate(steps, 1):
                step_name = step.get("name", f"Step {i}")
                steps_code += f'        # Шаг {i}: {step_name}\n'
                steps_code += f'        # TODO: Реализовать логику шага\n\n'
        
        return f'''"""
Сервис для процесса {process_id}: {process_name}
"""

from typing import Dict, Any
from .schemas import {class_name}Input, {class_name}Output
import logging

logger = logging.getLogger(__name__)


class {class_name}Service:
    """Сервис для выполнения бизнес-логики процесса {process_id}"""
    
    def execute(self, inputs: Dict[str, Any]) -> Dict[str, Any]:
        """
        Выполнение процесса {process_id}
        
        Args:
            inputs: Входные данные процесса
            
        Returns:
            Результат выполнения процесса
        """
        logger.info(f"Executing process {process_id}")
        
{steps_code}
        # TODO: Реализовать бизнес-логику процесса
        
        return {{
            "status": "completed",
            "process_id": "{process_id}",
            "result": {{}}
        }}
'''
    
    def _generate_api(self, process_id: str, process_name: str, process_type: str) -> str:
        """Генерация API эндпоинтов"""
        class_name = self._to_class_name(process_id)
        route_name = process_id.lower().replace(".", "_")
        tag_name = f"{process_type}:{process_id}"
        
        return f'''"""
API эндпоинты для процесса {process_id}: {process_name}
"""

from fastapi import APIRouter, Depends, HTTPException
from backend.middleware.rbac import get_current_user, require_permission
from backend.models.auth import User
from .schemas import {class_name}Input, {class_name}Output
from .service import {class_name}Service
import logging

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/{route_name}", tags=["{process_id}"])
PROCESS_TAG = "{tag_name}"

service = {class_name}Service()


@router.post("/execute", response_model={class_name}Output)
@require_permission("process", "execute")
async def execute_process(
    inputs: {class_name}Input,
    current_user: User = Depends(get_current_user)
):
    """Выполнение процесса {process_id}"""
    try:
        result = service.execute(inputs.dict())
        return {class_name}Output(**result)
    except Exception as e:
        logger.error(f"Process execution error: {{e}}")
        raise HTTPException(status_code=500, detail=str(e))
'''
    
    def _generate_readme(
        self,
        process_id: str,
        process_name: str,
        description: str,
        inputs: Optional[List[Dict[str, str]]],
        outputs: Optional[List[Dict[str, str]]],
        steps: Optional[List[Dict[str, Any]]]
    ) -> str:
        """Генерация README для процесса"""
        readme = f'''# {process_id}: {process_name}

{description}

## Входные данные

'''
        if inputs:
            for inp in inputs:
                readme += f'- **{inp.get("name")}** ({inp.get("type", "str")}): {inp.get("description", "")}\n'
        else:
            readme += "- Нет входных данных\n"
        
        readme += "\n## Выходные данные\n\n"
        if outputs:
            for out in outputs:
                readme += f'- **{out.get("name")}** ({out.get("type", "str")}): {out.get("description", "")}\n'
        else:
            readme += "- Нет выходных данных\n"
        
        readme += "\n## Шаги процесса\n\n"
        if steps:
            for i, step in enumerate(steps, 1):
                readme += f'{i}. {step.get("name", f"Step {i}")}\n'
        else:
            readme += "1. TODO: Определить шаги процесса\n"
        
        readme += f'''
## API

- `POST /api/{process_id.lower().replace(".", "_")}/execute` - Выполнение процесса

## Использование

```python
from backend.modules.{process_id.lower().replace(".", "_")} import {self._to_class_name(process_id)}Agent

agent = {self._to_class_name(process_id)}Agent()
result = agent.execute("task", {{"inputs": {{}}}})
```
'''
        return readme
    
    def parse_excel_spec(self, excel_path: str | Path) -> ProcessSpecification:
        """Парсинг Excel-файла с описанием процесса."""
        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        workbook = load_workbook(excel_path, data_only=True)
        sheets = {sheet.title.lower(): sheet for sheet in workbook.worksheets}

        overview_sheet = sheets.get("overview") or workbook.active
        overview = self._parse_overview_sheet(overview_sheet)

        inputs = self._parse_table_sheet(sheets.get("inputs"))
        outputs = self._parse_table_sheet(sheets.get("outputs"))
        steps = self._parse_steps_sheet(sheets.get("steps"))

        return ProcessSpecification(
            process_id=overview["process_id"],
            process_name=overview["process_name"],
            description=overview["description"],
            process_type=overview.get("process_type", "standard"),
            inputs=inputs,
            outputs=outputs,
            steps=steps,
        )

    def _parse_overview_sheet(self, sheet) -> Dict[str, str]:
        data: Dict[str, str] = {
            "process_id": "",
            "process_name": "",
            "description": "",
            "process_type": "standard",
        }
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            key = str(row[0]).strip().lower()
            value = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
            match key:
                case "id" | "process_id":
                    data["process_id"] = value
                case "name" | "process_name":
                    data["process_name"] = value
                case "description":
                    data["description"] = value
                case "type" | "process_type":
                    data["process_type"] = value or data["process_type"]
        missing = [k for k, v in data.items() if not v and k in {"process_id", "process_name", "description"}]
        if missing:
            raise ValueError(f"Missing required overview fields: {', '.join(missing)}")
        return data

    def _parse_table_sheet(self, sheet) -> List[Dict[str, Any]]:
        if sheet is None:
            return []
        headers = [str(cell.value).strip().lower() for cell in sheet[1] if cell.value]
        records: List[Dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            record = {}
            for header, value in zip(headers, row):
                record[header] = value
            records.append(record)
        return records

    def _parse_steps_sheet(self, sheet) -> List[Dict[str, Any]]:
        steps = self._parse_table_sheet(sheet)
        enriched_steps: List[Dict[str, Any]] = []
        for index, step in enumerate(steps, start=1):
            name = step.get("name") or step.get("step") or f"Step {index}"
            description = step.get("description", "")
            owner = step.get("owner") or step.get("responsible")
            enriched_steps.append(
                {
                    "name": str(name),
                    "description": description,
                    "owner": owner,
                    "order": index,
                }
            )
        return enriched_steps

    def validate_process(self, process_dir: Path | str) -> Dict[str, Any]:
        """Базовая валидация созданного процесса."""
        process_path = Path(process_dir)
        results: Dict[str, Any] = {"directory": str(process_path), "files": []}
        required_files = {"__init__.py", "agent.py", "schemas.py", "service.py", "api.py", "README.md"}

        for path in process_path.glob("*.py"):
            status = {"file": str(path), "exists": True, "compiled": True}
            try:
                compile(path.read_text(encoding="utf-8"), str(path), "exec")
            except SyntaxError as exc:
                status["compiled"] = False
                status["error"] = str(exc)
            results["files"].append(status)

        for name in required_files:
            file_path = process_path / name
            if not file_path.exists():
                results.setdefault("missing", []).append(str(file_path))

        results["status"] = "ok" if not results.get("missing") and all(f["compiled"] for f in results["files"]) else "attention"
        logger.info("Validation report for %s: %s", process_path, results["status"])
        return results

    def register_process_in_git(self, file_paths: List[str]) -> None:
        """Добавление созданных файлов в git-индекс."""
        if not file_paths:
            return
        try:
            subprocess.run(["git", "add", *file_paths], check=True)
            logger.info("Registered process files in git: %s", file_paths)
        except (subprocess.CalledProcessError, FileNotFoundError) as exc:
            logger.warning("Failed to register files in git: %s", exc)

    def _to_class_name(self, process_id: str) -> str:
        """Преобразование ID процесса в имя класса"""
        # F1.01 -> F1_01 -> F101
        return process_id.replace(".", "_").replace("-", "_").title().replace("_", "")


# Глобальный экземпляр фабрики
process_factory = ProcessFactory()

