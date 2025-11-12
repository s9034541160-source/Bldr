"""
Модуль загрузки ведомостей объёмов из табличных источников (CSV, XLS, XLSX).

Поддерживает автоматическое определение колонок и формирование WorkVolumeResult.
"""

from __future__ import annotations

import csv
import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import ezdxf
import fitz
import ifcopenshell
import xlrd
from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook

from backend.services.ocr.deepseek_service import deepseek_ocr_service
from backend.services.work_volume_extractor import (
    CATEGORY_HINTS,
    WorkVolumeEntry,
    WorkVolumeResult,
    WorkVolumeSummary,
)
from backend.services.gesn_ingestion import GESNCatalogLoader, GESNImportBundle

logger = logging.getLogger(__name__)


NAME_ALIASES = {"name", "work", "работа", "наименование", "описание", "description"}
QUANTITY_ALIASES = {"quantity", "qty", "объем", "объём", "количество", "volume"}
UNIT_ALIASES = {"unit", "ед", "ед.изм.", "ед. изм.", "единица", "единица измерения", "ед. измерения"}
CODE_ALIASES = {"code", "код", "шифр"}


def _parse_decimal(value: str) -> Decimal:
    normalized = value.replace(" ", "").replace(",", ".")
    return Decimal(normalized)


def _detect_category(code: Optional[str]) -> Optional[str]:
    if not code:
        return None
    digits = "".join(filter(str.isdigit, code))
    if len(digits) < 2:
        return None
    return CATEGORY_HINTS.get(digits[:2])


def _detect_columns(headers: Iterable[str]) -> Dict[str, str]:
    normalized = {header: header.strip().lower() for header in headers if header}
    mapping: Dict[str, str] = {}
    for header, normalized_name in normalized.items():
        if normalized_name in NAME_ALIASES:
            mapping["name"] = header
        elif normalized_name in QUANTITY_ALIASES:
            mapping["quantity"] = header
        elif normalized_name in UNIT_ALIASES:
            mapping["unit"] = header
        elif normalized_name in CODE_ALIASES:
            mapping["code"] = header
    if "name" not in mapping or "quantity" not in mapping:
        raise ValueError("Не удалось определить колонки 'наименование' и 'количество' в ведомости.")
    return mapping


def _row_to_entry(row: Dict[str, str], columns: Dict[str, str], source: str) -> Optional[WorkVolumeEntry]:
    try:
        name = row[columns["name"]].strip()
        if not name:
            return None
        quantity_raw = row[columns["quantity"]]
        quantity = _parse_decimal(quantity_raw)
        unit = row[columns.get("unit", "")].strip() if columns.get("unit") else ""
        code = row[columns.get("code", "")].strip() if columns.get("code") else None
        category = _detect_category(code)
        metadata = {"source": source}
        return WorkVolumeEntry(
            name=name,
            quantity=quantity,
            unit=unit or "ед.",
            source_line=str(row),
            code=code,
            category=category,
            metadata=metadata,
        )
    except (KeyError, InvalidOperation, ValueError):
        return None


def _build_summary(entries: Sequence[WorkVolumeEntry]) -> WorkVolumeSummary:
    totals: Dict[str, Decimal] = {}
    total_quantity = Decimal("0")
    for entry in entries:
        total_quantity += entry.quantity
        category = entry.category or "Прочее"
        totals.setdefault(category, Decimal("0"))
        totals[category] += entry.quantity
    return WorkVolumeSummary(totals=totals, total_quantity=total_quantity)


class SpreadsheetVolumeExtractor:
    """
    Извлекает объёмы работ из CSV/XLS/XLSX ведомостей.
    """

    def extract(self, file_path: str) -> WorkVolumeResult:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Файл {file_path} не найден.")

        rows = list(self._read_rows(path))
        if not rows:
            logger.warning("В файле %s не найдено данных.", file_path)
            return WorkVolumeResult(entries=[], summary=WorkVolumeSummary(), warnings=["Пустой файл ведомости."])

        columns = _detect_columns(rows[0].keys())
        entries: List[WorkVolumeEntry] = []
        warnings: List[str] = []

        for row in rows:
            entry = _row_to_entry(row, columns, source="spreadsheet")
            if entry:
                entries.append(entry)
            else:
                warnings.append(f"Строка пропущена: {row}")

        summary = _build_summary(entries)
        return WorkVolumeResult(entries=entries, summary=summary, warnings=warnings)

    # ------------------------------------------------------------------ #
    # Чтение данных
    # ------------------------------------------------------------------ #

    def _read_rows(self, path: Path) -> Iterable[Dict[str, str]]:
        ext = path.suffix.lower()
        if ext == ".csv":
            yield from self._read_csv(path)
        elif ext in {".xlsx", ".xlsm", ".xltx"}:
            yield from self._read_openpyxl(path)
        elif ext == ".xls":
            yield from self._read_xls(path)
        else:
            raise ValueError(f"Неподдерживаемый формат для ведомости: {path.suffix}")

    def _read_csv(self, path: Path) -> Iterable[Dict[str, str]]:
        with path.open("r", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                yield {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}

    def _read_openpyxl(self, path: Path) -> Iterable[Dict[str, str]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for row in sheet.iter_rows(min_row=2):
            values = {}
            for header, cell in zip(headers, row):
                values[header] = str(cell.value).strip() if cell.value is not None else ""
            yield values

    def _read_xls(self, path: Path) -> Iterable[Dict[str, str]]:
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
        for row_idx in range(1, sheet.nrows):
            values = {}
            for col_idx, header in enumerate(headers):
                cell_value = sheet.cell_value(row_idx, col_idx)
                if isinstance(cell_value, float):
                    values[header] = str(cell_value)
                else:
                    values[header] = str(cell_value).strip()
            yield values

    # ------------------------------------------------------------------ #
    # Преобразование строки в WorkVolumeEntry
    # ------------------------------------------------------------------ #

spreadsheet_volume_extractor = SpreadsheetVolumeExtractor()

class WordVolumeExtractor:
    """
    Извлекает объёмы работ из DOC/DOCX документов (таблицы).
    """

    def extract(self, file_path: str) -> WorkVolumeResult:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Файл {file_path} не найден.")

        ext = path.suffix.lower()
        if ext == ".docx":
            rows_iter = self._read_docx(path)
        elif ext == ".doc":
            rows_iter = self._read_doc(path)
        else:
            raise ValueError(f"Неподдерживаемый формат Word документа: {ext}")

        entries: List[WorkVolumeEntry] = []
        warnings: List[str] = []

        for headers, rows in rows_iter:
            try:
                columns = _detect_columns(headers)
            except ValueError:
                warnings.append(f"Таблица пропущена: не найдены колонки в заголовке {headers}")
                continue

            for row in rows:
                entry = _row_to_entry(row, columns, source="word")
                if entry:
                    entries.append(entry)
                else:
                    warnings.append(f"Строка пропущена: {row}")

        summary = _build_summary(entries)
        return WorkVolumeResult(entries=entries, summary=summary, warnings=warnings)

    def _read_docx(self, path: Path) -> Iterable[tuple[List[str], List[Dict[str, str]]]]:
        document = Document(path)
        for table in document.tables:
            headers = [cell.text.strip() for cell in table.rows[0].cells]
            rows: List[Dict[str, str]] = []
            for row in table.rows[1:]:
                row_data = {}
                for idx, header in enumerate(headers):
                    text = row.cells[idx].text.strip() if idx < len(row.cells) else ""
                    row_data[header] = text
                rows.append(row_data)
            yield headers, rows

    def _read_doc(self, path: Path) -> Iterable[tuple[List[str], List[Dict[str, str]]]]:
        with path.open("rb") as doc_file:
            import mammoth

            html = mammoth.convert_to_html(doc_file).value
        soup = BeautifulSoup(html, "html.parser")
        for table in soup.find_all("table"):
            header_cells = table.find("tr")
            if not header_cells:
                continue
            headers = [cell.get_text(strip=True) for cell in header_cells.find_all(["th", "td"])]
            rows: List[Dict[str, str]] = []
            for tr in table.find_all("tr")[1:]:
                cells = tr.find_all(["td", "th"])
                row_data = {}
                for header, cell in zip(headers, cells):
                    row_data[header] = cell.get_text(strip=True)
                rows.append(row_data)
            yield headers, rows


word_volume_extractor = WordVolumeExtractor()


class PDFOCRVolumeExtractor:
    """
    Извлекает объёмы работ из PDF-документов (сканы/таблицы).
    """

    JSON_PROMPT = (
        "Ты инженер-сметчик. На изображении ведомость объёмов работ. "
        "Верни строго валидный JSON формата: {\"rows\": [{\"name\": \"...\", \"quantity\": \"...\", \"unit\": \"...\", \"code\": \"...\"}]}."
        "Если данных нет — верни {\"rows\": []}. Используй анализ таблицы, обработай все страницы."
    )

    def extract(self, file_path: str) -> WorkVolumeResult:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Файл {file_path} не найден.")

        entries: List[WorkVolumeEntry] = []
        warnings: List[str] = []

        with fitz.open(path) as doc:
            images: List[bytes] = []
            for page in doc:
                pixmap = page.get_pixmap(dpi=300)
                images.append(pixmap.tobytes("png"))

        try:
            raw_response = deepseek_ocr_service.extract_structured(
                b"".join(images), instruction=self.JSON_PROMPT, max_new_tokens=4096
            )
            data = self._safe_json_load(raw_response)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"OCR не удалось распознать документ: {exc}")
            return WorkVolumeResult(entries=[], summary=WorkVolumeSummary(), warnings=warnings)

        for row in data.get("rows", []):
            entry = self._row_to_entry(row)
            if entry:
                entries.append(entry)
            else:
                warnings.append(f"Строка пропущена: {row}")

        summary = _build_summary(entries)
        return WorkVolumeResult(entries=entries, summary=summary, warnings=warnings)

    def _safe_json_load(self, payload: str) -> Dict[str, object]:
        start = payload.find("{")
        end = payload.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise ValueError("Ответ LLM не содержит корректный JSON.")
        json_fragment = payload[start : end + 1]
        return json.loads(json_fragment)

    def _row_to_entry(self, row: Dict[str, str]) -> Optional[WorkVolumeEntry]:
        try:
            name = row.get("name", "").strip()
            if not name:
                return None
            quantity = _parse_decimal(row.get("quantity", "0"))
            unit = row.get("unit", "").strip() or "ед."
            code = row.get("code")
            category = _detect_category(code)
            metadata = {"source": "pdf_ocr"}
            return WorkVolumeEntry(
                name=name,
                quantity=quantity,
                unit=unit,
                source_line=str(row),
                code=code,
                category=category,
                metadata=metadata,
            )
        except (InvalidOperation, ValueError):
            return None


pdf_ocr_volume_extractor = PDFOCRVolumeExtractor()


class CADVolumeExtractor:
    """
    Извлекает ведомость объёмов работ из DWG/IFC документов (ezdxf/ifcopenshell).

    Стратегия:
        - DWG: ищем таблицы (CLASS 'TABLE'), извлекаем строки с единицами измерения.
        - IFC: читаем элементы и извлекаем объёмы (например, IfcQuantityVolume).
    """

    def extract(self, file_path: str) -> WorkVolumeResult:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Файл {file_path} не найден.")

        ext = path.suffix.lower()
        if ext == ".dwg":
            entries, warnings = self._extract_dwg(path)
        elif ext == ".ifc":
            entries, warnings = self._extract_ifc(path)
        else:
            raise ValueError(f"Неподдерживаемый формат CAD-файла: {ext}")

        summary = _build_summary(entries)
        return WorkVolumeResult(entries=entries, summary=summary, warnings=warnings)

    def _extract_dwg(self, path: Path) -> Tuple[List[WorkVolumeEntry], List[str]]:
        doc = ezdxf.readfile(path)
        tables = doc.objects.query("TABLE")
        entries: List[WorkVolumeEntry] = []
        warnings: List[str] = []

        for table in tables:
            for row in table.rows():
                cells = [cell.text for cell in row.cells]
                if len(cells) < 3:
                    continue
                name = cells[0].strip()
                quantity_raw = cells[1].strip()
                unit = cells[2].strip()
                try:
                    quantity = _parse_decimal(quantity_raw)
                except InvalidOperation:
                    continue
                entry = WorkVolumeEntry(
                    name=name,
                    quantity=quantity,
                    unit=unit or "ед.",
                    source_line=str(cells),
                    code=None,
                    category=_detect_category(None),
                    metadata={"source": "dwg"},
                )
                entries.append(entry)

        if not entries:
            warnings.append("Не удалось извлечь таблицы из DWG.")
        return entries, warnings

    def _extract_ifc(self, path: Path) -> Tuple[List[WorkVolumeEntry], List[str]]:
        model = ifcopenshell.open(str(path))
        entries: List[WorkVolumeEntry] = []
        warnings: List[str] = []

        quantities = model.by_type("IfcQuantityVolume")
        if not quantities:
            warnings.append("В модели IFC отсутствуют IfcQuantityVolume.")
            return entries, warnings

        for quantity in quantities:
            name = quantity.Name or "Без названия"
            volume = quantity.VolumeValue
            unit = quantity.Unit.name if quantity.Unit else "м³"
            entry = WorkVolumeEntry(
                name=name,
                quantity=Decimal(str(volume)),
                unit=unit,
                source_line=f"IFC GUID: {quantity.GlobalId}" if hasattr(quantity, "GlobalId") else "IFC",
                code=None,
                category=_detect_category(None),
                metadata={"source": "ifc"},
            )
            entries.append(entry)

        return entries, warnings


cad_volume_extractor = CADVolumeExtractor()

__all__ = [
    "SpreadsheetVolumeExtractor",
    "spreadsheet_volume_extractor",
    "WordVolumeExtractor",
    "word_volume_extractor",
    "PDFOCRVolumeExtractor",
    "pdf_ocr_volume_extractor",
    "CADVolumeExtractor",
    "cad_volume_extractor",
]


