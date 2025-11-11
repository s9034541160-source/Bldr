"""
Экспорт календарного плана в Excel с диаграммой Ганта.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.services.calendar_plan.cpm_builder import CPMResult
from backend.services.calendar_plan.resource_leveler import LevelingResult


class GanttExcelExporter:
    """
    Формирует Excel-файл с данными задач и визуальным представлением Ганта.
    """

    def __init__(self, base_date: Optional[dt.date] = None) -> None:
        self.base_date = base_date or dt.date.today()

    def export(
        self,
        cpm_result: CPMResult,
        output_path: str | Path,
        *,
        leveled: Optional[LevelingResult] = None,
        project_name: str = "BLDR Calendar Plan",
    ) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Gantt"

        headers = [
            "ID",
            "Название",
            "Код",
            "Уровень",
            "Начало",
            "Окончание",
            "Длительность (дни)",
            "Критическая",
            "Ранний старт",
            "Ранний финиш",
            "Поздний старт",
            "Поздний финиш",
            "Резерв (дни)",
        ]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        leveled_map = {task.task_id: task for task in (leveled.tasks if leveled else [])}
        project_days = int(round(cpm_result.project_duration)) + 1
        max_row = 1

        for idx, node_id in enumerate(cpm_result.graph.nodes, start=2):
            node = cpm_result.nodes[node_id]
            leveled_task = leveled_map.get(node_id)
            start = leveled_task.new_early_start if leveled_task else node.early_start
            finish = leveled_task.new_early_finish if leveled_task else node.early_finish

            row = [
                node_id,
                node.task.name,
                node.task.code,
                node.task.level,
                self._format_date(start),
                self._format_date(finish),
                round(finish - start, 2),
                "Да" if node.critical else "Нет",
                self._format_date(node.early_start),
                self._format_date(node.early_finish),
                self._format_date(node.late_start),
                self._format_date(node.late_finish),
                round(node.slack, 2),
            ]
            ws.append(row)

            for day_offset in range(project_days):
                current_col = len(headers) + 1 + day_offset
                date_label = self.base_date + dt.timedelta(days=day_offset)
                header_cell = ws.cell(row=1, column=current_col)
                header_cell.value = date_label.strftime("%d.%m")
                header_cell.font = Font(size=9)
                header_cell.alignment = Alignment(horizontal="center")

                if start <= day_offset < finish:
                    cell = ws.cell(row=idx, column=current_col)
                    cell.fill = PatternFill(
                        start_color="1F77B4" if node.critical else "A6CEE3",
                        end_color="1F77B4" if node.critical else "A6CEE3",
                        fill_type="solid",
                    )

            max_row = idx

        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        ws.freeze_panes = "A2"

        wb.save(path)
        return path

    def _format_date(self, project_days: float) -> str:
        target = self.base_date + dt.timedelta(days=project_days)
        return target.strftime("%d.%m.%Y")


__all__ = ["GanttExcelExporter"]


