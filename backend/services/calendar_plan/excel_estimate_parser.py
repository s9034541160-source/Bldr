"""
Парсер смет в формате Excel для подготовки календарного плана (F1.07).
"""

from __future__ import annotations

import datetime as dt
import logging
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class EstimateItem:
    """
    Представление строки из сметы, пригодной для построения календарного плана.
    """

    code: Optional[str]
    name: str
    unit: Optional[str]
    quantity: Optional[Decimal]
    duration_days: Optional[float]
    start_date: Optional[dt.date]
    finish_date: Optional[dt.date]
    level: int = 0
    raw_row: Dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class EstimateParseResult:
    """
    Результат разбора Excel-файла сметы.
    """

    items: List[EstimateItem]
    warnings: List[str]
    header_row: Optional[int]
    source_path: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        """
        Конвертирует результат в словарь для сериализации.
        """
        return {
            "items": [item.__dict__ for item in self.items],
            "warnings": self.warnings,
            "header_row": self.header_row,
            "source_path": self.source_path,
        }


class ExcelEstimateParser:
    """
    Считывает табличные сметы и извлекает ключевые столбцы для дальнейшего планирования.

    Поддерживаются следующие сценарии:
    - Пользовательские названия колонок (используем множество синонимов).
    - Иерархия сметы через код (1, 1.1, 1.1.1) или отступы в названии.
    - Даты в Excel-формате, строки или Python datetime.
    """

    # Карта синонимов заголовков -> унифицированное имя
    HEADER_ALIASES: Dict[str, str] = {
        "код": "code",
        "шифр": "code",
        "№": "code",
        "номер": "code",
        "поз": "code",
        "работа": "name",
        "наименование": "name",
        "наименование работ": "name",
        "вид работ": "name",
        "описание": "name",
        "ед": "unit",
        "ед.": "unit",
        "ед изм": "unit",
        "ед. изм.": "unit",
        "единица измерения": "unit",
        "количество": "quantity",
        "объем": "quantity",
        "объём": "quantity",
        "кол-во": "quantity",
        "кол": "quantity",
        "продолжительность": "duration",
        "длительность": "duration",
        "длительность (дн)": "duration",
        "продолжительность (дн)": "duration",
        "начало": "start",
        "start": "start",
        "дата начала": "start",
        "окончание": "finish",
        "завершение": "finish",
        "finish": "finish",
        "дата окончания": "finish",
    }

    REQUIRED_HEADERS = {"name"}

    def __init__(self, *, data_only: bool = True) -> None:
        self.data_only = data_only

    # ------------------------------------------------------------------ #
    # Публичные методы
    # ------------------------------------------------------------------ #

    def parse(self, file_path: str | Path) -> EstimateParseResult:
        """
        Загружает книгу Excel и возвращает структурированный список работ.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Файл сметы {path} не найден.")

        workbook = load_workbook(filename=path, data_only=self.data_only)
        sheet = workbook.active

        header_map, header_row = self._detect_header(sheet)
        if not header_map:
            warning = "Не удалось определить строку заголовков в смете."
            logger.warning("ExcelEstimateParser: %s (%s)", warning, path)
            return EstimateParseResult(items=[], warnings=[warning], header_row=None, source_path=str(path))

        items: List[EstimateItem] = []
        warnings: List[str] = []

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_data = self._extract_row(sheet, row_idx, header_map)
            if row_data is None:
                continue

            item, row_warnings = row_data
            if row_warnings:
                warnings.extend(row_warnings)
            items.append(item)

        if not items:
            warnings.append("После строки заголовков не найдено данных по работам.")

        return EstimateParseResult(
            items=items,
            warnings=warnings,
            header_row=header_row + 1,  # human-friendly (1-based)
            source_path=str(path),
        )

    # ------------------------------------------------------------------ #
    # Внутренние процедуры
    # ------------------------------------------------------------------ #

    def _detect_header(self, sheet) -> Tuple[Dict[str, int], Optional[int]]:
        """
        Находит строку заголовков и возвращает карту {canonical_name: column_index}.
        """
        for row in range(1, min(sheet.max_row, 20) + 1):
            header_candidates: Dict[str, int] = {}
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=col).value
                if value is None:
                    continue
                normalized = self._normalize_header(str(value))
                if not normalized:
                    continue
                canonical = self.HEADER_ALIASES.get(normalized)
                if canonical and canonical not in header_candidates:
                    header_candidates[canonical] = col
            if self.REQUIRED_HEADERS.issubset(header_candidates.keys()):
                return header_candidates, row
        return {}, None

    def _extract_row(
        self,
        sheet,
        row_idx: int,
        header_map: Dict[str, int],
    ) -> Optional[Tuple[EstimateItem, List[str]]]:
        """
        Извлекает данные из строки и формирует EstimateItem.
        """
        name_col = header_map.get("name")
        if not name_col:
            return None
        name_value = sheet.cell(row=row_idx, column=name_col).value
        if name_value is None:
            return None

        name_text = str(name_value).strip()
        if not name_text:
            return None

        warnings: List[str] = []
        raw: Dict[str, Any] = {}

        code = self._read_cell(sheet, row_idx, header_map.get("code"), raw, "code")
        unit = self._read_cell(sheet, row_idx, header_map.get("unit"), raw, "unit")
        quantity = self._read_decimal(sheet, row_idx, header_map.get("quantity"), raw, "quantity", warnings)
        duration = self._read_float(sheet, row_idx, header_map.get("duration"), raw, "duration_days", warnings)
        start = self._read_date(sheet, row_idx, header_map.get("start"), raw, "start_date", warnings)
        finish = self._read_date(sheet, row_idx, header_map.get("finish"), raw, "finish_date", warnings)

        level = self._infer_level(code, name_text)

        item = EstimateItem(
            code=code,
            name=name_text.strip(),
            unit=unit,
            quantity=quantity,
            duration_days=duration,
            start_date=start,
            finish_date=finish,
            level=level,
            raw_row=raw,
        )
        return item, warnings

    def _read_cell(
        self,
        sheet,
        row_idx: int,
        column: Optional[int],
        raw: Dict[str, Any],
        key: str,
    ) -> Optional[str]:
        if column is None:
            return None
        value = sheet.cell(row=row_idx, column=column).value
        raw[key] = value
        if value is None:
            return None
        return str(value).strip()

    def _read_decimal(
        self,
        sheet,
        row_idx: int,
        column: Optional[int],
        raw: Dict[str, Any],
        key: str,
        warnings: List[str],
    ) -> Optional[Decimal]:
        if column is None:
            return None
        value = sheet.cell(row=row_idx, column=column).value
        raw[key] = value
        if value is None:
            return None
        try:
            return Decimal(str(value)).quantize(Decimal("0.0001"))
        except (InvalidOperation, ValueError):
            warnings.append(f"Не удалось преобразовать значение '{value}' в Decimal (строка {row_idx}).")
            return None

    def _read_float(
        self,
        sheet,
        row_idx: int,
        column: Optional[int],
        raw: Dict[str, Any],
        key: str,
        warnings: List[str],
    ) -> Optional[float]:
        if column is None:
            return None
        value = sheet.cell(row=row_idx, column=column).value
        raw[key] = value
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            warnings.append(f"Не удалось прочитать продолжительность '{value}' (строка {row_idx}).")
            return None

    def _read_date(
        self,
        sheet,
        row_idx: int,
        column: Optional[int],
        raw: Dict[str, Any],
        key: str,
        warnings: List[str],
    ) -> Optional[dt.date]:
        if column is None:
            return None
        value = sheet.cell(row=row_idx, column=column).value
        raw[key] = value
        if value is None:
            return None
        if isinstance(value, dt.datetime):
            return value.date()
        if isinstance(value, dt.date):
            return value
        if isinstance(value, str):
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return dt.datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue
        warnings.append(f"Не удалось распознать дату '{value}' (строка {row_idx}).")
        return None

    @staticmethod
    def _normalize_header(value: str) -> str:
        """
        Приводит заголовок к нижнему регистру и убирает лишние символы.
        """
        return (
            value.strip()
            .lower()
            .replace("\n", " ")
            .replace("\r", " ")
            .replace("  ", " ")
            .replace(".", "")
            .replace(",", "")
        )

    @staticmethod
    def _infer_level(code: Optional[str], name: str) -> int:
        """
        Определяет уровень вложенности на основе кода или отступа.
        """
        if code:
            stripped = code.strip()
            if stripped:
                return stripped.count(".")
        leading_spaces = len(name) - len(name.lstrip())
        if leading_spaces > 0:
            return max(0, leading_spaces // 4)
        return 0


__all__ = ["ExcelEstimateParser", "EstimateItem", "EstimateParseResult"]


