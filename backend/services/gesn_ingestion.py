"""
GESN catalog ingestion utilities.

Supports importing normative data from CSV/Excel, XML and PDF sources.
"""

from __future__ import annotations

import csv
import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

import fitz
from lxml import etree
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.models.gesn import (
    GESNNorm,
    GESNNormComponent,
    GESNNormResource,
    GESNSection,
)
from backend.services.ocr.deepseek_service import deepseek_ocr_service

logger = logging.getLogger(__name__)


# ------------------------------------------------------------------------------
# Data models used by importers
# ------------------------------------------------------------------------------


@dataclass(slots=True)
class SectionRecord:
    code: str
    name: str
    parent_code: Optional[str] = None
    description: Optional[str] = None
    metadata: Dict[str, object] = field(default_factory=dict)


@dataclass(slots=True)
class NormRecord:
    code: str
    name: str
    unit: str
    section_code: str
    unit_description: Optional[str] = None
    work_type: Optional[str] = None
    description: Optional[str] = None
    composition: Optional[str] = None
    parameters: Dict[str, object] = field(default_factory=dict)
    metadata: Dict[str, object] = field(default_factory=dict)
    components: List[str] = field(default_factory=list)


@dataclass(slots=True)
class ResourceRecord:
    norm_code: str
    resource_type: str  # material / labor / machine
    name: str
    unit: str
    quantity: Optional[float] = None
    code: Optional[str] = None
    metadata: Dict[str, object] = field(default_factory=dict)


@dataclass(slots=True)
class GESNImportBundle:
    sections: List[SectionRecord] = field(default_factory=list)
    norms: List[NormRecord] = field(default_factory=list)
    resources: List[ResourceRecord] = field(default_factory=list)

    def extend(self, other: "GESNImportBundle") -> None:
        self.sections.extend(other.sections)
        self.norms.extend(other.norms)
        self.resources.extend(other.resources)


# ------------------------------------------------------------------------------
# Base importer interface
# ------------------------------------------------------------------------------


class BaseGESNImporter:
    """Base importer returning a bundle of records."""

    def parse(self, path: Path) -> GESNImportBundle:
        raise NotImplementedError

    @staticmethod
    def _normalize_code(code: str) -> str:
        return re.sub(r"\s+", "", code.upper())


# ------------------------------------------------------------------------------
# CSV / Excel importer
# ------------------------------------------------------------------------------


class SpreadsheetGESNImporter(BaseGESNImporter):
    """Importer for CSV or Excel sources based on official Rosstandart releases."""

    def parse(self, path: Path) -> GESNImportBundle:
        if path.suffix.lower() in {".csv"}:
            rows = self._read_csv(path)
        else:
            rows = self._read_excel(path)

        bundle = GESNImportBundle()
        for row in rows:
            try:
                section_code = self._normalize_code(row.get("section_code") or row.get("section", ""))
                if section_code:
                    bundle.sections.append(
                        SectionRecord(
                            code=section_code,
                            name=row.get("section_name") or row.get("section"),
                            parent_code=self._normalize_code(row.get("parent_section") or "") or None,
                        )
                    )

                norm_code = self._normalize_code(row["norm_code"])
                bundle.norms.append(
                    NormRecord(
                        code=norm_code,
                        name=row["norm_name"],
                        unit=row.get("unit") or row.get("unit_code") or "ед.",
                        section_code=section_code or self._normalize_code(row.get("section_code") or ""),
                        unit_description=row.get("unit_description"),
                        work_type=row.get("work_type"),
                        description=row.get("description"),
                        composition=row.get("composition"),
                        parameters=self._extract_parameters(row.get("parameters")),
                        metadata=self._extract_metadata(row),
                        components=self._split_components(row.get("components")),
                    )
                )

                resources_raw = row.get("resources")
                for res in self._parse_resources(norm_code, resources_raw):
                    bundle.resources.append(res)

            except KeyError as exc:
                logger.error("Missing required column in %s: %s", path, exc)
        return bundle

    def _read_csv(self, path: Path) -> Iterable[Dict[str, str]]:
        with path.open("r", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                yield {k.strip(): v.strip() if isinstance(v, str) else v for k, v in row.items()}

    def _read_excel(self, path: Path) -> Iterable[Dict[str, str]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        for row in sheet.iter_rows(min_row=2):
            values = {}
            for header, cell in zip(headers, row):
                values[header] = str(cell.value).strip() if cell.value is not None else ""
            yield values

    def _extract_parameters(self, raw: Optional[str]) -> Dict[str, object]:
        if not raw:
            return {}
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return {"raw": raw}

    def _extract_metadata(self, row: Dict[str, str]) -> Dict[str, object]:
        metadata = {}
        for key in ("source", "chapter", "notes"):
            if row.get(key):
                metadata[key] = row[key]
        return metadata

    def _split_components(self, raw: Optional[str]) -> List[str]:
        if not raw:
            return []
        parts = [part.strip() for part in re.split(r"[;\n]", raw) if part.strip()]
        return parts

    def _parse_resources(self, norm_code: str, raw: Optional[str]) -> List[ResourceRecord]:
        if not raw:
            return []

        resources: List[ResourceRecord] = []
        try:
            data = json.loads(raw)
            for item in data:
                resources.append(
                    ResourceRecord(
                        norm_code=norm_code,
                        resource_type=item.get("type", "material"),
                        name=item["name"],
                        unit=item.get("unit", "ед."),
                        quantity=float(item["quantity"]) if item.get("quantity") else None,
                        code=item.get("code"),
                        metadata={k: v for k, v in item.items() if k not in {"type", "name", "unit", "quantity", "code"}},
                    )
                )
        except json.JSONDecodeError:
            # Fallback: semicolon separated entries
            for item in raw.split(";"):
                item = item.strip()
                if not item:
                    continue
                match = re.match(r"(?P<name>.+?)(?:\s+\((?P<unit>[^)]+)\))?(?:\s+-\s*(?P<quantity>[\d.,]+))?", item)
                if not match:
                    continue
                quantity = match.group("quantity")
                resources.append(
                    ResourceRecord(
                        norm_code=norm_code,
                        resource_type="material",
                        name=match.group("name").strip(),
                        unit=(match.group("unit") or "ед.").strip(),
                        quantity=float(quantity.replace(",", ".")) if quantity else None,
                    )
                )
        return resources


# ------------------------------------------------------------------------------
# XML importer
# ------------------------------------------------------------------------------


class XMLGESNImporter(BaseGESNImporter):
    """Importer for official XML releases (Bldr2 legacy compatible)."""

    def parse(self, path: Path) -> GESNImportBundle:
        tree = etree.parse(str(path))
        root = tree.getroot()

        namespace = root.nsmap.get(None, "")
        ns = {"ns": namespace} if namespace else {}

        bundle = GESNImportBundle()
        for section in root.findall(".//ns:Section", namespaces=ns) if namespace else root.findall(".//Section"):
            code = self._normalize_code(section.get("Code") or "")
            parent_code = self._normalize_code(section.get("ParentCode") or "") or None
            name = section.get("Name") or (section.findtext("ns:Name", namespaces=ns) if namespace else section.findtext("Name"))
            if not code or not name:
                continue
            bundle.sections.append(
                SectionRecord(
                    code=code,
                    name=name,
                    parent_code=parent_code,
                    description=section.findtext("ns:Description", namespaces=ns) if namespace else section.findtext("Description"),
                )
            )

        for norm in root.findall(".//ns:Norm", namespaces=ns) if namespace else root.findall(".//Norm"):
            code = self._normalize_code(norm.get("Code") or "")
            if not code:
                continue
            section_code = self._normalize_code(norm.get("SectionCode") or norm.findtext("ns:SectionCode", namespaces=ns) or "")
            components = [
                comp.text.strip()
                for comp in norm.findall("ns:Component", namespaces=ns)
                if comp.text and comp.text.strip()
            ] if namespace else [
                comp.text.strip() for comp in norm.findall("Component") if comp.text and comp.text.strip()
            ]
            parameters = {}
            for param in norm.findall("ns:Parameter", namespaces=ns) if namespace else norm.findall("Parameter"):
                name = param.get("Name")
                value = param.get("Value")
                if name:
                    parameters[name] = value

            resources = norm.findall("ns:Resource", namespaces=ns) if namespace else norm.findall("Resource")

            bundle.norms.append(
                NormRecord(
                    code=code,
                    name=norm.get("Name") or norm.findtext("ns:Name", namespaces=ns) or "",
                    unit=norm.get("Unit") or norm.findtext("ns:Unit", namespaces=ns) or "ед.",
                    section_code=section_code,
                    unit_description=norm.get("UnitDescription"),
                    work_type=norm.get("WorkType"),
                    description=norm.findtext("ns:Description", namespaces=ns) if namespace else norm.findtext("Description"),
                    composition=norm.findtext("ns:Composition", namespaces=ns) if namespace else norm.findtext("Composition"),
                    parameters=parameters,
                    metadata={"source": "xml"},
                    components=components,
                )
            )

            for resource in resources:
                bundle.resources.append(
                    ResourceRecord(
                        norm_code=code,
                        resource_type=resource.get("Type", "material"),
                        name=resource.get("Name") or "",
                        unit=resource.get("Unit") or "ед.",
                        quantity=float(resource.get("Quantity").replace(",", ".")) if resource.get("Quantity") else None,
                        code=resource.get("Code"),
                        metadata={"source": "xml"},
                    )
                )
        return bundle


# ------------------------------------------------------------------------------
# PDF importer (heuristic)
# ------------------------------------------------------------------------------


class PDFGESNImporter(BaseGESNImporter):
    """
    Импортёр PDF-книжек ГЭСН с использованием DeepSeek OCR и LLM-интерпретации.

    Алгоритм:
      1. Каждая страница конвертируется в изображение и передаётся в DeepSeek.
      2. Модель получает инструкцию вернуть JSON-структуру с нормами, ресурсами и составом работ.
      3. Ответ проверяется и преобразуется в NormRecord/ResourceRecord.
      4. При ошибках парсинга применяется гарантированный фолбэк — эвристический анализ текста.
    """

    JSON_PROMPT = (
        "Ты инженер-сметчик. На изображении находится страница официального сборника ГЭСН. "
        "Проанализируй текст и верни строго валидный JSON без комментариев и пояснений. "
        "Структура: {\"norms\": [{\"code\": \"16-02-005-01\", \"name\": \"Прокладка труб...\", "
        "\"unit\": \"100 м\", \"section_code\": \"16\", "
        "\"components\": [\"Разметка трассы\", \"Сварка\"], "
        "\"resources\": [{\"type\": \"labor\", \"name\": \"Монтажники\", \"unit\": \"чел·ч\", \"quantity\": 12.5}, "
        "{\"type\": \"material\", \"name\": \"Труба стальная\", \"unit\": \"м\", \"quantity\": 100}]}]}. "
        "Заполни поле section_code исходя из первых двух цифр кода нормы. "
        "Если на странице нет норм — верни {\"norms\": []}."
    )

    HEADER_PATTERN = re.compile(r"^(?P<code>\d{2}\-\d{2}\-\d{3}(?:\-[\dA-Z]+)?)\s+(?P<name>.+)$")
    UNIT_PATTERN = re.compile(r"Ед\.\s*изм\.\s*:\s*(?P<unit>[^\s]+)", re.IGNORECASE)

    def parse(self, path: Path) -> GESNImportBundle:
        bundle = GESNImportBundle()
        with fitz.open(path) as doc:
            page_groups = self._group_pages_by_layout(doc)
            for group in page_groups:
                try:
                    ocr_bundle = self._parse_group_with_ocr(doc, group)
                    bundle.extend(ocr_bundle)
                except Exception as exc:
                    logger.warning(
                        "OCR-парсинг страницы(-ц) %s завершился ошибкой: %s",
                        ", ".join(str(idx + 1) for idx in group),
                        exc,
                    )
                    for page_index in group:
                        fallback_bundle = self._parse_page_fallback(doc[page_index])
                        bundle.extend(fallback_bundle)
    def _group_pages_by_layout(self, doc: fitz.Document, *, dpi: int = 300) -> List[List[int]]:
        """
        Определяет группы страниц, где таблица "перетекает" на следующую страницу.

        Стратегия:
            - Превращаем страницу в текст.
            - Ищем признаки незавершённой таблицы в конце (строки без итоговой строки и без пустой строки).
            - Если таблица продолжается, объединяем текущую страницу с следующей.
        """
        groups: List[List[int]] = []
        current_group: List[int] = []

        for index in range(len(doc)):
            page = doc[index]
            text = page.get_text("text").strip()
            if not current_group:
                current_group.append(index)
            else:
                # проверяем, закончилась ли таблица на предыдущей странице
                prev_page = doc[current_group[-1]]
                prev_text = prev_page.get_text("text")
                if self._table_continues(prev_text):
                    current_group.append(index)
                else:
                    groups.append(current_group)
                    current_group = [index]

            # если это последняя страница, добавляем текущую группу
            if index == len(doc) - 1:
                groups.append(current_group)

        return groups

    def _table_continues(self, text: str) -> bool:
        """
        Эвристика: если последняя строка страницы заканчивается числом или единицей измерения, таблица идёт дальше.
        """
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if not lines:
            return False
        last_line = lines[-1]
        return bool(re.search(r"(?:\d|\bшт\b|\bм\b|\bм²\b|\bм³\b)$", last_line))

    def _parse_group_with_ocr(self, doc: fitz.Document, group: List[int]) -> GESNImportBundle:
        """
        Выполняет OCR для одной или нескольких страниц, объединённых в таблицу.

        Returns:
            GESNImportBundle: результат разбора.
        """
        images: List[bytes] = []
        for page_index in group:
            page = doc[page_index]
            pixmap = page.get_pixmap(dpi=300)
            images.append(pixmap.tobytes("png"))

        instruction = self.JSON_PROMPT + " Учитывай, что таблица может продолжаться на нескольких страницах."
        raw_response = deepseek_ocr_service.extract_structured(
            b"".join(images),  # объединённое изображение
            instruction=instruction,
            max_new_tokens=4096,
        )

        data = self._safe_json_load(raw_response)
        bundle = GESNImportBundle()

        for entry in data.get("norms", []):
            code = self._normalize_code(entry.get("code", ""))
            if not code:
                continue

            section_code = self._normalize_code(entry.get("section_code") or code[:2])
            bundle.sections.append(
                SectionRecord(
                    code=section_code,
                    name=f"Раздел {section_code}",
                    parent_code=None,
                    metadata={"source": "pdf_llm", "pages": [idx + 1 for idx in group]},
                )
            )

            norm = NormRecord(
                code=code,
                name=entry.get("name", "").strip(),
                unit=(entry.get("unit") or "ед.").strip(),
                section_code=section_code,
                description=entry.get("description"),
                composition=None,
                parameters=self._collect_parameters(entry),
                metadata={"source": "pdf_llm", "pages": [idx + 1 for idx in group]},
                components=entry.get("components") or [],
            )
            bundle.norms.append(norm)

            for resource in entry.get("resources") or []:
                try:
                    resource_record = ResourceRecord(
                        norm_code=code,
                        resource_type=(resource.get("type") or "material").strip(),
                        name=resource.get("name", "").strip(),
                        unit=(resource.get("unit") or "ед.").strip(),
                        quantity=float(resource.get("quantity")) if resource.get("quantity") is not None else None,
                        code=(resource.get("code") or "").strip() or None,
                    )
                    bundle.resources.append(resource_record)
                except (TypeError, ValueError):
                    logger.debug("Пропуск ресурса из-за неверного формата: %s", resource)

        return bundle

        self._append_inferred_sections(bundle)
        return bundle

    def _parse_page_with_ocr(self, page: fitz.Page) -> GESNImportBundle:
        """
        Распознаёт страницу PDF с помощью DeepSeek и возвращает институт данных.
        """
        pixmap = page.get_pixmap(dpi=300)
        image_bytes = pixmap.tobytes("png")
        raw_response = deepseek_ocr_service.extract_structured(
            image_bytes,
            instruction=self.JSON_PROMPT,
            max_new_tokens=2048,
        )
        data = self._safe_json_load(raw_response)
        bundle = GESNImportBundle()

        for entry in data.get("norms", []):
            code = self._normalize_code(entry.get("code", ""))
            if not code:
                continue
            section_code = self._normalize_code(entry.get("section_code") or code[:2])

            bundle.sections.append(
                SectionRecord(
                    code=section_code,
                    name=f"Раздел {section_code}",
                    parent_code=None,
                    metadata={"source": "pdf_llm"},
                )
            )

            norm = NormRecord(
                code=code,
                name=entry.get("name", "").strip(),
                unit=(entry.get("unit") or "ед.").strip(),
                section_code=section_code,
                description=entry.get("description"),
                composition=None,
                parameters=self._collect_parameters(entry),
                metadata={"source": "pdf_llm", "page": page.number + 1},
                components=entry.get("components") or [],
            )
            bundle.norms.append(norm)

            for resource in entry.get("resources") or []:
                try:
                    resource_record = ResourceRecord(
                        norm_code=code,
                        resource_type=(resource.get("type") or "material").strip(),
                        name=resource.get("name", "").strip(),
                        unit=(resource.get("unit") or "ед.").strip(),
                        quantity=float(resource.get("quantity")) if resource.get("quantity") is not None else None,
                        code=(resource.get("code") or "").strip() or None,
                        metadata={"source": "pdf_llm"},
                    )
                    bundle.resources.append(resource_record)
                except (TypeError, ValueError):
                    logger.debug("Пропуск ресурса из-за неверного формата: %s", resource)
        return bundle

    def _parse_page_fallback(self, page: fitz.Page) -> GESNImportBundle:
        """
        Резервная стратегия: эвристический анализ текста без LLM, чтобы не терять данные.
        """
        bundle = GESNImportBundle()
        text = page.get_text("text")
        current_norm: Optional[NormRecord] = None
        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue

            header_match = self.HEADER_PATTERN.match(line)
            if header_match:
                if current_norm:
                    bundle.norms.append(current_norm)
                code = self._normalize_code(header_match.group("code"))
                name = header_match.group("name").strip()
                section_code = code[:2] if len(code) >= 2 else ""
                current_norm = NormRecord(
                    code=code,
                    name=name,
                    unit="ед.",
                    section_code=section_code,
                    metadata={"source": "pdf_fallback", "page": page.number + 1},
                    components=[],
                )
                bundle.sections.append(
                    SectionRecord(
                        code=section_code,
                        name=f"Раздел {section_code}",
                        metadata={"source": "pdf_fallback"},
                    )
                )
                continue

            if current_norm:
                unit_match = self.UNIT_PATTERN.search(line)
                if unit_match:
                    current_norm.unit = unit_match.group("unit")
                    continue
                current_norm.components.append(line)

        if current_norm:
            bundle.norms.append(current_norm)
        return bundle

    def _safe_json_load(self, payload: str) -> Dict[str, object]:
        """
        Аккуратно извлекает JSON-объект из произвольного ответа модели.
        """
        start = payload.find("{")
        end = payload.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise ValueError("Ответ LLM не содержит корректного JSON.")
        json_fragment = payload[start : end + 1]
        return json.loads(json_fragment)

    def _collect_parameters(self, entry: Dict[str, object]) -> Dict[str, object]:
        """
        Собирает дополнительные параметры нормы, если они указаны моделью.
        """
        params = entry.get("parameters")
        if isinstance(params, dict):
            return params
        return {}

    def _append_inferred_sections(self, bundle: GESNImportBundle) -> None:
        """
        Добавляет недостающие разделы, опираясь на коды норм.
        """
        existing = {section.code for section in bundle.sections if section.code}
        for norm in bundle.norms:
            if norm.section_code and norm.section_code not in existing:
                bundle.sections.append(
                    SectionRecord(
                        code=norm.section_code,
                        name=f"Раздел {norm.section_code}",
                        metadata={"source": "pdf_inferred"},
                    )
                )
                existing.add(norm.section_code)


# ------------------------------------------------------------------------------
# Catalog loader
# ------------------------------------------------------------------------------


class GESNCatalogLoader:
    """
    Loads bundles into the relational schema, handling upserts.
    """

    def __init__(self, session: Session):
        self.session = session
        self._section_cache: Dict[str, GESNSection] = {}
        self._norm_cache: Dict[str, GESNNorm] = {}

    def load(self, bundle: GESNImportBundle, *, commit: bool = True) -> None:
        self._ensure_sections(bundle.sections)
        self._ensure_norms(bundle.norms)
        self._ensure_resources(bundle.resources)
        if commit:
            self.session.commit()

    # -- helpers ----------------------------------------------------------------

    def _ensure_sections(self, records: Sequence[SectionRecord]) -> None:
        for record in records:
            code = record.code
            if not code:
                continue
            section = self._section_cache.get(code) or self.session.query(GESNSection).filter_by(code=code).one_or_none()
            if section:
                updated = False
                if record.name and section.name != record.name:
                    section.name = record.name
                    updated = True
                if record.description and section.description != record.description:
                    section.description = record.description
                    updated = True
                if record.metadata:
                    section.metadata_json = {**(section.metadata_json or {}), **record.metadata}
                    updated = True
                if updated:
                    self.session.add(section)
            else:
                parent = None
                if record.parent_code:
                    parent = self._section_cache.get(record.parent_code) or self.session.query(GESNSection).filter_by(
                        code=record.parent_code
                    ).one_or_none()
                section = GESNSection(
                    code=code,
                    name=record.name,
                    parent=parent,
                    description=record.description,
                    metadata_json=record.metadata or None,
                )
                self.session.add(section)
            self._section_cache[code] = section

    def _ensure_norms(self, records: Sequence[NormRecord]) -> None:
        for record in records:
            if not record.code:
                continue
            section = self._section_cache.get(record.section_code) or self.session.query(GESNSection).filter_by(
                code=record.section_code
            ).one_or_none()
            if not section:
                logger.warning("Section %s not found for norm %s", record.section_code, record.code)
                continue

            norm = self._norm_cache.get(record.code) or self.session.query(GESNNorm).filter_by(code=record.code).one_or_none()
            if norm:
                norm.name = record.name or norm.name
                norm.unit = record.unit or norm.unit
                norm.unit_description = record.unit_description or norm.unit_description
                norm.work_type = record.work_type or norm.work_type
                norm.description = record.description or norm.description
                norm.composition = record.composition or norm.composition
                if record.parameters:
                    norm.parameters = {**(norm.parameters or {}), **record.parameters}
                if record.metadata:
                    norm.extra_metadata = {**(norm.extra_metadata or {}), **record.metadata}
                if record.components:
                    norm.components = []
                    for idx, component in enumerate(record.components):
                        norm.components.append(GESNNormComponent(order_index=idx, description=component))
                self.session.add(norm)
            else:
                norm = GESNNorm(
                    code=record.code,
                    name=record.name,
                    unit=record.unit,
                    unit_description=record.unit_description,
                    work_type=record.work_type,
                    description=record.description,
                    composition=record.composition,
                    parameters=record.parameters or None,
                    extra_metadata=record.metadata or None,
                    section=section,
                )
                for idx, component in enumerate(record.components):
                    norm.components.append(GESNNormComponent(order_index=idx, description=component))
                self.session.add(norm)
            self._norm_cache[record.code] = norm

    def _ensure_resources(self, records: Sequence[ResourceRecord]) -> None:
        for record in records:
            norm = self._norm_cache.get(record.norm_code) or self.session.query(GESNNorm).filter_by(
                code=record.norm_code
            ).one_or_none()
            if not norm:
                logger.warning("Norm %s not found for resource %s", record.norm_code, record.name)
                continue

            norm.resources.append(
                GESNNormResource(
                    resource_type=record.resource_type,
                    name=record.name,
                    unit=record.unit,
                    quantity=record.quantity,
                    code=record.code,
                    metadata_json=record.metadata or None,
                )
            )


# ------------------------------------------------------------------------------
# Helper function
# ------------------------------------------------------------------------------


def detect_importer(path: Path) -> BaseGESNImporter:
    ext = path.suffix.lower()
    if ext in {".csv", ".xlsx", ".xlsm"}:
        return SpreadsheetGESNImporter()
    if ext in {".xml"}:
        return XMLGESNImporter()
    if ext in {".pdf"}:
        return PDFGESNImporter()
    raise ValueError(f"Unsupported GESN source format: {ext}")


