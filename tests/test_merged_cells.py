#!/usr/bin/env python3
"""
Тест обработки объединенных ячеек в enhanced_smeta_parser.py
"""

import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment
from core.enhanced_smeta_parser import EnhancedSmetaParser

def create_test_file_with_merged_cells():
    """Создаем тестовый Excel файл с объединенными ячейками"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Тест объединенных ячеек"
    
    # Заголовки
    ws['A1'] = 'Код'
    ws['B1'] = 'Наименование работ'
    ws['C1'] = 'Ед. изм.'
    ws['D1'] = 'Количество'
    ws['E1'] = 'Цена'
    ws['F1'] = 'Сумма'
    
    # Данные с объединенными ячейками
    # Строка 1: Устройство фундамента (объединено на 3 строки)
    ws['A2'] = 'ГЭСН 8-1-1'
    ws.merge_cells('B2:B4')  # Объединяем "Наименование работ" на 3 строки
    ws['B2'] = 'Устройство бетонного фундамента'
    ws['C2'] = 'м3'
    ws['D2'] = 100
    ws['E2'] = 15000
    ws['F2'] = 1500000
    
    # Строка 2: Детали (без объединения)
    ws['A3'] = 'ГЭСН 8-1-1.1'
    ws['C3'] = 'м3'
    ws['D3'] = 50
    ws['E3'] = 12000
    ws['F3'] = 600000
    
    # Строка 3: Детали (без объединения)
    ws['A4'] = 'ГЭСН 8-1-1.2'
    ws['C4'] = 'м3'
    ws['D4'] = 30
    ws['E4'] = 18000
    ws['F4'] = 540000
    
    # Строка 4: Устройство стен (объединено на 2 строки)
    ws['A5'] = 'ГЭСН 8-2-1'
    ws.merge_cells('B5:B6')  # Объединяем "Наименование работ" на 2 строки
    ws['B5'] = 'Устройство кирпичных стен'
    ws['C5'] = 'м2'
    ws['D5'] = 200
    ws['E5'] = 8000
    ws['F5'] = 1600000
    
    # Строка 5: Детали (без объединения)
    ws['A6'] = 'ГЭСН 8-2-1.1'
    ws['C6'] = 'м2'
    ws['D6'] = 100
    ws['E6'] = 7500
    ws['F6'] = 750000
    
    # Итого
    ws['A7'] = 'ИТОГО'
    ws['B7'] = 'Общая стоимость'
    ws['F7'] = 4990000
    
    # Применяем стили
    for row in ws.iter_rows(min_row=1, max_row=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Сохраняем файл
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file_path = temp_file.name
    temp_file.close()
    
    wb.save(temp_file_path)
    print(f"Создан тестовый файл: {temp_file_path}")
    return temp_file_path

def test_merged_cells_parsing():
    """Тестируем парсинг файла с объединенными ячейками"""
    
    print("ТЕСТ: Обработка объединенных ячеек")
    print("=" * 50)
    
    # Создаем тестовый файл
    test_file = create_test_file_with_merged_cells()
    
    try:
        # Парсим файл
        parser = EnhancedSmetaParser()
        result = parser.parse_excel_estimate(test_file)
        
        print(f"Статус парсинга: {result.get('format', 'unknown')}")
        print(f"Найдено позиций: {result.get('positions_count', 0)}")
        print(f"Общая стоимость: {result.get('total_cost', 0):,.2f} руб")
        
        print("\nДетали позиций:")
        for i, pos in enumerate(result.get('positions', []), 1):
            print(f"  {i}. {pos.get('code', 'N/A')} - {pos.get('description', 'N/A')}")
            print(f"     Количество: {pos.get('quantity', 0)} {pos.get('unit', '').replace('м³', 'м3').replace('м²', 'м2')}")
            print(f"     Цена: {pos.get('price', 0):,.2f} руб")
            print(f"     Сумма: {pos.get('total_cost', 0):,.2f} руб")
            print()
        
        # Проверяем метаданные
        metadata = result.get('metadata', {})
        if metadata.get('merged_cells_processed'):
            print("Объединенные ячейки были обработаны!")
        else:
            print("Объединенные ячейки НЕ были обработаны")
        
        return result
        
    except Exception as e:
        print(f"Ошибка тестирования: {e}")
        return None
    
    finally:
        # Удаляем тестовый файл
        try:
            os.unlink(test_file)
            print(f"Удален тестовый файл: {test_file}")
        except:
            pass

if __name__ == "__main__":
    test_merged_cells_parsing()
